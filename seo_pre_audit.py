#!/usr/bin/env python3
"""
SEO Audit Script — v1.0
Analizza tre aree chiave:
  1. Rendering e visibilità dei contenuti (contenuti nascosti, lazy load, JS-rendered)
  2. Struttura heading H2/H3 e coerenza con snippet
  3. robots.txt — direttive non supportate e typo

Uso:
    python seo_audit.py --url https://esempio.com
    python seo_audit.py --url https://esempio.com --output report.xlsx
    python seo_audit.py --sitemap https://esempio.com/sitemap.xml --max 20

Dipendenze: requests, beautifulsoup4, lxml, playwright, openpyxl
"""

import argparse
import sys
import os
import re
import json
import time
from urllib.parse import urlparse, urljoin
from datetime import datetime

# ── Playwright: installa browser se mancante ─────────────────────────────────
# Su Streamlit Cloud il browser va scaricato a runtime con os.system.
# È il metodo confermato dalla community ufficiale Streamlit.
import glob as _glob

import requests
from bs4 import BeautifulSoup

# Playwright opzionale — funziona senza per il fetch statico
import warnings as _warnings
_warnings.filterwarnings("ignore")
try:
    import logging as _logging
    _logging.getLogger("playwright").setLevel(_logging.ERROR)
    from playwright.sync_api import sync_playwright as _sync_playwright
    PLAYWRIGHT_AVAILABLE = True
except ImportError:
    PLAYWRIGHT_AVAILABLE = False
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter


# ─── CONFIGURAZIONE ──────────────────────────────────────────────────────────

# Se usi la Search Console API per recuperare snippet reali, inserisci qui.
# Lascia vuoto per saltare il confronto snippet (funziona comunque tutto il resto).
GOOGLE_API_KEY = ""
SEARCH_CONSOLE_SITE = ""   # es. "https://esempio.com/"

REQUEST_TIMEOUT = 15
USER_AGENT = (
    "Mozilla/5.0 (compatible; SEOAuditBot/1.0; "
    "+https://github.com/emiliano/seo-audit)"
)

# Direttive robots.txt NON supportate da Google (fonte: Google docs + SEJ)
UNSUPPORTED_DIRECTIVES = {
    "crawl-delay", "noindex", "nofollow", "noarchive", "nosnippet",
    "noodp", "noydir", "nocache", "nopreview", "notranslate",
    "noimageindex", "unavailable_after", "visit-time", "request-rate",
    "cache-delay", "acap-crawler", "acap-labels",
    "host",         # usato da Yandex, ignorato da Google
    "clean-param",  # usato da Yandex
    "pattern",
    "length",
    "index",        # non standard
    "follow",       # non standard
}

# Typo comuni di "Disallow" riscontrati nei robots.txt reali
DISALLOW_TYPOS = [
    "disalow", "dissallow", "dissalow", "disallov", "disalloow",
    "disallow ", "Disallow",  "DISALLOW",
    "Disalow", "diasllow", "disllow", "diallow",
]

# Selettori CSS che indicano contenuto nascosto/espandibile
HIDDEN_CONTENT_SELECTORS = [
    # Accordion / FAQ
    "[aria-expanded]", "[aria-hidden='true']",
    ".accordion", ".accordion-item", ".accordion-body",
    ".faq", ".faq-item", ".faq-answer",
    # Bootstrap collapse
    ".collapse:not(.show)", ".collapsible",
    # Tab panels
    '[role="tabpanel"]', ".tab-content", ".tab-pane:not(.active)",
    # Toggle / show-more
    ".hidden", ".is-hidden", ".d-none", "[hidden]",
    ".read-more-content", ".show-more",
    # Custom espandibili comuni
    ".expandable", ".expand-content", "[data-toggle='collapse']",
    ".js-hidden", ".initially-hidden",
]

# Selettori per lazy-load / scroll-triggered
LAZY_LOAD_SELECTORS = [
    "[data-src]", "[data-lazy]", "[data-lazyload]",
    "[loading='lazy']", ".lazyload", ".lazy",
    "[data-scroll]", "[data-aos]",  # AOS animate on scroll
    "[data-sal]",                    # Scroll Animation Library
    "[data-wow]",                    # WOW.js
    "[data-inview]",
]

# Pattern JS che suggeriscono contenuto iniettato dopo interazione
JS_INJECT_PATTERNS = [
    r"IntersectionObserver",
    r"addEventListener\(['\"]scroll",
    r"addEventListener\(['\"]click",
    r"\.toggle\(",
    r"\.slideDown\(",
    r"\.slideToggle\(",
    r"classList\.toggle",
    r"classList\.remove\(['\"]hidden",
    r"classList\.add\(['\"]visible",
    r"style\.display\s*=\s*['\"]block",
    r"fetch\(|axios\.|XMLHttpRequest",
    r"loadMore|load_more|lazy.?load",
]


# ─── HELPERS ─────────────────────────────────────────────────────────────────

def get_session():
    s = requests.Session()
    s.headers.update({"User-Agent": USER_AGENT})
    return s


def fetch_html_static(url, session):
    """Fetch statico (senza JS)."""
    try:
        r = session.get(url, timeout=REQUEST_TIMEOUT, allow_redirects=True)
        r.raise_for_status()
        return r.text, r.status_code, r.url
    except Exception as e:
        return None, None, str(e)


def fetch_html_rendered(url):
    """Fetch con Playwright (JS eseguito). Ritorna None se non disponibile."""
    if not PLAYWRIGHT_AVAILABLE:
        return None
    try:
        import os as _os
        _os.environ["PYTHONWARNINGS"] = "ignore"
        from playwright.sync_api import sync_playwright
        with sync_playwright() as p:
            browser = p.chromium.launch(
                headless=True,
                args=[
                    "--no-sandbox",
                    "--disable-setuid-sandbox",
                    "--disable-dev-shm-usage",
                    "--disable-gpu",
                    "--single-process",
                    "--disable-extensions",
                    "--log-level=3",
                ],
            )
            page = browser.new_page(user_agent=USER_AGENT)
            page.goto(url, wait_until="domcontentloaded", timeout=30000)
            page.wait_for_timeout(1500)
            content = page.content()
            browser.close()
            return content
    except Exception as e:
        return None


# ─── AREA 1: RENDERING E VISIBILITÀ CONTENUTI ────────────────────────────────

def audit_content_visibility(url, static_html, rendered_html):
    """
    Confronta DOM statico vs renderizzato.
    Identifica contenuti nascosti, tab, accordion, lazy-load.
    """
    issues = []
    details = []

    static_soup = BeautifulSoup(static_html, "lxml") if static_html else None
    rendered_soup = BeautifulSoup(rendered_html, "lxml") if rendered_html else None

    # 1a. Contenuti nascosti/espandibili nel DOM statico
    hidden_found = []
    if static_soup:
        for selector in HIDDEN_CONTENT_SELECTORS:
            try:
                elements = static_soup.select(selector)
                for el in elements:
                    text = el.get_text(strip=True)
                    if len(text) > 30:  # solo se c'è testo significativo
                        hidden_found.append({
                            "selector": selector,
                            "tag": el.name,
                            "text_preview": text[:100] + ("…" if len(text) > 100 else ""),
                            "has_heading": bool(el.find(["h2", "h3", "h4"])),
                        })
            except Exception:
                pass

    if hidden_found:
        issues.append("WARN")
        details.append(
            f"Trovati {len(hidden_found)} blocchi con contenuto nascosto/espandibile "
            f"(accordion, tab, collapse). Questi riducono la probabilità di deep link 'Read more'."
        )
    else:
        issues.append("OK")
        details.append("Nessun contenuto nascosto significativo rilevato nel DOM statico.")

    # 1b. Elementi lazy-load
    lazy_found = []
    if static_soup:
        for selector in LAZY_LOAD_SELECTORS:
            try:
                elements = static_soup.select(selector)
                if elements:
                    lazy_found.extend([(selector, el.name) for el in elements])
            except Exception:
                pass

    if lazy_found:
        issues.append("WARN")
        details.append(
            f"Trovati {len(lazy_found)} elementi con attributi lazy-load/scroll-trigger "
            f"({', '.join(set(s for s,_ in lazy_found[:5]))}). "
            f"Il contenuto caricato dopo scroll riduce la probabilità di deep link."
        )
    else:
        issues.append("OK")
        details.append("Nessun attributo lazy-load problematico rilevato.")

    # 1c. Pattern JS sospetti negli script inline
    js_patterns_found = []
    if static_soup:
        scripts = static_soup.find_all("script")
        inline_js = " ".join(s.string or "" for s in scripts if s.string)
        for pattern in JS_INJECT_PATTERNS:
            if re.search(pattern, inline_js):
                js_patterns_found.append(pattern)

    if js_patterns_found:
        issues.append("WARN")
        details.append(
            f"Trovati {len(js_patterns_found)} pattern JS che suggeriscono contenuto "
            f"iniettato dopo interazione utente: {', '.join(js_patterns_found[:4])}."
        )
    else:
        issues.append("OK")
        details.append("Nessun pattern JS di iniezione contenuto post-interazione rilevato.")

    # 1d. Delta testo statico vs renderizzato
    delta_info = "N/A (rendering JS non disponibile)"
    if static_soup and rendered_soup:
        static_text = static_soup.get_text(separator=" ", strip=True)
        rendered_text = rendered_soup.get_text(separator=" ", strip=True)
        static_words = len(static_text.split())
        rendered_words = len(rendered_text.split())
        delta_pct = ((rendered_words - static_words) / max(static_words, 1)) * 100
        delta_info = (
            f"Testo statico: {static_words} parole | "
            f"Testo renderizzato: {rendered_words} parole | "
            f"Delta: {delta_pct:+.1f}%"
        )
        if delta_pct > 15:
            issues.append("WARN")
            details.append(
                f"Il rendering JS aggiunge il {delta_pct:.1f}% di testo in più rispetto "
                f"al DOM statico. Googlebot potrebbe non vedere tutto questo contenuto "
                f"al primo crawl. ({delta_info})"
            )
        else:
            issues.append("OK")
            details.append(f"Delta testo statico/renderizzato contenuto ({delta_info}).")
    else:
        details.append(delta_info)

    # Riepilogo sezione
    hidden_detail = []
    for h in hidden_found[:10]:
        hidden_detail.append(
            f"  [{h['selector']}] <{h['tag']}> — "
            f"{'⚠ contiene heading' if h['has_heading'] else 'no heading'} — "
            f"«{h['text_preview']}»"
        )

    overall = "FAIL" if issues.count("WARN") >= 2 else ("WARN" if "WARN" in issues else "OK")

    return {
        "area": "1. Rendering & Visibilità Contenuti",
        "overall": overall,
        "issues": issues,
        "summary": " | ".join(details),
        "hidden_blocks": hidden_found,
        "hidden_detail_text": "\n".join(hidden_detail) if hidden_detail else "—",
        "lazy_elements": len(lazy_found),
        "js_patterns": js_patterns_found,
        "dom_delta": delta_info,
    }


# ─── AREA 2: STRUTTURA HEADING ────────────────────────────────────────────────

def audit_heading_structure(url, static_html):
    """
    Verifica:
    - Presenza e unicità di H1
    - Uso corretto di H2/H3 per le sezioni principali
    - Sezioni di testo lungo prive di heading
    - Salti di livello (es. H1→H3 senza H2)
    """
    issues = []
    details = []
    heading_map = []

    if not static_html:
        return {"area": "2. Struttura Heading", "overall": "ERROR", "summary": "HTML non disponibile"}

    soup = BeautifulSoup(static_html, "lxml")

    # Rimuovi header/footer/nav per l'analisi del contenuto principale
    for tag in soup.find_all(["nav", "header", "footer", "aside"]):
        tag.decompose()

    headings = soup.find_all(["h1", "h2", "h3", "h4", "h5", "h6"])

    # 2a. H1
    h1_list = [h for h in headings if h.name == "h1"]
    if len(h1_list) == 0:
        issues.append("FAIL")
        details.append("Nessun H1 trovato nella pagina.")
    elif len(h1_list) > 1:
        issues.append("WARN")
        details.append(f"Trovati {len(h1_list)} H1 — dovrebbe essere uno solo.")
    else:
        issues.append("OK")
        details.append(f"H1 presente e unico: «{h1_list[0].get_text(strip=True)[:80]}»")

    # 2b. Mappa heading e salti di livello
    level_sequence = []
    for h in headings:
        level = int(h.name[1])
        text = h.get_text(strip=True)
        level_sequence.append(level)
        heading_map.append({
            "level": level,
            "tag": h.name.upper(),
            "text": text[:100] + ("…" if len(text) > 100 else ""),
        })

    jumps = []
    for i in range(1, len(level_sequence)):
        if level_sequence[i] - level_sequence[i - 1] > 1:
            jumps.append(
                f"H{level_sequence[i-1]}→H{level_sequence[i]} "
                f"(«{heading_map[i]['text'][:50]}»)"
            )

    if jumps:
        issues.append("WARN")
        details.append(
            f"Rilevati {len(jumps)} salti di livello heading: {'; '.join(jumps[:3])}."
        )
    else:
        issues.append("OK")
        details.append("Nessun salto di livello heading rilevato.")

    # 2c. H2/H3 nelle sezioni principali
    h2_list = [h for h in headings if h.name == "h2"]
    h3_list = [h for h in headings if h.name == "h3"]
    if not h2_list:
        issues.append("WARN")
        details.append(
            "Nessun H2 trovato. Google richiede H2/H3 per le sezioni dei deep link."
        )
    else:
        issues.append("OK")
        details.append(f"Trovati {len(h2_list)} H2 e {len(h3_list)} H3.")

    # 2d. Blocchi di testo lungo senza heading vicino (>300 parole tra un heading e l'altro)
    orphan_text_blocks = []
    body = soup.find("body") or soup
    all_nodes = body.find_all(["h1","h2","h3","h4","h5","h6","p","div","section","article"])
    current_text = []
    current_words = 0
    for node in all_nodes:
        if node.name in ["h1","h2","h3","h4","h5","h6"]:
            if current_words > 300:
                snippet = " ".join(current_text)[:120]
                orphan_text_blocks.append(f"…{snippet}…")
            current_text = []
            current_words = 0
        elif node.name in ["p"]:
            t = node.get_text(strip=True)
            words = len(t.split())
            if words > 10:
                current_text.append(t)
                current_words += words

    if orphan_text_blocks:
        issues.append("WARN")
        details.append(
            f"Trovati {len(orphan_text_blocks)} blocchi di testo >300 parole privi di heading. "
            f"Strutturare con H2/H3 migliora eligibilità deep link."
        )
    else:
        issues.append("OK")
        details.append("Nessun blocco di testo lungo senza heading rilevato.")

    overall_score = issues.count("FAIL") * 2 + issues.count("WARN")
    if issues.count("FAIL") > 0:
        overall = "FAIL"
    elif overall_score >= 2:
        overall = "WARN"
    else:
        overall = "OK"

    return {
        "area": "2. Struttura Heading H2/H3",
        "overall": overall,
        "issues": issues,
        "summary": " | ".join(details),
        "heading_map": heading_map,
        "h1_count": len(h1_list),
        "h2_count": len(h2_list),
        "h3_count": len(h3_list),
        "level_jumps": jumps,
        "orphan_blocks": orphan_text_blocks,
    }


# ─── AREA 3: ROBOTS.TXT ──────────────────────────────────────────────────────

def fetch_robots_txt(base_url, session):
    parsed = urlparse(base_url)
    robots_url = f"{parsed.scheme}://{parsed.netloc}/robots.txt"
    try:
        r = session.get(robots_url, timeout=REQUEST_TIMEOUT)
        if r.status_code == 200:
            return r.text, robots_url
        else:
            return None, robots_url
    except Exception as e:
        return None, robots_url


def parse_robots_txt(content):
    """Parse robots.txt e restituisce struttura per l'analisi."""
    lines = content.splitlines()
    parsed = []
    for i, raw_line in enumerate(lines, 1):
        line = raw_line.strip()
        if not line or line.startswith("#"):
            continue
        if ":" in line:
            parts = line.split(":", 1)
            directive = parts[0].strip()
            value = parts[1].strip() if len(parts) > 1 else ""
        else:
            directive = line
            value = ""
        parsed.append({"line": i, "raw": raw_line, "directive": directive, "value": value})
    return parsed


def audit_robots_txt(base_url, session):
    """
    Verifica:
    - Presenza del file
    - Direttive non supportate da Google
    - Typo su Disallow
    - Sitemap dichiarata
    - Regole per Googlebot
    """
    issues = []
    details = []
    findings = []

    content, robots_url = fetch_robots_txt(base_url, session)

    if content is None:
        return {
            "area": "3. Robots.txt",
            "overall": "WARN",
            "summary": f"robots.txt non trovato o non accessibile ({robots_url})",
            "robots_url": robots_url,
            "findings": [],
        }

    parsed = parse_robots_txt(content)

    # 3a. Direttive non supportate
    unsupported_found = []
    for entry in parsed:
        d_lower = entry["directive"].lower()
        if d_lower in UNSUPPORTED_DIRECTIVES:
            unsupported_found.append(entry)
            findings.append({
                "type": "UNSUPPORTED_DIRECTIVE",
                "severity": "WARN",
                "line": entry["line"],
                "directive": entry["directive"],
                "value": entry["value"],
                "note": f"'{entry['directive']}' è ignorato da Google. "
                        f"Non usarlo per controllare l'accesso di Googlebot.",
            })

    if unsupported_found:
        issues.append("WARN")
        directives_list = list({e["directive"].lower() for e in unsupported_found})
        details.append(
            f"Trovate {len(unsupported_found)} direttive non supportate da Google: "
            f"{', '.join(directives_list)}."
        )
    else:
        issues.append("OK")
        details.append("Nessuna direttiva non supportata rilevata.")

    # 3b. Typo su Disallow
    typos_found = []
    for entry in parsed:
        d = entry["directive"]
        d_lower = d.lower()
        for typo in DISALLOW_TYPOS:
            if d_lower == typo.lower() and d_lower != "disallow":
                typos_found.append(entry)
                findings.append({
                    "type": "TYPO",
                    "severity": "WARN",
                    "line": entry["line"],
                    "directive": d,
                    "value": entry["value"],
                    "note": f"Possibile typo di 'Disallow': «{d}». "
                            f"Google potrebbe ignorare questa regola.",
                })
                break
        # Controlla anche maiuscole miste
        if d != "Disallow" and d != "disallow" and d_lower == "disallow":
            if d not in [t for t in DISALLOW_TYPOS]:
                typos_found.append(entry)
                findings.append({
                    "type": "TYPO",
                    "severity": "INFO",
                    "line": entry["line"],
                    "directive": d,
                    "value": entry["value"],
                    "note": f"Casing non standard: «{d}». Usa 'Disallow' (con D maiuscola).",
                })

    if typos_found:
        issues.append("WARN")
        details.append(f"Trovati {len(typos_found)} possibili typo di Disallow.")
    else:
        issues.append("OK")
        details.append("Nessun typo di 'Disallow' rilevato.")

    # 3c. Sitemap
    sitemaps = [e for e in parsed if e["directive"].lower() == "sitemap"]
    if not sitemaps:
        issues.append("WARN")
        details.append("Nessuna dichiarazione Sitemap nel robots.txt.")
        findings.append({
            "type": "MISSING_SITEMAP",
            "severity": "WARN",
            "line": "—",
            "directive": "Sitemap",
            "value": "—",
            "note": "Aggiungi 'Sitemap: https://tuodominio.com/sitemap.xml' nel robots.txt.",
        })
    else:
        issues.append("OK")
        details.append(f"Sitemap dichiarata: {sitemaps[0]['value']}")

    # 3d. Regole per Googlebot
    lines_content = content.lower()
    has_googlebot_section = "user-agent: googlebot" in lines_content
    has_wildcard = "user-agent: *" in lines_content

    if not has_googlebot_section and not has_wildcard:
        issues.append("WARN")
        details.append(
            "Nessuna sezione 'User-agent: Googlebot' né 'User-agent: *' trovata."
        )
        findings.append({
            "type": "MISSING_GOOGLEBOT",
            "severity": "WARN",
            "line": "—",
            "directive": "User-agent",
            "value": "Googlebot",
            "note": "Assicurati di avere almeno 'User-agent: *' con le regole appropriate.",
        })
    else:
        issues.append("OK")
        agent = "Googlebot" if has_googlebot_section else "* (wildcard)"
        details.append(f"Regole per {agent} presenti.")

    # 3e. Blocchi su URL importanti (verifica /sitemap.xml non bloccato)
    disallow_rules = [e for e in parsed if e["directive"].lower() == "disallow"]
    blocked_sitemap = any("/sitemap" in e["value"] for e in disallow_rules)
    if blocked_sitemap:
        issues.append("FAIL")
        details.append("⚠ ATTENZIONE: /sitemap sembra bloccato da una regola Disallow!")
        findings.append({
            "type": "BLOCKED_SITEMAP",
            "severity": "FAIL",
            "line": "—",
            "directive": "Disallow",
            "value": "/sitemap",
            "note": "La sitemap non deve essere bloccata da Disallow.",
        })

    overall_score = issues.count("FAIL") * 2 + issues.count("WARN")
    if "FAIL" in issues:
        overall = "FAIL"
    elif overall_score >= 2:
        overall = "WARN"
    else:
        overall = "OK"

    return {
        "area": "3. Robots.txt",
        "overall": overall,
        "issues": issues,
        "summary": " | ".join(details),
        "robots_url": robots_url,
        "robots_content": content,
        "total_lines": len(parsed),
        "unsupported_count": len(unsupported_found),
        "typos_count": len(typos_found),
        "sitemap_declared": bool(sitemaps),
        "findings": findings,
    }



# ─── AREA 8: E-E-A-T SIGNALS ─────────────────────────────────────────────────

# Schemi che indicano autorevolezza autore/organizzazione
EEAT_AUTHOR_SCHEMA_TYPES = {"Person", "Author", "Organization", "NewsMediaOrganization"}

# Selettori CSS comuni per firma autore
AUTHOR_CSS_SELECTORS = [
    "[rel='author']", "[class*='author']", "[id*='author']",
    "[class*='byline']", "[class*='writer']", "[class*='editor']",
    "[itemprop='author']", "[class*='contributor']",
    ".post-author", ".entry-author", ".article-author",
]

# Selettori per bio autore
BIO_CSS_SELECTORS = [
    "[class*='bio']", "[class*='about-author']", "[class*='author-info']",
    "[class*='author-box']", "[class*='author-card']", "[class*='author-profile']",
]

# Selettori per segnali di fiducia/trust
TRUST_CSS_SELECTORS = [
    "[class*='trust']", "[class*='certified']", "[class*='verified']",
    "[class*='award']", "[class*='badge']", "[class*='accredit']",
    "[class*='partner']", "[class*='press']", "[class*='mention']",
    "[class*='as-seen']", "[class*='featured-in']",
]

# Keyword YMYL che alzano l'asticella E-E-A-T
YMYL_KEYWORDS = [
    "prestito", "mutuo", "investimento", "finanziamento", "assicurazione",
    "salute", "medico", "farmaco", "diagnosi", "terapia", "legale", "avvocato",
    "fiscal", "tasse", "reddito", "pensione", "loan", "mortgage", "insurance",
    "health", "medical", "doctor", "legal", "lawyer", "tax", "investment",
]

def audit_eeat_signals(url, static_html):
    """
    Rileva segnali on-page di Experience, Expertise, Authoritativeness, Trustworthiness.
    Verifica: firma autore, bio, data pubblicazione/aggiornamento, schema Person/Organization,
    link a fonti esterne autorevoli, segnali trust, profondità contenuto.
    """
    issues = []
    details = []
    signals = []
    missing = []

    if not static_html:
        return {"area": "8. E-E-A-T Signals", "overall": "ERROR",
                "summary": "HTML non disponibile", "signals": [], "missing": []}

    soup = BeautifulSoup(static_html, "lxml")
    score = 0  # punteggio interno 0-100

    # ── 8a. Firma autore ─────────────────────────────────────────────────────
    author_found = []
    for sel in AUTHOR_CSS_SELECTORS:
        try:
            els = soup.select(sel)
            for el in els:
                txt = el.get_text(strip=True)
                if len(txt) > 2:
                    author_found.append(txt[:80])
        except Exception:
            pass

    # Cerca anche in JSON-LD
    author_schema = []
    for block in soup.find_all("script", type="application/ld+json"):
        try:
            data = json.loads(block.string or "")
            items = data if isinstance(data, list) else [data]
            for item in items:
                if not isinstance(item, dict):
                    continue
                author = item.get("author", {})
                if isinstance(author, dict):
                    name = author.get("name", "")
                    if name:
                        author_schema.append(name)
                elif isinstance(author, list):
                    for a in author:
                        if isinstance(a, dict) and a.get("name"):
                            author_schema.append(a["name"])
        except Exception:
            pass

    all_authors = list(set(author_found + author_schema))
    if all_authors:
        score += 20
        signals.append(f"Autore identificato: {', '.join(all_authors[:3])}")
    else:
        missing.append("Firma autore assente — aggiungere rel='author' o schema Person")

    # ── 8b. Bio autore ───────────────────────────────────────────────────────
    bio_found = False
    for sel in BIO_CSS_SELECTORS:
        try:
            if soup.select(sel):
                bio_found = True
                break
        except Exception:
            pass

    if bio_found:
        score += 10
        signals.append("Bio autore presente nella pagina")
    else:
        missing.append("Bio autore assente — aggiungere sezione con credenziali e profilo autore")

    # ── 8c. Data pubblicazione e aggiornamento ───────────────────────────────
    date_pub = soup.find(attrs={"itemprop": "datePublished"}) or \
               soup.find("time", attrs={"datetime": True}) or \
               soup.select_one("[class*='publish']") or \
               soup.select_one("[class*='posted']")

    date_mod = soup.find(attrs={"itemprop": "dateModified"}) or \
               soup.select_one("[class*='updated']") or \
               soup.select_one("[class*='modified']")

    if date_pub:
        score += 10
        dt = date_pub.get("datetime", date_pub.get_text(strip=True))[:20]
        signals.append(f"Data di pubblicazione presente: {dt}")
    else:
        missing.append("Data pubblicazione assente — aiuta Google a valutare freschezza contenuto")

    if date_mod:
        score += 5
        signals.append("Data di aggiornamento presente")

    # ── 8d. Link a fonti esterne autorevoli ──────────────────────────────────
    authoritative_domains = [
        ".gov", ".edu", ".org", "wikipedia.org", "pubmed.ncbi", "scholar.google",
        "who.int", "ema.europa", "agcm.it", "bancaditalia.it", "istat.it",
        "mise.gov.it", "mef.gov.it", "consob.it", "inps.it",
    ]
    external_links = soup.find_all("a", href=True)
    auth_links = []
    for a in external_links:
        href = a.get("href", "")
        if any(d in href for d in authoritative_domains):
            auth_links.append({"url": href, "text": a.get_text(strip=True)[:60]})

    if auth_links:
        score += 15
        signals.append(f"{len(auth_links)} link verso fonti autorevoli ({', '.join(set(l['url'].split('/')[2] for l in auth_links[:3]))})")
    else:
        missing.append("Nessun link a fonti esterne autorevoli — citare studi, dati ufficiali, fonti primarie")

    # ── 8e. Schema Organization / Person ────────────────────────────────────
    org_schema = False
    for block in soup.find_all("script", type="application/ld+json"):
        try:
            data = json.loads(block.string or "")
            items = data if isinstance(data, list) else ([data.get("@graph", data)] if isinstance(data, dict) else [])
            if isinstance(items, dict):
                items = [items]
            flat = []
            for i in items:
                if isinstance(i, list):
                    flat.extend(i)
                elif isinstance(i, dict):
                    flat.append(i)
            for item in flat:
                if not isinstance(item, dict):
                    continue
                t = item.get("@type", "")
                if isinstance(t, list):
                    t = t[0] if t else ""
                if any(s in t for s in EEAT_AUTHOR_SCHEMA_TYPES):
                    org_schema = True
                    break
        except Exception:
            pass

    if org_schema:
        score += 15
        signals.append("Schema Organization/Person presente — identità verificabile dai motori")
    else:
        missing.append("Schema Organization/Person assente — aggiungere JSON-LD con dati aziendali")

    # ── 8f. Segnali trust e riconoscimenti ───────────────────────────────────
    trust_els = []
    for sel in TRUST_CSS_SELECTORS:
        try:
            els = soup.select(sel)
            trust_els.extend(els)
        except Exception:
            pass

    # Cerca "as seen in", "featured in", menzioni press
    press_keywords = ["as seen", "featured in", "come visto", "citato da",
                      "premi", "riconoscimenti", "certificazioni", "menzionato"]
    body_text = soup.get_text(" ", strip=True).lower()
    press_found = [kw for kw in press_keywords if kw in body_text]

    if trust_els or press_found:
        score += 10
        signals.append(f"Segnali di trust/riconoscimento presenti ({len(trust_els)} elementi)")
    else:
        missing.append("Nessun segnale di trust visibile (premi, citazioni press, loghi partner)")

    # ── 8g. Profondità contenuto testuale ────────────────────────────────────
    main_content = soup.find("main") or soup.find("article") or soup.find("body")
    if main_content:
        for tag in main_content.find_all(["nav", "header", "footer", "aside", "script", "style"]):
            tag.decompose()
        word_count = len(main_content.get_text(" ", strip=True).split())
    else:
        word_count = 0

    if word_count >= 800:
        score += 10
        signals.append(f"Contenuto sostanzioso: {word_count} parole")
    elif word_count >= 300:
        score += 5
        signals.append(f"Contenuto di media lunghezza: {word_count} parole")
    else:
        missing.append(f"Contenuto scarso: solo {word_count} parole — Google preferisce pagine esaurienti")

    # ── 8h. Segnali YMYL ─────────────────────────────────────────────────────
    is_ymyl = any(kw in body_text for kw in YMYL_KEYWORDS)
    ymyl_note = ""
    if is_ymyl:
        ymyl_note = "⚠ Pagina YMYL rilevata — E-E-A-T è critico per questo settore"
        if score < 50:
            issues.append("FAIL")
        elif score < 70:
            issues.append("WARN")

    # ── Overall ──────────────────────────────────────────────────────────────
    if score >= 70:
        overall = "OK"
        details.append(f"Buoni segnali E-E-A-T ({score}/100): {len(signals)} indicatori positivi.")
    elif score >= 40:
        overall = "WARN"
        details.append(f"Segnali E-E-A-T parziali ({score}/100): {len(missing)} aree da migliorare.")
    else:
        overall = "FAIL"
        details.append(f"Segnali E-E-A-T insufficienti ({score}/100): {len(missing)} elementi critici mancanti.")

    if ymyl_note:
        details.append(ymyl_note)

    return {
        "area": "8. E-E-A-T Signals",
        "overall": overall,
        "summary": " | ".join(details),
        "score": score,
        "signals": signals,
        "missing": missing,
        "author_found": all_authors,
        "bio_found": bio_found,
        "date_pub": bool(date_pub),
        "date_mod": bool(date_mod),
        "auth_links": auth_links[:10],
        "org_schema": org_schema,
        "word_count": word_count,
        "is_ymyl": is_ymyl,
        "trust_signals": len(trust_els),
    }


# ─── AREA 9: PERFORMANCE SIGNALS ─────────────────────────────────────────────

# Framework JS pesanti (indicatori di potenziale INP alto)
JS_HEAVY_FRAMEWORKS = {
    "react": "React (SSR consigliato per LCP)",
    "vue": "Vue.js (verificare hydration impact su INP)",
    "angular": "Angular (bundle tipicamente pesante)",
    "next": "Next.js (verificare modalità rendering)",
    "nuxt": "Nuxt.js (verificare modalità rendering)",
    "gatsby": "Gatsby (build statico, generalmente OK)",
    "ember": "Ember.js (bundle pesante)",
    "backbone": "Backbone.js",
    "knockout": "Knockout.js",
}

# CDN e ottimizzatori noti
CDN_PROVIDERS = [
    "cloudflare", "fastly", "akamai", "cloudfront", "cdn.jsdelivr",
    "unpkg.com", "cdnjs.cloudflare", "jsdelivr", "bunny.net",
]

def audit_performance_signals(url, static_html):
    """
    Rileva segnali indiretti di performance senza Lighthouse.
    Verifica: immagini senza dimensioni (CLS), immagini hero senza preload,
    font bloccanti, framework JS pesanti, CDN, lazy loading corretto,
    numero script/CSS, presenza viewport meta.
    NON sostituisce Lighthouse/CrUX — fornisce segnali rapidi da DOM statico.
    """
    issues = []
    details = []
    findings = []
    positives = []

    if not static_html:
        return {"area": "9. Performance Signals", "overall": "ERROR",
                "summary": "HTML non disponibile", "findings": [], "positives": []}

    soup = BeautifulSoup(static_html, "lxml")
    score = 100  # si parte da 100 e si scala

    # ── 9a. Viewport meta (prerequisito mobile-first) ─────────────────────────
    viewport = soup.find("meta", attrs={"name": "viewport"})
    if viewport:
        positives.append("Meta viewport presente")
    else:
        score -= 15
        findings.append({
            "tipo": "Meta viewport assente",
            "severity": "FAIL",
            "impatto": "LCP / Mobile-first indexing",
            "nota": "Aggiungere <meta name='viewport' content='width=device-width,initial-scale=1'>",
        })

    # ── 9b. Immagini senza width/height (Cumulative Layout Shift) ────────────
    images = soup.find_all("img")
    imgs_no_size = [img for img in images
                    if not (img.get("width") and img.get("height"))
                    and not img.get("loading") == "lazy"
                    and img.get("src", "").startswith("http")]
    imgs_no_alt = [img for img in images if not img.get("alt")]

    if imgs_no_size:
        penalty = min(20, len(imgs_no_size) * 3)
        score -= penalty
        findings.append({
            "tipo": f"Immagini senza dimensioni esplicite ({len(imgs_no_size)})",
            "severity": "WARN",
            "impatto": "CLS (Cumulative Layout Shift)",
            "nota": "Aggiungere width e height a ogni <img> per evitare spostamenti layout durante il caricamento",
        })
    else:
        positives.append(f"Tutte le immagini above-the-fold hanno dimensioni dichiarate")

    if imgs_no_alt:
        findings.append({
            "tipo": f"Immagini senza attributo alt ({len(imgs_no_alt)})",
            "severity": "WARN",
            "impatto": "Accessibilità e indicizzazione immagini",
            "nota": "Aggiungere alt descrittivo a ogni immagine — Google Image Search e screen reader",
        })

    # ── 9c. Immagine hero senza fetchpriority ─────────────────────────────────
    first_img = soup.find("img")
    if first_img:
        has_priority = first_img.get("fetchpriority") == "high" or \
                       first_img.get("loading") != "lazy"
        if not has_priority:
            score -= 10
            findings.append({
                "tipo": "Prima immagine senza fetchpriority=high",
                "severity": "WARN",
                "impatto": "LCP (Largest Contentful Paint)",
                "nota": "Aggiungere fetchpriority='high' all'immagine hero per accelerare LCP",
            })
        else:
            positives.append("Prima immagine ha priorità di caricamento corretta")

    # ── 9d. Font bloccanti ────────────────────────────────────────────────────
    link_tags = soup.find_all("link", rel=True)
    blocking_fonts = []
    preloaded_fonts = []
    for link in link_tags:
        rels = link.get("rel", [])
        if isinstance(rels, str):
            rels = [rels]
        href = link.get("href", "")
        if "stylesheet" in rels and ("fonts.googleapis" in href or "typekit" in href or "fonts.adobe" in href):
            if "preload" not in rels:
                blocking_fonts.append(href[:80])
        if "preload" in rels and link.get("as") == "font":
            preloaded_fonts.append(href[:60])

    if blocking_fonts:
        score -= 10
        findings.append({
            "tipo": f"Font Google/Adobe caricati in modo bloccante ({len(blocking_fonts)})",
            "severity": "WARN",
            "impatto": "FCP (First Contentful Paint) / LCP",
            "nota": "Usare display=swap e preload per i font critici. Valutare font-system come alternativa.",
        })
    if preloaded_fonts:
        positives.append(f"{len(preloaded_fonts)} font con preload corretto")

    # ── 9e. Framework JS pesanti ──────────────────────────────────────────────
    all_scripts = soup.find_all("script")
    script_srcs = [s.get("src", "") for s in all_scripts if s.get("src")]
    inline_scripts = [s for s in all_scripts if s.string and len(s.string or "") > 100]

    detected_frameworks = []
    for src in script_srcs:
        src_low = src.lower()
        for fw_key, fw_label in JS_HEAVY_FRAMEWORKS.items():
            if fw_key in src_low:
                detected_frameworks.append(fw_label)
                break

    # Controlla anche nei tag inline
    inline_text = " ".join(s.string or "" for s in inline_scripts).lower()
    for fw_key, fw_label in JS_HEAVY_FRAMEWORKS.items():
        if fw_key in inline_text and fw_label not in detected_frameworks:
            detected_frameworks.append(fw_label)

    if detected_frameworks:
        findings.append({
            "tipo": f"Framework JS rilevati: {', '.join(detected_frameworks)}",
            "severity": "INFO",
            "impatto": "INP (Interaction to Next Paint) / TTI",
            "nota": "Verificare INP con PageSpeed Insights. Priorità a SSR/SSG e code splitting.",
        })

    # Numero script esterni
    ext_scripts = [s for s in script_srcs if s.startswith("http")]
    if len(ext_scripts) > 10:
        score -= 10
        findings.append({
            "tipo": f"{len(ext_scripts)} script esterni (>10)",
            "severity": "WARN",
            "impatto": "TBT (Total Blocking Time) / INP",
            "nota": "Ridurre script di terze parti o caricarli con defer/async",
        })
    else:
        positives.append(f"Script esterni contenuti ({len(ext_scripts)})")

    # ── 9f. Script senza defer/async ─────────────────────────────────────────
    blocking_scripts = [s for s in all_scripts
                        if s.get("src")
                        and not s.get("defer")
                        and not s.get("async")
                        and not s.get("type") == "module"]
    if blocking_scripts:
        score -= 10
        findings.append({
            "tipo": f"{len(blocking_scripts)} script bloccanti (senza defer/async)",
            "severity": "WARN",
            "impatto": "FCP / TBT",
            "nota": "Aggiungere defer o async agli script non critici per il render iniziale",
        })
    else:
        positives.append("Tutti gli script usano defer/async/module")

    # ── 9g. CDN rilevato ──────────────────────────────────────────────────────
    page_src = static_html.lower()
    cdn_found = [cdn for cdn in CDN_PROVIDERS if cdn in page_src]
    if cdn_found:
        positives.append(f"CDN rilevato: {cdn_found[0]}")

    # ── 9h. Preconnect / DNS prefetch ─────────────────────────────────────────
    preconnect = [l for l in link_tags
                  if "preconnect" in (l.get("rel") or [])
                  or "dns-prefetch" in (l.get("rel") or [])]
    if preconnect:
        positives.append(f"{len(preconnect)} hint preconnect/dns-prefetch presenti")
    else:
        findings.append({
            "tipo": "Nessun preconnect/dns-prefetch dichiarato",
            "severity": "INFO",
            "impatto": "TTFB verso domini terze parti",
            "nota": "Aggiungere <link rel='preconnect'> per Google Fonts, CDN e analytics",
        })

    # ── 9i. CSS bloccanti inline ──────────────────────────────────────────────
    style_tags = soup.find_all("style")
    large_inline_css = [s for s in style_tags if len(s.string or "") > 50000]
    if large_inline_css:
        score -= 5
        findings.append({
            "tipo": f"{len(large_inline_css)} blocchi CSS inline molto grandi (>50KB)",
            "severity": "INFO",
            "impatto": "Parse CSS / FCP",
            "nota": "Valutare se estrarre CSS non critico in file separati con lazy load",
        })

    # ── Overall ───────────────────────────────────────────────────────────────
    score = max(0, min(100, score))
    fail_count = sum(1 for f in findings if f["severity"] == "FAIL")
    warn_count = sum(1 for f in findings if f["severity"] == "WARN")

    if fail_count > 0 or score < 40:
        overall = "FAIL"
        details.append(f"Segnali performance critici: {fail_count} FAIL, {warn_count} WARN. Score stimato: {score}/100.")
    elif warn_count >= 3 or score < 65:
        overall = "WARN"
        details.append(f"Segnali performance da migliorare: {warn_count} WARN. Score stimato: {score}/100.")
    else:
        overall = "OK"
        details.append(f"Segnali performance positivi. Score stimato: {score}/100. {len(positives)} elementi ottimizzati.")

    details.append("⚠ Verifica con PageSpeed Insights per dati reali LCP/INP/CLS.")

    return {
        "area": "9. Performance Signals",
        "overall": overall,
        "summary": " | ".join(details),
        "score": score,
        "findings": findings,
        "positives": positives,
        "images_total": len(images),
        "imgs_no_size": len(imgs_no_size),
        "imgs_no_alt": len(imgs_no_alt),
        "blocking_scripts": len(blocking_scripts),
        "ext_scripts": len(ext_scripts),
        "frameworks": detected_frameworks,
        "blocking_fonts": blocking_fonts,
        "preconnect_count": len(preconnect),
    }


# ─── AREA 10: AUTOREVOLEZZA TOPICA ───────────────────────────────────────────

# Selettori per segnali di profondità tematica
TOPICAL_DEPTH_SELECTORS = [
    "table", "figure", "figcaption",
    "[class*='statistic']", "[class*='data']", "[class*='chart']",
    "[class*='infographic']", "[class*='research']", "[class*='study']",
    "blockquote", "cite",
]

# Pattern per citazioni e riferimenti
CITATION_PATTERNS = [
    r"fonte[:\s]", r"secondo\s+(?:uno studio|la ricerca|i dati|il rapporto)",
    r"studi(?:o)?\s+(?:del|della|di)", r"ricerca(?:tori)?\s+(?:del|di)",
    r"dati\s+(?:del|della|di)", r"\d{4}\s+study", r"according to",
    r"source[:\s]", r"cited\s+by", r"research\s+(?:by|from|shows)",
    r"\[\d+\]",  # note a piè di pagina stile accademico
]

# Segnali di internal linking tematico
INTERNAL_LINK_SIGNALS = [
    "[class*='related']", "[class*='correlat']", "[class*='similar']",
    "[class*='further-reading']", "[class*='see-also']", "[class*='approfondim']",
    "[class*='leggi-anche']", "[class*='ti-potrebbe']",
]

def audit_topical_authority(url, static_html, session):
    """
    Rileva segnali di autorevolezza topica on-page.
    Verifica: profondità contenuto, citazioni fonti, uso tabelle/dati,
    internal linking contestuale, breadcrumb, copertura semantica argomento,
    presenza di meta description ottimizzata.
    """
    issues = []
    details = []
    signals = []
    opportunities = []

    if not static_html:
        return {"area": "10. Autorevolezza Topica", "overall": "ERROR",
                "summary": "HTML non disponibile", "signals": [], "opportunities": []}

    soup = BeautifulSoup(static_html, "lxml")
    parsed = urlparse(url)
    base_domain = f"{parsed.scheme}://{parsed.netloc}"
    score = 0

    # ── 10a. Meta title e description ─────────────────────────────────────────
    title_tag = soup.find("title")
    meta_desc = soup.find("meta", attrs={"name": "description"})
    canonical = soup.find("link", attrs={"rel": "canonical"})

    title_text = title_tag.get_text(strip=True) if title_tag else ""
    title_len = len(title_text)
    desc_text = meta_desc.get("content", "") if meta_desc else ""
    desc_len = len(desc_text)

    meta_issues = []
    if not title_text:
        meta_issues.append("Title assente")
        opportunities.append("Title tag mancante — elemento SEO fondamentale")
    elif title_len < 30:
        meta_issues.append(f"Title troppo corto ({title_len} car.)")
        opportunities.append(f"Title troppo corto ({title_len} car.) — ottimale 50-60 caratteri")
    elif title_len > 65:
        meta_issues.append(f"Title troppo lungo ({title_len} car., troncato in SERP)")
        opportunities.append(f"Title troppo lungo ({title_len} car.) — ridurre a 50-60 caratteri")
    else:
        score += 10
        signals.append(f"Title ottimizzato ({title_len} car.): «{title_text[:60]}»")

    if not desc_text:
        meta_issues.append("Meta description assente")
        opportunities.append("Meta description mancante — impatta CTR in SERP")
    elif desc_len < 70:
        meta_issues.append(f"Meta description corta ({desc_len} car.)")
        opportunities.append(f"Meta description corta ({desc_len} car.) — ottimale 120-160 caratteri")
    elif desc_len > 165:
        meta_issues.append(f"Meta description lunga ({desc_len} car., troncata in SERP)")
    else:
        score += 10
        signals.append(f"Meta description ottimizzata ({desc_len} car.)")

    if canonical:
        score += 5
        canon_href = canonical.get("href", "")
        signals.append(f"Canonical tag presente: {canon_href[:80]}")
    else:
        opportunities.append("Canonical tag assente — rischio contenuto duplicato")

    # ── 10b. Breadcrumb navigation ────────────────────────────────────────────
    breadcrumb_selectors = [
        "[class*='breadcrumb']", "[aria-label*='breadcrumb']",
        "[class*='fil-']", "nav[aria-label]", "[itemtype*='BreadcrumbList']",
        "[class*='percorso']",
    ]
    breadcrumb_found = False
    for sel in breadcrumb_selectors:
        try:
            if soup.select(sel):
                breadcrumb_found = True
                break
        except Exception:
            pass

    if breadcrumb_found:
        score += 10
        signals.append("Breadcrumb navigation presente — aiuta Google a capire la struttura del sito")
    else:
        opportunities.append("Breadcrumb assente — utile per struttura sito e rich results in SERP")

    # ── 10c. Citazioni e riferimenti a fonti ──────────────────────────────────
    body_text = soup.get_text(" ", strip=True)
    citation_matches = []
    for pattern in CITATION_PATTERNS:
        matches = re.findall(pattern, body_text.lower())
        citation_matches.extend(matches)

    # Cerca anche cite/blockquote
    cite_tags = soup.find_all(["cite", "blockquote"])
    external_links = [a for a in soup.find_all("a", href=True)
                      if a.get("href", "").startswith("http")
                      and parsed.netloc not in a.get("href", "")]

    if len(citation_matches) >= 2 or len(cite_tags) >= 1:
        score += 15
        signals.append(f"Contenuto cita fonti esterne ({len(citation_matches)} pattern, {len(cite_tags)} tag cite/blockquote)")
    elif external_links:
        score += 5
        signals.append(f"{len(external_links)} link a risorse esterne presenti")
    else:
        opportunities.append("Nessuna citazione di fonti — aggiungere riferimenti a dati/ricerche per E-E-A-T")

    # ── 10d. Uso di dati strutturati visivi (tabelle, grafici, figure) ────────
    tables = soup.find_all("table")
    figures = soup.find_all("figure")
    figcaptions = soup.find_all("figcaption")

    depth_els = []
    for sel in TOPICAL_DEPTH_SELECTORS:
        try:
            depth_els.extend(soup.select(sel))
        except Exception:
            pass

    if tables or figures:
        score += 10
        signals.append(f"Contenuto arricchito con dati visivi: {len(tables)} tabelle, {len(figures)} figure")
    else:
        opportunities.append("Nessuna tabella o figura — dati visivi aumentano engagement e citabilità da AI")

    # ── 10e. Internal linking tematico ────────────────────────────────────────
    internal_links = [a for a in soup.find_all("a", href=True)
                      if a.get("href", "").startswith("/")
                      or (a.get("href", "").startswith("http") and parsed.netloc in a.get("href", ""))]

    related_sections = []
    for sel in INTERNAL_LINK_SIGNALS:
        try:
            els = soup.select(sel)
            related_sections.extend(els)
        except Exception:
            pass

    if len(internal_links) >= 5:
        score += 10
        signals.append(f"{len(internal_links)} link interni presenti")
    else:
        opportunities.append(f"Pochi link interni ({len(internal_links)}) — aumentare per distribuire autorità e guidare crawling")

    if related_sections:
        score += 5
        signals.append(f"Sezione 'articoli correlati' o 'leggi anche' presente ({len(related_sections)} elementi)")
    else:
        opportunities.append("Nessuna sezione 'articoli correlati' — aiuta utenti e crawler a scoprire contenuti correlati")

    # ── 10f. Copertura semantica (lista di H2/H3 come proxy) ─────────────────
    h2_list = [h.get_text(strip=True) for h in soup.find_all("h2")]
    h3_list = [h.get_text(strip=True) for h in soup.find_all("h3")]
    total_sections = len(h2_list) + len(h3_list)

    if total_sections >= 5:
        score += 10
        signals.append(f"Copertura semantica ampia: {len(h2_list)} H2 + {len(h3_list)} H3 = {total_sections} sezioni")
    elif total_sections >= 2:
        score += 5
        signals.append(f"Copertura semantica parziale: {total_sections} sezioni totali (H2+H3)")
    else:
        opportunities.append("Struttura contenuto piatta — aumentare numero di sezioni tematiche con H2/H3")

    # ── 10g. Lunghezza contenuto come proxy di completezza ────────────────────
    main_el = soup.find("main") or soup.find("article") or soup.find("body")
    if main_el:
        for tag in main_el.find_all(["nav","header","footer","aside","script","style"]):
            tag.decompose()
        word_count = len(main_el.get_text(" ", strip=True).split())
    else:
        word_count = len(body_text.split())

    if word_count >= 1500:
        score += 15
        signals.append(f"Contenuto esaustivo: {word_count} parole — ottimo segnale topicale")
    elif word_count >= 600:
        score += 8
        signals.append(f"Contenuto di buona lunghezza: {word_count} parole")
    elif word_count >= 200:
        score += 3
        opportunities.append(f"Contenuto corto ({word_count} parole) — espandere per coprire l'argomento in profondità")
    else:
        opportunities.append(f"Contenuto molto scarso ({word_count} parole) — rischio di essere considerato thin content")

    # ── Overall ───────────────────────────────────────────────────────────────
    score = min(100, score)

    if meta_issues:
        issues.append("WARN")
        details.append(f"Meta tag: {', '.join(meta_issues)}.")

    if score >= 65:
        overall = "OK"
        details.append(f"Buona autorevolezza topica ({score}/100): {len(signals)} segnali positivi.")
    elif score >= 35:
        overall = "WARN"
        details.append(f"Autorevolezza topica parziale ({score}/100): {len(opportunities)} aree da rafforzare.")
    else:
        overall = "FAIL"
        details.append(f"Autorevolezza topica insufficiente ({score}/100): {len(opportunities)} opportunità critiche.")

    return {
        "area": "10. Autorevolezza Topica",
        "overall": overall,
        "summary": " | ".join(details),
        "score": score,
        "signals": signals,
        "opportunities": opportunities,
        "title_text": title_text,
        "title_len": title_len,
        "desc_text": desc_text,
        "desc_len": desc_len,
        "has_canonical": bool(canonical),
        "has_breadcrumb": breadcrumb_found,
        "internal_links_count": len(internal_links),
        "external_links_count": len(external_links),
        "tables_count": len(tables),
        "figures_count": len(figures),
        "word_count": word_count,
        "h2_list": h2_list[:10],
        "h3_list": h3_list[:10],
        "citation_found": len(citation_matches),
    }

# ─── OUTPUT EXCEL ─────────────────────────────────────────────────────────────

COLOR_OK   = "C6EFCE"   # verde
COLOR_WARN = "FFEB9C"   # giallo
COLOR_FAIL = "FFC7CE"   # rosso
COLOR_INFO = "DDEEFF"   # blu chiaro
COLOR_HEAD = "2F4F8F"   # intestazione

FONT_HEAD = Font(bold=True, color="FFFFFF", size=11)
FONT_BOLD = Font(bold=True, size=10)
FONT_NORM = Font(size=10)

ALIGN_WRAP = Alignment(wrap_text=True, vertical="top")
ALIGN_CTR  = Alignment(horizontal="center", vertical="center")

thin = Side(style="thin", color="AAAAAA")
BORDER = Border(left=thin, right=thin, top=thin, bottom=thin)


def cell_style(ws, row, col, value, fill_color=None, bold=False, center=False):
    c = ws.cell(row=row, column=col, value=value)
    if fill_color:
        c.fill = PatternFill("solid", fgColor=fill_color)
    c.font = Font(bold=bold, size=10)
    c.alignment = ALIGN_CTR if center else ALIGN_WRAP
    c.border = BORDER
    return c


def severity_color(sev):
    return {"OK": COLOR_OK, "WARN": COLOR_WARN, "FAIL": COLOR_FAIL, "INFO": COLOR_INFO}.get(sev, None)


def write_summary_sheet(wb, url, results, timestamp):
    ws = wb.create_sheet("📋 Riepilogo", 0)
    ws.column_dimensions["A"].width = 35
    ws.column_dimensions["B"].width = 12
    ws.column_dimensions["C"].width = 80

    # Titolo
    ws.merge_cells("A1:C1")
    c = ws["A1"]
    c.value = f"SEO Audit Report — {url}"
    c.font = Font(bold=True, size=14, color="FFFFFF")
    c.fill = PatternFill("solid", fgColor=COLOR_HEAD)
    c.alignment = ALIGN_WRAP

    ws.merge_cells("A2:C2")
    c = ws["A2"]
    c.value = f"Generato il: {timestamp}"
    c.font = FONT_NORM
    c.alignment = ALIGN_WRAP

    # Header tabella
    row = 4
    for col, hdr in enumerate(["Area di Analisi", "Esito", "Sintesi"], 1):
        c = ws.cell(row=row, column=col, value=hdr)
        c.fill = PatternFill("solid", fgColor=COLOR_HEAD)
        c.font = FONT_HEAD
        c.alignment = ALIGN_CTR
        c.border = BORDER

    row = 5
    for r in results:
        cell_style(ws, row, 1, r["area"], bold=True)
        cell_style(ws, row, 2, r["overall"], fill_color=severity_color(r["overall"]), center=True, bold=True)
        cell_style(ws, row, 3, r.get("summary", ""))
        ws.row_dimensions[row].height = 60
        row += 1

    ws.row_dimensions[4].height = 20


def write_visibility_sheet(wb, result):
    ws = wb.create_sheet("1. Visibilità Contenuti")
    ws.column_dimensions["A"].width = 25
    ws.column_dimensions["B"].width = 20
    ws.column_dimensions["C"].width = 20
    ws.column_dimensions["D"].width = 70

    # Header
    ws.merge_cells("A1:D1")
    c = ws["A1"]
    c.value = result["area"] + f"  —  Esito: {result['overall']}"
    c.font = Font(bold=True, size=12, color="FFFFFF")
    c.fill = PatternFill("solid", fgColor=COLOR_HEAD)
    c.alignment = ALIGN_WRAP

    row = 3
    # KPI
    kpis = [
        ("Blocchi contenuto nascosto", len(result.get("hidden_blocks", []))),
        ("Elementi lazy-load", result.get("lazy_elements", 0)),
        ("Pattern JS post-interazione", len(result.get("js_patterns", []))),
        ("Delta testo statico/JS", result.get("dom_delta", "N/A")),
    ]
    for label, val in kpis:
        cell_style(ws, row, 1, label, bold=True)
        cell_style(ws, row, 2, str(val), fill_color=COLOR_WARN if (isinstance(val, int) and val > 0) else COLOR_OK)
        ws.merge_cells(f"C{row}:D{row}")
        row += 1

    # Dettaglio blocchi nascosti
    row += 1
    for col, hdr in enumerate(["Selettore CSS", "Tag", "Contiene Heading?", "Anteprima Testo"], 1):
        c = ws.cell(row=row, column=col, value=hdr)
        c.fill = PatternFill("solid", fgColor=COLOR_HEAD)
        c.font = FONT_HEAD
        c.alignment = ALIGN_CTR
        c.border = BORDER

    row += 1
    for b in result.get("hidden_blocks", [])[:50]:
        cell_style(ws, row, 1, b["selector"])
        cell_style(ws, row, 2, b["tag"])
        cell_style(ws, row, 3, "Sì ⚠" if b["has_heading"] else "No",
                   fill_color=COLOR_WARN if b["has_heading"] else None)
        cell_style(ws, row, 4, b["text_preview"])
        ws.row_dimensions[row].height = 40
        row += 1

    if not result.get("hidden_blocks"):
        ws.cell(row=row, column=1, value="Nessun blocco nascosto rilevato.").font = FONT_NORM

    # Pattern JS
    row += 2
    ws.cell(row=row, column=1, value="Pattern JS rilevati:").font = FONT_BOLD
    row += 1
    for p in result.get("js_patterns", []):
        cell_style(ws, row, 1, p, fill_color=COLOR_WARN)
        row += 1
    if not result.get("js_patterns"):
        ws.cell(row=row, column=1, value="Nessuno").font = FONT_NORM


def write_heading_sheet(wb, result):
    ws = wb.create_sheet("2. Struttura Heading")
    ws.column_dimensions["A"].width = 10
    ws.column_dimensions["B"].width = 8
    ws.column_dimensions["C"].width = 80

    ws.merge_cells("A1:C1")
    c = ws["A1"]
    c.value = result["area"] + f"  —  Esito: {result['overall']}"
    c.font = Font(bold=True, size=12, color="FFFFFF")
    c.fill = PatternFill("solid", fgColor=COLOR_HEAD)
    c.alignment = ALIGN_WRAP

    row = 3
    kpis = [
        ("H1 trovati", result.get("h1_count", 0), result.get("h1_count", 0) != 1),
        ("H2 trovati", result.get("h2_count", 0), result.get("h2_count", 0) == 0),
        ("H3 trovati", result.get("h3_count", 0), False),
        ("Salti di livello", len(result.get("level_jumps", [])), len(result.get("level_jumps", [])) > 0),
        ("Blocchi testo orfani (>300 parole)", len(result.get("orphan_blocks", [])), len(result.get("orphan_blocks", [])) > 0),
    ]
    for label, val, is_issue in kpis:
        cell_style(ws, row, 1, label, bold=True)
        cell_style(ws, row, 2, str(val), fill_color=COLOR_WARN if is_issue else COLOR_OK, center=True)
        ws.merge_cells(f"C{row}:C{row}")
        row += 1

    # Mappa heading
    row += 1
    for col, hdr in enumerate(["Livello", "Tag", "Testo"], 1):
        c = ws.cell(row=row, column=col, value=hdr)
        c.fill = PatternFill("solid", fgColor=COLOR_HEAD)
        c.font = FONT_HEAD
        c.alignment = ALIGN_CTR
        c.border = BORDER

    row += 1
    for h in result.get("heading_map", []):
        indent = "  " * (h["level"] - 1)
        fill = None
        if h["level"] == 1:
            fill = COLOR_INFO
        elif h["level"] == 2:
            fill = "EEF4FF"
        cell_style(ws, row, 1, h["level"], center=True, fill_color=fill)
        cell_style(ws, row, 2, h["tag"], center=True, fill_color=fill, bold=(h["level"] <= 2))
        cell_style(ws, row, 3, indent + h["text"], fill_color=fill)
        ws.row_dimensions[row].height = 30
        row += 1

    # Salti di livello
    if result.get("level_jumps"):
        row += 1
        ws.cell(row=row, column=1, value="⚠ Salti di livello:").font = FONT_BOLD
        row += 1
        for j in result["level_jumps"]:
            cell_style(ws, row, 1, j, fill_color=COLOR_WARN)
            ws.merge_cells(f"A{row}:C{row}")
            row += 1


def write_robots_sheet(wb, result):
    ws = wb.create_sheet("3. Robots.txt")
    ws.column_dimensions["A"].width = 22
    ws.column_dimensions["B"].width = 10
    ws.column_dimensions["C"].width = 12
    ws.column_dimensions["D"].width = 20
    ws.column_dimensions["E"].width = 60

    ws.merge_cells("A1:E1")
    c = ws["A1"]
    c.value = result["area"] + f"  —  Esito: {result['overall']}"
    c.font = Font(bold=True, size=12, color="FFFFFF")
    c.fill = PatternFill("solid", fgColor=COLOR_HEAD)
    c.alignment = ALIGN_WRAP

    row = 3
    ws.cell(row=row, column=1, value=f"URL robots.txt: {result.get('robots_url', '—')}").font = FONT_BOLD
    row += 1

    kpis = [
        ("Direttive non supportate", result.get("unsupported_count", 0)),
        ("Typo Disallow", result.get("typos_count", 0)),
        ("Sitemap dichiarata", "Sì" if result.get("sitemap_declared") else "No"),
    ]
    for label, val in kpis:
        cell_style(ws, row, 1, label, bold=True)
        is_issue = (isinstance(val, int) and val > 0) or val == "No"
        cell_style(ws, row, 2, str(val), fill_color=COLOR_WARN if is_issue else COLOR_OK, center=True)
        row += 1

    # Findings
    row += 1
    for col, hdr in enumerate(["Tipo", "Severity", "Linea", "Direttiva", "Note"], 1):
        c = ws.cell(row=row, column=col, value=hdr)
        c.fill = PatternFill("solid", fgColor=COLOR_HEAD)
        c.font = FONT_HEAD
        c.alignment = ALIGN_CTR
        c.border = BORDER

    row += 1
    for f in result.get("findings", []):
        sev = f.get("severity", "INFO")
        cell_style(ws, row, 1, f["type"])
        cell_style(ws, row, 2, sev, fill_color=severity_color(sev), center=True, bold=True)
        cell_style(ws, row, 3, str(f["line"]), center=True)
        cell_style(ws, row, 4, f["directive"])
        cell_style(ws, row, 5, f["note"])
        ws.row_dimensions[row].height = 50
        row += 1

    if not result.get("findings"):
        ws.cell(row=row, column=1, value="Nessun problema rilevato. ✓").font = FONT_BOLD

    # Contenuto raw robots.txt
    row += 2
    ws.cell(row=row, column=1, value="Contenuto robots.txt:").font = FONT_BOLD
    row += 1
    if result.get("robots_content"):
        for line in result["robots_content"].splitlines()[:100]:
            ws.cell(row=row, column=1, value=line).font = Font(name="Courier New", size=9)
            ws.merge_cells(f"A{row}:E{row}")
            row += 1


def export_html(url, grouped_results, output_path):
    """Genera report HTML con layout ottimizzato, tabella pagine e tabella problemi."""
    timestamp = datetime.now().strftime("%d/%m/%Y %H:%M")

    score_map   = {"OK": 100, "WARN": 50, "FAIL": 0, "ERROR": 0}
    badge_color = {"OK": "#22c55e", "WARN": "#f59e0b", "FAIL": "#ef4444", "ERROR": "#6b7280"}
    badge_label = {"OK": "OK", "WARN": "WARN", "FAIL": "FAIL", "ERROR": "ERR"}

    # ── Per ogni URL raccogliamo tutti i risultati ──
    pages = []
    for pg in grouped_results:
        r1  = next((r for r in pg if "1."  in r.get("area", "")), {})
        r2  = next((r for r in pg if "2."  in r.get("area", "")), {})
        r3  = next((r for r in pg if "3."  in r.get("area", "")), {})
        r8  = next((r for r in pg if "8."  in r.get("area", "")), {})
        r9  = next((r for r in pg if "9."  in r.get("area", "")), {})
        r10 = next((r for r in pg if "10." in r.get("area", "")), {})
        pg_url = r1.get("url") or r2.get("url") or r3.get("url") or "—"
        o1  = r1.get("overall","ERROR");  o2  = r2.get("overall","ERROR")
        o3  = r3.get("overall","ERROR");  o8  = r8.get("overall","ERROR")
        o9  = r9.get("overall","ERROR");  o10 = r10.get("overall","ERROR")
        all_ov = [o1, o2, o3, o8, o9, o10]
        gs = round(sum(score_map.get(o, 0) for o in all_ov) / len(all_ov))
        pages.append({"url": pg_url,
                      "r1": r1, "r2": r2, "r3": r3, "r8": r8, "r9": r9, "r10": r10,
                      "o1": o1, "o2": o2, "o3": o3, "o8": o8, "o9": o9, "o10": o10,
                      "score": gs})

    # ── Global score (media di tutte le pagine) ──
    global_score = round(sum(p["score"] for p in pages) / len(pages)) if pages else 0
    gauge_color  = "#22c55e" if global_score >= 80 else ("#f59e0b" if global_score >= 50 else "#ef4444")

    # ── Tabella pagine analizzate ──
    pages_rows = ""
    for i, p in enumerate(pages, 1):
        def bc(o): return badge_color.get(o, "#6b7280")
        def bl(o): return badge_label.get(o, o)
        score_col = "#22c55e" if p["score"] >= 80 else ("#f59e0b" if p["score"] >= 50 else "#ef4444")
        pages_rows += f"""
        <tr>
          <td style="text-align:center;color:var(--muted);font-size:12px">{i}</td>
          <td style="font-family:monospace;font-size:12px;word-break:break-all">{p['url']}</td>
          <td style="text-align:center"><span class="badge" style="background:{bc(p['o1'])}">{bl(p['o1'])}</span></td>
          <td style="text-align:center"><span class="badge" style="background:{bc(p['o2'])}">{bl(p['o2'])}</span></td>
          <td style="text-align:center"><span class="badge" style="background:{bc(p['o3'])}">{bl(p['o3'])}</span></td>
          <td style="text-align:center"><span class="badge" style="background:{bc(p['o8'])}">{bl(p['o8'])}</span></td>
          <td style="text-align:center"><span class="badge" style="background:{bc(p['o9'])}">{bl(p['o9'])}</span></td>
          <td style="text-align:center"><span class="badge" style="background:{bc(p['o10'])}">{bl(p['o10'])}</span></td>
          <td style="text-align:center;font-weight:700;color:{score_col}">{p['score']}/100</td>
        </tr>"""

    # ── Tabella problemi per area ──
    problems = []
    for p in pages:
        u = p["url"]
        r1, r2, r3 = p["r1"], p["r2"], p["r3"]
        for b in r1.get("hidden_blocks", []):
            problems.append(("Visibilità", u, "Contenuto nascosto", "WARN",
                             f"{b['selector']} — {b['text_preview'][:80]}"))
        if r1.get("lazy_elements", 0):
            problems.append(("Visibilità", u, "Lazy-load / scroll", "WARN",
                             f"{r1['lazy_elements']} elementi con attributi lazy-load"))
        for pat in r1.get("js_patterns", []):
            problems.append(("Visibilità", u, "Pattern JS", "WARN", pat))
        if r2.get("h1_count", 1) == 0:
            problems.append(("Heading", u, "H1 assente", "FAIL", "Nessun H1 trovato nella pagina"))
        elif r2.get("h1_count", 1) > 1:
            problems.append(("Heading", u, "H1 multipli", "WARN", f"{r2['h1_count']} H1 trovati"))
        for jmp in r2.get("level_jumps", []):
            problems.append(("Heading", u, "Salto livello", "WARN", jmp))
        for orph in r2.get("orphan_blocks", []):
            problems.append(("Heading", u, "Blocco orfano", "WARN", orph[:100]))
        for f in r3.get("findings", []):
            if f.get("severity") in ("WARN", "FAIL"):
                problems.append(("Robots.txt", u, f.get("type","—"), f.get("severity","WARN"),
                                 f"L.{f.get('line','—')} {f.get('directive','')} — {f.get('note','')}"))

    problems_rows = ""
    for area, pu, tipo, sev, note in problems:
        sc = badge_color.get(sev, "#6b7280")
        problems_rows += f"""
        <tr>
          <td><span class="area-tag">{area}</span></td>
          <td style="font-family:monospace;font-size:11px;word-break:break-all">{pu}</td>
          <td style="font-size:13px">{tipo}</td>
          <td style="text-align:center"><span class="badge" style="background:{sc}">{sev}</span></td>
          <td style="font-size:12px;color:var(--muted)">{note}</td>
        </tr>"""
    if not problems_rows:
        problems_rows = '<tr><td colspan="5" style="text-align:center;color:#22c55e;padding:20px">✓ Nessun problema rilevato</td></tr>'

    # ── Sezione dettaglio prima URL ──
    first = pages[0] if pages else {}
    r1f   = first.get("r1",  {})
    r2f   = first.get("r2",  {})
    r3f   = first.get("r3",  {})
    r8f   = first.get("r8",  {})
    r9f   = first.get("r9",  {})
    r10f  = first.get("r10", {})

    # heading map
    heading_rows = ""
    for h in r2f.get("heading_map", [])[:50]:
        indent = "&nbsp;" * ((h["level"] - 1) * 4)
        col = {"1":"#6366f1","2":"#0ea5e9","3":"#14b8a6"}.get(str(h["level"]), "#94a3b8")
        heading_rows += f'<tr><td style="text-align:center;font-weight:700;color:{col}">H{h["level"]}</td><td>{indent}<span style="color:{col}">{h["text"]}</span></td></tr>\n'
    if not heading_rows:
        heading_rows = '<tr><td colspan="2" style="text-align:center;color:var(--muted)">—</td></tr>'

    # robots findings
    findings_rows = ""
    sev_icon = {"OK":"✓","WARN":"⚠","FAIL":"✗","INFO":"ℹ"}
    for f in r3f.get("findings", []):
        sev = f.get("severity","INFO")
        col = badge_color.get(sev, "#6b7280")
        findings_rows += f'<tr><td><span class="badge" style="background:{col}">{sev_icon.get(sev,"?")} {sev}</span></td><td style="font-family:monospace;font-size:12px">{f.get("directive","")} <em style="color:#94a3b8">L.{f.get("line","—")}</em></td><td style="font-size:13px">{f.get("note","")}</td></tr>\n'
    if not findings_rows:
        findings_rows = '<tr><td colspan="3" style="text-align:center;color:#22c55e;padding:16px">✓ Nessun problema rilevato</td></tr>'

    # hidden blocks
    hidden_rows = ""
    for b in r1f.get("hidden_blocks", [])[:20]:
        warn = "⚠ contiene heading" if b["has_heading"] else ""
        hidden_rows += f'<tr><td style="font-family:monospace;font-size:11px;color:#f59e0b">{b["selector"]}</td><td style="font-family:monospace">&lt;{b["tag"]}&gt;</td><td style="color:#f59e0b;font-size:12px">{warn}</td><td style="font-size:12px;color:#94a3b8">{b["text_preview"]}</td></tr>\n'
    if not hidden_rows:
        hidden_rows = '<tr><td colspan="4" style="text-align:center;color:#22c55e;padding:16px">✓ Nessun blocco nascosto</td></tr>'

    js_html = "".join(f'<code class="tag-pill">{p}</code>' for p in r1f.get("js_patterns",[])) or '<span style="color:#22c55e">Nessuno</span>'
    sum1 = r1f.get("summary","").replace("|","<br>")
    sum2 = r2f.get("summary","").replace("|","<br>")
    sum3 = r3f.get("summary","").replace("|","<br>")
    dom_delta = r1f.get("dom_delta","N/A")
    robots_raw = r3f.get("robots_content","Non disponibile")[:2000]
    robots_url_str = r3f.get("robots_url","—")

    # chart data (prima pagina)
    o1f, o2f, o3f = first.get("o1","ERROR"), first.get("o2","ERROR"), first.get("o3","ERROR")
    s1f, s2f, s3f = score_map.get(o1f,0), score_map.get(o2f,0), score_map.get(o3f,0)
    hidden_n = len(r1f.get("hidden_blocks",[]))
    lazy_n   = r1f.get("lazy_elements",0)
    js_n     = len(r1f.get("js_patterns",[]))
    h2c      = r2f.get("h2_count",0)
    h3c      = r2f.get("h3_count",0)
    jumps_n  = len(r2f.get("level_jumps",[]))
    orphans_n= len(r2f.get("orphan_blocks",[]))

    # ── Render cards prima pagina ──
    def card(area_label, overall, summary_text, accent):
        bc_ = badge_color.get(overall,"#6b7280")
        bl_ = badge_label.get(overall, overall)
        return f"""<div class="card" style="--accent:{accent}">
          <div class="card-label">{area_label}</div>
          <span class="badge" style="background:{bc_}">{bl_}</span>
          <p class="card-summary">{summary_text[:200]}</p>
        </div>"""

    o8f  = first.get("o8",  "ERROR")
    o9f  = first.get("o9",  "ERROR")
    o10f = first.get("o10", "ERROR")
    sum8  = r8f.get("summary",  "").replace("|","<br>")
    sum9  = r9f.get("summary",  "").replace("|","<br>")
    sum10 = r10f.get("summary", "").replace("|","<br>")
    cards_html = (
        card("Area 1 — Visibilità Contenuti", o1f,  sum1,  badge_color.get(o1f, "#6b7280")) +
        card("Area 2 — Struttura Heading",    o2f,  sum2,  badge_color.get(o2f, "#6b7280")) +
        card("Area 3 — Robots.txt",           o3f,  sum3,  badge_color.get(o3f, "#6b7280")) +
        card("Area 8 — E-E-A-T",              o8f,  sum8,  badge_color.get(o8f, "#6b7280")) +
        card("Area 9 — Performance",          o9f,  sum9,  badge_color.get(o9f, "#6b7280")) +
        card("Area 10 — Topical Authority",   o10f, sum10, badge_color.get(o10f,"#6b7280"))
    )

    first_url_label = first.get("url", url)

    html = f"""<!DOCTYPE html>
<html lang="it">
<head>
<meta charset="UTF-8">
<meta name="viewport" content="width=device-width, initial-scale=1.0">
<title>SEO Audit — {url}</title>
<script src="https://cdnjs.cloudflare.com/ajax/libs/Chart.js/4.4.1/chart.umd.min.js"></script>
<style>
  @import url('https://fonts.googleapis.com/css2?family=Inter:wght@400;500;600;700&family=JetBrains+Mono:wght@400;500&display=swap');

  *, *::before, *::after {{ box-sizing: border-box; margin: 0; padding: 0; }}

  :root {{
    --bg:       #0f1117;
    --surface:  #161b27;
    --surface2: #1d2333;
    --border:   rgba(255,255,255,0.08);
    --text:     #e2e8f0;
    --muted:    #64748b;
    --ok:       #22c55e;
    --warn:     #f59e0b;
    --fail:     #ef4444;
    --accent:   #6366f1;
    --radius:   10px;
  }}

  html {{ scroll-behavior: smooth; }}
  body {{ background: var(--bg); color: var(--text); font-family: 'Inter', sans-serif; font-size: 14px; line-height: 1.6; }}

  /* topbar */
  .topbar {{
    position: sticky; top: 0; z-index: 100;
    background: rgba(15,17,23,0.9); backdrop-filter: blur(10px);
    border-bottom: 1px solid var(--border);
    display: flex; align-items: center; gap: 20px;
    padding: 0 36px; height: 52px;
  }}
  .topbar-logo {{ font-weight: 700; font-size: 13px; color: var(--accent); letter-spacing: 0.06em; text-transform: uppercase; }}
  .topbar-url  {{ font-family: 'JetBrains Mono', monospace; font-size: 11px; color: var(--muted); flex: 1; overflow: hidden; text-overflow: ellipsis; white-space: nowrap; }}
  .topbar-ts   {{ font-size: 11px; color: var(--muted); white-space: nowrap; }}

  /* nav */
  .nav {{ display: flex; gap: 6px; padding: 24px 36px 0; flex-wrap: wrap; }}
  .nav a {{
    font-size: 12px; font-weight: 600; padding: 5px 14px; border-radius: 6px;
    border: 1px solid var(--border); color: var(--muted);
    text-decoration: none; transition: all .15s;
  }}
  .nav a:hover {{ border-color: var(--accent); color: var(--accent); background: rgba(99,102,241,.06); }}

  /* layout */
  .wrap {{ max-width: 1240px; margin: 0 auto; padding: 36px; }}
  .section {{ margin-bottom: 48px; }}
  .section-title {{
    font-size: 10px; font-weight: 700; letter-spacing: 0.14em; text-transform: uppercase;
    color: var(--muted); margin-bottom: 20px;
    display: flex; align-items: center; gap: 10px;
  }}
  .section-title::after {{ content:''; flex:1; height:1px; background:var(--border); }}

  /* score global */
  .hero {{ display: grid; grid-template-columns: 180px 1fr 1fr 1fr; gap: 16px; margin-bottom: 40px; align-items: start; }}
  @media(max-width:880px) {{ .hero {{ grid-template-columns: 1fr 1fr; }} }}

  .score-globe {{
    background: var(--surface); border: 1px solid var(--border); border-radius: var(--radius);
    padding: 28px 20px; display: flex; flex-direction: column; align-items: center; gap: 8px; text-align: center;
  }}
  .score-globe-num {{ font-size: 56px; font-weight: 700; line-height: 1; color: {gauge_color}; letter-spacing: -2px; }}
  .score-globe-lbl {{ font-size: 10px; font-weight: 600; letter-spacing: 0.1em; text-transform: uppercase; color: var(--muted); }}
  .score-bar {{ width: 100%; height: 6px; background: var(--border); border-radius: 999px; overflow: hidden; margin-top: 4px; }}
  .score-bar-fill {{ height: 100%; border-radius: 999px; background: {gauge_color}; width: {global_score}%; }}

  /* cards */
  .card {{
    background: var(--surface); border: 1px solid var(--border); border-radius: var(--radius);
    padding: 22px 24px; display: flex; flex-direction: column; gap: 10px;
    border-top: 3px solid var(--accent);
  }}
  .card-label {{ font-size: 11px; font-weight: 600; color: var(--muted); text-transform: uppercase; letter-spacing: 0.07em; }}
  .card-summary {{ font-size: 12px; color: var(--muted); line-height: 1.6; }}

  /* badge */
  .badge {{
    display: inline-flex; align-items: center;
    padding: 2px 10px; border-radius: 5px;
    font-size: 11px; font-weight: 700; letter-spacing: 0.05em; color: #fff;
  }}
  .area-tag {{
    display: inline-block; padding: 2px 8px; border-radius: 4px;
    font-size: 11px; font-weight: 600; background: var(--surface2); color: var(--muted);
  }}

  /* panel */
  .panel {{
    background: var(--surface); border: 1px solid var(--border);
    border-radius: var(--radius); padding: 28px; margin-bottom: 16px;
  }}
  .panel h3 {{ font-size: 15px; font-weight: 700; margin-bottom: 6px; }}
  .panel-meta {{ font-size: 13px; color: var(--muted); margin-bottom: 20px; line-height: 1.7; }}

  /* charts */
  .chart-grid {{ display: grid; grid-template-columns: 1fr 1fr; gap: 16px; margin-bottom: 40px; }}
  @media(max-width:768px) {{ .chart-grid {{ grid-template-columns: 1fr; }} }}
  .chart-card {{ background: var(--surface); border: 1px solid var(--border); border-radius: var(--radius); padding: 24px; }}
  .chart-card h3 {{ font-size: 13px; font-weight: 700; margin-bottom: 16px; color: var(--text); }}
  .chart-wrap {{ position: relative; height: 240px; }}

  /* kpi */
  .kpi-row {{ display: flex; gap: 12px; flex-wrap: wrap; margin-bottom: 20px; }}
  .kpi {{ background: var(--surface2); border: 1px solid var(--border); border-radius: 8px; padding: 14px 18px; flex: 1; min-width: 110px; }}
  .kpi-val {{ font-size: 26px; font-weight: 700; line-height: 1; }}
  .kpi-lbl {{ font-size: 11px; color: var(--muted); margin-top: 4px; }}

  /* tables */
  .tbl-wrap {{ overflow-x: auto; }}
  table {{ width: 100%; border-collapse: collapse; font-size: 13px; }}
  th {{
    font-size: 11px; font-weight: 700; letter-spacing: 0.07em; text-transform: uppercase;
    color: var(--muted); text-align: left; padding: 10px 14px;
    border-bottom: 1px solid var(--border); white-space: nowrap;
  }}
  td {{ padding: 11px 14px; border-bottom: 1px solid var(--border); vertical-align: top; }}
  tr:last-child td {{ border-bottom: none; }}
  tr:hover td {{ background: rgba(255,255,255,0.015); }}

  /* code */
  pre {{
    background: #0d1117; border: 1px solid var(--border); border-radius: 8px;
    padding: 16px; font-family: 'JetBrains Mono', monospace; font-size: 12px;
    color: #7dd3fc; overflow-x: auto; max-height: 280px; overflow-y: auto; line-height: 1.8;
  }}
  .tag-pill {{
    display: inline-block; background: rgba(245,158,11,.1); border: 1px solid rgba(245,158,11,.2);
    color: var(--warn); font-family: 'JetBrains Mono', monospace; font-size: 11px;
    padding: 2px 8px; border-radius: 4px; margin: 2px;
  }}

  .footer {{ margin-top: 64px; padding: 20px 36px; border-top: 1px solid var(--border); text-align: center; font-size: 11px; color: var(--muted); }}
</style>
</head>
<body>

<div class="topbar">
  <span class="topbar-logo">SEO AUDIT</span>
  <span class="topbar-url">{url}</span>
  <span class="topbar-ts">{timestamp}</span>
</div>

<div class="nav">
  <a href="#overview">Overview</a>
  <a href="#pages">Pagine</a>
  <a href="#problems">Problemi</a>
  <a href="#charts">Grafici</a>
  <a href="#area1">Visibilità</a>
  <a href="#area2">Heading</a>
  <a href="#area3">Robots.txt</a>
  <a href="#area8">E-E-A-T</a>
  <a href="#area9">Performance</a>
  <a href="#area10">Topical</a>
</div>

<div class="wrap">

  <!-- ── OVERVIEW ── -->
  <div id="overview" class="section" style="padding-top:36px">
    <div class="section-title">Panoramica — {first_url_label}</div>
    <div class="hero">
      <div class="score-globe">
        <div class="score-globe-lbl">Score globale</div>
        <div class="score-globe-num">{global_score}</div>
        <div class="score-bar"><div class="score-bar-fill"></div></div>
        <div class="score-globe-lbl" style="font-size:9px">su 100 punti</div>
      </div>
      {cards_html}
    </div>
  </div>

  <!-- ── PAGINE ANALIZZATE ── -->
  <div id="pages" class="section">
    <div class="section-title">Pagine Analizzate — {len(pages)} URL</div>
    <div class="panel">
      <div class="tbl-wrap">
        <table>
          <thead>
            <tr>
              <th>#</th>
              <th>URL</th>
              <th>Area 1 — Visibilità</th>
              <th>Area 2 — Heading</th>
              <th>Area 3 — Robots.txt</th>
              <th>Score</th>
            </tr>
          </thead>
          <tbody>
            {pages_rows}
          </tbody>
        </table>
      </div>
    </div>
  </div>

  <!-- ── PROBLEMI PER AREA ── -->
  <div id="problems" class="section">
    <div class="section-title">Problemi Rilevati — {len(problems)} finding{'s' if len(problems) != 1 else ''}</div>
    <div class="panel">
      <div class="tbl-wrap">
        <table>
          <thead>
            <tr>
              <th>Area</th>
              <th>URL</th>
              <th>Tipo</th>
              <th>Severity</th>
              <th>Dettaglio</th>
            </tr>
          </thead>
          <tbody>
            {problems_rows}
          </tbody>
        </table>
      </div>
    </div>
  </div>

  <!-- ── GRAFICI ── -->
  <div id="charts" class="section">
    <div class="section-title">Grafici — prima URL analizzata</div>
    <div class="chart-grid">

      <div class="chart-card">
        <h3>Score per area</h3>
        <div class="chart-wrap"><canvas id="barScore"></canvas></div>
      </div>

      <div class="chart-card">
        <h3>Distribuzione esiti</h3>
        <div class="chart-wrap"><canvas id="pieChart"></canvas></div>
      </div>

      <div class="chart-card">
        <h3>Area 1 — Metriche visibilità</h3>
        <div class="chart-wrap"><canvas id="barVisibility"></canvas></div>
      </div>

      <div class="chart-card">
        <h3>Area 2 — Metriche heading</h3>
        <div class="chart-wrap"><canvas id="barHeading"></canvas></div>
      </div>

    </div>
  </div>

  <!-- ── AREA 1 ── -->
  <div id="area1" class="section">
    <div class="section-title">Area 1 — Rendering & Visibilità Contenuti</div>
    <div class="panel">
      <h3>Visibilità Contenuti <span class="badge" style="background:{badge_color.get(o1f,'#6b7280')};margin-left:8px">{badge_label.get(o1f,o1f)}</span></h3>
      <p class="panel-meta">{sum1}</p>

      <div class="kpi-row">
        <div class="kpi"><div class="kpi-val" style="color:{'#ef4444' if hidden_n>0 else '#22c55e'}">{hidden_n}</div><div class="kpi-lbl">Blocchi nascosti</div></div>
        <div class="kpi"><div class="kpi-val" style="color:{'#f59e0b' if lazy_n>0 else '#22c55e'}">{lazy_n}</div><div class="kpi-lbl">Elementi lazy-load</div></div>
        <div class="kpi"><div class="kpi-val" style="color:{'#f59e0b' if js_n>0 else '#22c55e'}">{js_n}</div><div class="kpi-lbl">Pattern JS</div></div>
      </div>

      <p style="font-size:12px;color:var(--muted);margin-bottom:12px">Delta DOM statico / renderizzato: <strong style="color:var(--text)">{dom_delta}</strong></p>

      <p style="font-size:12px;color:var(--muted);margin-bottom:8px">Pattern JS rilevati:</p>
      <div style="margin-bottom:20px">{js_html}</div>

      <div class="tbl-wrap">
        <table>
          <thead><tr><th>Selettore</th><th>Tag</th><th>Note</th><th>Anteprima testo</th></tr></thead>
          <tbody>{hidden_rows}</tbody>
        </table>
      </div>
    </div>
  </div>

  <!-- ── AREA 2 ── -->
  <div id="area2" class="section">
    <div class="section-title">Area 2 — Struttura Heading H2/H3</div>
    <div class="panel">
      <h3>Struttura Heading <span class="badge" style="background:{badge_color.get(o2f,'#6b7280')};margin-left:8px">{badge_label.get(o2f,o2f)}</span></h3>
      <p class="panel-meta">{sum2}</p>

      <div class="kpi-row">
        <div class="kpi"><div class="kpi-val">{r2f.get('h1_count',0)}</div><div class="kpi-lbl">H1</div></div>
        <div class="kpi"><div class="kpi-val">{h2c}</div><div class="kpi-lbl">H2</div></div>
        <div class="kpi"><div class="kpi-val">{h3c}</div><div class="kpi-lbl">H3</div></div>
        <div class="kpi"><div class="kpi-val" style="color:{'#f59e0b' if jumps_n>0 else '#22c55e'}">{jumps_n}</div><div class="kpi-lbl">Salti livello</div></div>
        <div class="kpi"><div class="kpi-val" style="color:{'#f59e0b' if orphans_n>0 else '#22c55e'}">{orphans_n}</div><div class="kpi-lbl">Blocchi orfani</div></div>
      </div>

      <div class="tbl-wrap">
        <table>
          <thead><tr><th>Livello</th><th>Testo</th></tr></thead>
          <tbody>{heading_rows}</tbody>
        </table>
      </div>
    </div>
  </div>

  <!-- ── AREA 3 ── -->
  <div id="area3" class="section">
    <div class="section-title">Area 3 — Robots.txt</div>
    <div class="panel">
      <h3>Robots.txt <span class="badge" style="background:{badge_color.get(o3f,'#6b7280')};margin-left:8px">{badge_label.get(o3f,o3f)}</span></h3>
      <p class="panel-meta">{sum3}</p>
      <p style="font-size:12px;color:var(--muted);margin-bottom:16px">File: <code style="font-family:monospace">{robots_url_str}</code></p>

      <div class="tbl-wrap" style="margin-bottom:20px">
        <table>
          <thead><tr><th>Severity</th><th>Direttiva</th><th>Nota</th></tr></thead>
          <tbody>{findings_rows}</tbody>
        </table>
      </div>

      <p style="font-size:11px;color:var(--muted);margin-bottom:8px">Contenuto robots.txt:</p>
      <pre>{robots_raw}</pre>
    </div>
  </div>



  <!-- ── AREA 8 — E-E-A-T ────────────────────────────────────────────── -->
  <div id="area8" class="section">
    <div class="section-title">Area 8 — E-E-A-T Signals</div>
    <div class="panel">
      <h3>Experience · Expertise · Authoritativeness · Trustworthiness
        <span class="badge" style="background:{badge_color.get(o8f,'#6b7280')};margin-left:8px">{badge_label.get(o8f,o8f)}</span>
      </h3>
      <p class="panel-meta">{sum8}</p>

      <div class="kpi-row">
        <div class="kpi"><div class="kpi-val" style="color:{'#22c55e' if r8f.get('author_found') else '#ef4444'}">{len(r8f.get('author_found',[]))}</div><div class="kpi-lbl">Autori identificati</div></div>
        <div class="kpi"><div class="kpi-val" style="color:{'#22c55e' if r8f.get('bio_found') else '#f59e0b'}">{'Sì' if r8f.get('bio_found') else 'No'}</div><div class="kpi-lbl">Bio autore</div></div>
        <div class="kpi"><div class="kpi-val" style="color:{'#22c55e' if r8f.get('date_pub') else '#f59e0b'}">{'Sì' if r8f.get('date_pub') else 'No'}</div><div class="kpi-lbl">Data pubblicazione</div></div>
        <div class="kpi"><div class="kpi-val" style="color:{'#22c55e' if r8f.get('org_schema') else '#f59e0b'}">{'Sì' if r8f.get('org_schema') else 'No'}</div><div class="kpi-lbl">Schema Organization</div></div>
        <div class="kpi"><div class="kpi-val" style="color:{'#22c55e' if r8f.get('auth_links') else '#f59e0b'}">{len(r8f.get('auth_links',[]))}</div><div class="kpi-lbl">Link fonti autorevoli</div></div>
        <div class="kpi"><div class="kpi-val">{r8f.get('word_count',0)}</div><div class="kpi-lbl">Parole contenuto</div></div>
      </div>

      {'<div style="background:rgba(239,68,68,.08);border:1px solid rgba(239,68,68,.2);border-radius:8px;padding:14px 18px;margin-bottom:16px;font-size:13px;color:#991b1b"><strong>⚠ Pagina YMYL rilevata</strong> — E-E-A-T è critico: motori di ricerca e sistemi AI applicano standard più severi per contenuti su finanza, salute e diritto.</div>' if r8f.get('is_ymyl') else ''}

      <div style="display:grid;grid-template-columns:1fr 1fr;gap:20px">
        <div>
          <p style="font-size:12px;font-weight:700;text-transform:uppercase;letter-spacing:.06em;color:var(--muted);margin-bottom:10px">✓ Segnali positivi</p>
          {''.join(f'<div style="font-size:13px;padding:7px 0;border-bottom:1px solid var(--border);color:var(--text)">{s}</div>' for s in r8f.get("signals",[])) or '<p style="font-size:13px;color:var(--muted)">Nessun segnale positivo rilevato</p>'}
        </div>
        <div>
          <p style="font-size:12px;font-weight:700;text-transform:uppercase;letter-spacing:.06em;color:var(--muted);margin-bottom:10px">⚠ Da migliorare</p>
          {''.join(f'<div style="font-size:13px;padding:7px 0;border-bottom:1px solid var(--border);color:#b45309">{m}</div>' for m in r8f.get("missing",[])) or '<p style="font-size:13px;color:#22c55e">Nessuna criticità rilevata</p>'}
        </div>
      </div>

      {('<div style="margin-top:20px"><p style="font-size:12px;font-weight:700;color:var(--muted);margin-bottom:8px">Link verso fonti autorevoli:</p><div class="tbl-wrap"><table><thead><tr><th>URL</th><th>Testo</th></tr></thead><tbody>' + "".join(f'<tr><td style="font-family:monospace;font-size:12px;word-break:break-all">{l["url"][:80]}</td><td style="font-size:13px">{l["text"]}</td></tr>' for l in r8f.get("auth_links",[])) + "</tbody></table></div></div>") if r8f.get("auth_links") else ""}
    </div>
  </div>

  <!-- ── AREA 9 — PERFORMANCE ─────────────────────────────────────────── -->
  <div id="area9" class="section">
    <div class="section-title">Area 9 — Performance Signals</div>
    <div class="panel">
      <h3>Segnali di performance da DOM statico
        <span class="badge" style="background:{badge_color.get(o9f,'#6b7280')};margin-left:8px">{badge_label.get(o9f,o9f)}</span>
      </h3>
      <p class="panel-meta">{sum9}</p>
      <p style="font-size:12px;color:var(--muted);background:rgba(99,102,241,.06);border:1px solid rgba(99,102,241,.15);border-radius:6px;padding:10px 14px;margin-bottom:20px">
        ℹ Questi sono segnali rilevabili dal DOM senza browser reale. Per dati LCP/INP/CLS verificati usare
        <a href="https://pagespeed.web.dev/analysis?url={first.get('url','')}" target="_blank" style="color:var(--accent)">PageSpeed Insights →</a>
      </p>

      <div class="kpi-row">
        <div class="kpi"><div class="kpi-val">{r9f.get('images_total',0)}</div><div class="kpi-lbl">Immagini totali</div></div>
        <div class="kpi"><div class="kpi-val" style="color:{'#ef4444' if r9f.get('imgs_no_size',0)>0 else '#22c55e'}">{r9f.get('imgs_no_size',0)}</div><div class="kpi-lbl">Senza dimensioni (CLS)</div></div>
        <div class="kpi"><div class="kpi-val" style="color:{'#f59e0b' if r9f.get('imgs_no_alt',0)>0 else '#22c55e'}">{r9f.get('imgs_no_alt',0)}</div><div class="kpi-lbl">Senza alt text</div></div>
        <div class="kpi"><div class="kpi-val" style="color:{'#f59e0b' if r9f.get('blocking_scripts',0)>0 else '#22c55e'}">{r9f.get('blocking_scripts',0)}</div><div class="kpi-lbl">Script bloccanti</div></div>
        <div class="kpi"><div class="kpi-val">{r9f.get('ext_scripts',0)}</div><div class="kpi-lbl">Script esterni</div></div>
        <div class="kpi"><div class="kpi-val">{r9f.get('preconnect_count',0)}</div><div class="kpi-lbl">Preconnect/DNS hint</div></div>
      </div>

      {'<div style="margin-bottom:16px"><p style="font-size:12px;font-weight:700;color:var(--muted);margin-bottom:8px">Framework JS rilevati:</p>' + "".join(f'<span class="tag-pill">{fw}</span>' for fw in r9f.get("frameworks",[])) + "</div>" if r9f.get("frameworks") else ""}

      <div class="tbl-wrap" style="margin-bottom:20px">
        <table>
          <thead><tr><th>Problema</th><th>Severity</th><th>Impatto metrica</th><th>Raccomandazione</th></tr></thead>
          <tbody>
            {''.join(f'<tr><td style="font-size:13px;font-weight:600">{f["tipo"]}</td><td style="text-align:center"><span class="badge" style="background:{badge_color.get(f["severity"],"#6b7280")}">{f["severity"]}</span></td><td style="font-size:12px;font-family:monospace;color:var(--accent)">{f["impatto"]}</td><td style="font-size:12px;color:var(--muted)">{f["nota"]}</td></tr>' for f in r9f.get("findings",[]))}
            {'<tr><td colspan="4" style="text-align:center;color:#22c55e;padding:16px">✓ Nessun problema rilevato</td></tr>' if not r9f.get("findings") else ""}
          </tbody>
        </table>
      </div>

      <p style="font-size:12px;font-weight:700;color:var(--muted);margin-bottom:10px">✓ Elementi ottimizzati:</p>
      {''.join(f'<div style="font-size:13px;padding:6px 0;border-bottom:1px solid var(--border);color:#22c55e">✓ {p}</div>' for p in r9f.get("positives",[])) or '<p style="font-size:13px;color:var(--muted)">—</p>'}
    </div>
  </div>

  <!-- ── AREA 10 — TOPICAL AUTHORITY ─────────────────────────────────── -->
  <div id="area10" class="section">
    <div class="section-title">Area 10 — Autorevolezza Topica</div>
    <div class="panel">
      <h3>Meta tag · Struttura contenuto · Internal linking · Citazioni
        <span class="badge" style="background:{badge_color.get(o10f,'#6b7280')};margin-left:8px">{badge_label.get(o10f,o10f)}</span>
      </h3>
      <p class="panel-meta">{sum10}</p>

      <div class="kpi-row">
        <div class="kpi">
          <div class="kpi-val" style="color:{'#22c55e' if 50<=r10f.get('title_len',0)<=65 else '#f59e0b' if r10f.get('title_len',0)>0 else '#ef4444'}">{r10f.get('title_len',0)}</div>
          <div class="kpi-lbl">Car. Title (opt. 50-60)</div>
        </div>
        <div class="kpi">
          <div class="kpi-val" style="color:{'#22c55e' if 120<=r10f.get('desc_len',0)<=165 else '#f59e0b' if r10f.get('desc_len',0)>0 else '#ef4444'}">{r10f.get('desc_len',0)}</div>
          <div class="kpi-lbl">Car. Meta desc (opt. 120-160)</div>
        </div>
        <div class="kpi"><div class="kpi-val" style="color:{'#22c55e' if r10f.get('has_canonical') else '#f59e0b'}">{'Sì' if r10f.get('has_canonical') else 'No'}</div><div class="kpi-lbl">Canonical tag</div></div>
        <div class="kpi"><div class="kpi-val" style="color:{'#22c55e' if r10f.get('has_breadcrumb') else '#f59e0b'}">{'Sì' if r10f.get('has_breadcrumb') else 'No'}</div><div class="kpi-lbl">Breadcrumb</div></div>
        <div class="kpi"><div class="kpi-val">{r10f.get('internal_links_count',0)}</div><div class="kpi-lbl">Link interni</div></div>
        <div class="kpi"><div class="kpi-val">{r10f.get('word_count',0)}</div><div class="kpi-lbl">Parole</div></div>
      </div>

      {('<div style="background:var(--surface2);border-radius:8px;padding:14px 18px;margin-bottom:20px"><p style="font-size:12px;font-weight:700;color:var(--muted);margin-bottom:4px">Title tag:</p><p style="font-family:monospace;font-size:13px">' + r10f.get("title_text","—")[:80] + '</p></div>') if r10f.get("title_text") else ""}
      {('<div style="background:var(--surface2);border-radius:8px;padding:14px 18px;margin-bottom:20px"><p style="font-size:12px;font-weight:700;color:var(--muted);margin-bottom:4px">Meta description:</p><p style="font-size:13px;color:var(--muted)">' + r10f.get("desc_text","—")[:200] + '</p></div>') if r10f.get("desc_text") else ""}

      <div style="display:grid;grid-template-columns:1fr 1fr;gap:20px;margin-bottom:20px">
        <div>
          <p style="font-size:12px;font-weight:700;text-transform:uppercase;letter-spacing:.06em;color:var(--muted);margin-bottom:10px">✓ Punti di forza</p>
          {''.join(f'<div style="font-size:13px;padding:7px 0;border-bottom:1px solid var(--border)">{s}</div>' for s in r10f.get("signals",[])) or '<p style="font-size:13px;color:var(--muted)">—</p>'}
        </div>
        <div>
          <p style="font-size:12px;font-weight:700;text-transform:uppercase;letter-spacing:.06em;color:var(--muted);margin-bottom:10px">💡 Opportunità</p>
          {''.join(f'<div style="font-size:13px;padding:7px 0;border-bottom:1px solid var(--border);color:#b45309">{o}</div>' for o in r10f.get("opportunities",[])) or '<p style="font-size:13px;color:#22c55e">Nessuna opportunità critica</p>'}
        </div>
      </div>

      {('<div><p style="font-size:12px;font-weight:700;color:var(--muted);margin-bottom:10px">Sezioni H2 della pagina:</p>' + "".join(f'<div style="font-size:13px;padding:6px 12px;border-left:3px solid var(--accent);margin-bottom:6px;background:var(--surface2);border-radius:0 6px 6px 0">{h}</div>' for h in r10f.get("h2_list",[])) + "</div>") if r10f.get("h2_list") else ""}
    </div>
  </div>

</div><!-- /wrap -->

<div class="footer">SEO Audit Script — generato il {timestamp}</div>

<script>
const COLORS = {{
  ok: '#22c55e', warn: '#f59e0b', fail: '#ef4444',
  accent: '#6366f1', accent2: '#0ea5e9', accent3: '#14b8a6',
  muted: '#334155', text: '#e2e8f0', grid: 'rgba(255,255,255,0.06)'
}};
const chartDefaults = {{
  plugins: {{ legend: {{ labels: {{ color: COLORS.text, font: {{ size: 12 }} }} }} }},
  scales: {{
    x: {{ ticks: {{ color: COLORS.muted }}, grid: {{ color: COLORS.grid }} }},
    y: {{ ticks: {{ color: COLORS.muted }}, grid: {{ color: COLORS.grid }} }}
  }}
}};

// Bar: score per area
new Chart(document.getElementById('barScore'), {{
  type: 'bar',
  data: {{
    labels: ['Visibilità', 'Heading', 'Robots.txt'],
    datasets: [{{
      data: [{s1f}, {s2f}, {s3f}],
      backgroundColor: ['{badge_color.get(o1f,"#6b7280")}', '{badge_color.get(o2f,"#6b7280")}', '{badge_color.get(o3f,"#6b7280")}'],
      borderRadius: 6, borderSkipped: false
    }}]
  }},
  options: {{
    ...chartDefaults,
    plugins: {{ legend: {{ display: false }} }},
    scales: {{
      x: {{ ticks: {{ color: COLORS.muted }}, grid: {{ color: COLORS.grid }} }},
      y: {{ min: 0, max: 100, ticks: {{ color: COLORS.muted }}, grid: {{ color: COLORS.grid }} }}
    }}
  }}
}});

// Pie: distribuzione esiti
new Chart(document.getElementById('pieChart'), {{
  type: 'doughnut',
  data: {{
    labels: ['Visibilità', 'Heading', 'Robots.txt'],
    datasets: [{{
      data: [1, 1, 1],
      backgroundColor: ['{badge_color.get(o1f,"#6b7280")}', '{badge_color.get(o2f,"#6b7280")}', '{badge_color.get(o3f,"#6b7280")}'],
      borderWidth: 2, borderColor: '#0f1117'
    }}]
  }},
  options: {{
    plugins: {{ legend: {{ position: 'bottom', labels: {{ color: COLORS.text, font: {{ size: 12 }} }} }} }},
    cutout: '60%'
  }}
}});

// Bar: visibilità
new Chart(document.getElementById('barVisibility'), {{
  type: 'bar',
  data: {{
    labels: ['Blocchi nascosti', 'Lazy-load', 'Pattern JS'],
    datasets: [{{
      data: [{hidden_n}, {lazy_n}, {js_n}],
      backgroundColor: [COLORS.fail, COLORS.warn, COLORS.warn],
      borderRadius: 6, borderSkipped: false
    }}]
  }},
  options: {{
    ...chartDefaults,
    plugins: {{ legend: {{ display: false }} }},
    scales: {{
      x: {{ ticks: {{ color: COLORS.muted }}, grid: {{ color: COLORS.grid }} }},
      y: {{ ticks: {{ color: COLORS.muted, stepSize: 1 }}, grid: {{ color: COLORS.grid }} }}
    }}
  }}
}});

// Bar: heading
new Chart(document.getElementById('barHeading'), {{
  type: 'bar',
  data: {{
    labels: ['H2', 'H3', 'Salti liv.', 'Orfani'],
    datasets: [{{
      data: [{h2c}, {h3c}, {jumps_n}, {orphans_n}],
      backgroundColor: [COLORS.accent, COLORS.accent2, COLORS.warn, COLORS.warn],
      borderRadius: 6, borderSkipped: false
    }}]
  }},
  options: {{
    ...chartDefaults,
    plugins: {{ legend: {{ display: false }} }},
    scales: {{
      x: {{ ticks: {{ color: COLORS.muted }}, grid: {{ color: COLORS.grid }} }},
      y: {{ ticks: {{ color: COLORS.muted, stepSize: 1 }}, grid: {{ color: COLORS.grid }} }}
    }}
  }}
}});
</script>
</body>
</html>"""

    with open(output_path, "w", encoding="utf-8") as f:
        f.write(html)
    print(f"✅ Report HTML salvato: {output_path}")


def write_sintesi_sheet(wb, url, results, timestamp):
    """Foglio Sintesi: dettaglio numerico di tutte le metriche per URL analizzata."""
    ws = wb.create_sheet("📊 Sintesi", 1)

    # Larghezze colonne
    col_widths = [40, 18, 14, 16, 14, 16, 14, 14, 14, 16, 14, 14, 55]
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    # ── Intestazione ──
    ws.merge_cells("A1:M1")
    c = ws["A1"]
    c.value = f"Sintesi analisi — {url}"
    c.font = Font(bold=True, size=13, color="FFFFFF")
    c.fill = PatternFill("solid", fgColor=COLOR_HEAD)
    c.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    ws.row_dimensions[1].height = 24

    ws.merge_cells("A2:M2")
    c = ws["A2"]
    c.value = f"Generato il: {timestamp}"
    c.font = Font(size=9, color="888888")
    c.alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[2].height = 16

    # ── Headers tabella (riga 4) ──
    headers = [
        "URL",
        # Area 1
        "A1 — Esito", "Blocchi nascosti", "Lazy-load", "Pattern JS",
        # Area 2
        "A2 — Esito", "H1", "H2", "H3", "Salti livello", "Blocchi orfani",
        # Area 3
        "A3 — Esito", "Note / Findings robots.txt",
    ]

    row = 4
    # Gruppo header area 1
    ws.merge_cells("B4:E4")
    c = ws["B4"]
    c.value = "Area 1 — Visibilità Contenuti"
    c.fill = PatternFill("solid", fgColor="1A3A5C")
    c.font = Font(bold=True, size=10, color="FFFFFF")
    c.alignment = ALIGN_CTR
    c.border = BORDER

    ws.merge_cells("F4:K4")
    c = ws["F4"]
    c.value = "Area 2 — Struttura Heading"
    c.fill = PatternFill("solid", fgColor="1A3A5C")
    c.font = Font(bold=True, size=10, color="FFFFFF")
    c.alignment = ALIGN_CTR
    c.border = BORDER

    ws.merge_cells("L4:M4")
    c = ws["L4"]
    c.value = "Area 3 — Robots.txt"
    c.fill = PatternFill("solid", fgColor="1A3A5C")
    c.font = Font(bold=True, size=10, color="FFFFFF")
    c.alignment = ALIGN_CTR
    c.border = BORDER

    # Celle A4 (URL label)
    c = ws["A4"]
    c.value = "URL"
    c.fill = PatternFill("solid", fgColor=COLOR_HEAD)
    c.font = FONT_HEAD
    c.alignment = ALIGN_CTR
    c.border = BORDER

    row = 5
    sub_headers = [
        "URL analizzata",
        "Esito", "Blocchi nascosti", "Lazy-load", "Pattern JS",
        "Esito", "H1", "H2", "H3", "Salti livello", "Blocchi orfani",
        "Esito", "Findings (sintesi)",
    ]
    for col, hdr in enumerate(sub_headers, 1):
        c = ws.cell(row=row, column=col, value=hdr)
        c.fill = PatternFill("solid", fgColor="1E3A52")
        c.font = Font(bold=True, size=9, color="CBD5E1")
        c.alignment = ALIGN_CTR
        c.border = BORDER
    ws.row_dimensions[row].height = 28

    # ── Dati ──
    r1 = next((r for r in results if "1." in r.get("area", "")), {})
    r2 = next((r for r in results if "2." in r.get("area", "")), {})
    r3 = next((r for r in results if "3." in r.get("area", "")), {})

    def esito_fill(o):
        return {"OK": COLOR_OK, "WARN": COLOR_WARN, "FAIL": COLOR_FAIL}.get(o, "DDDDDD")

    o1 = r1.get("overall", "N/A")
    o2 = r2.get("overall", "N/A")
    o3 = r3.get("overall", "N/A")

    findings_short = "; ".join(
        f"[{f.get('severity','?')}] {f.get('directive','')} — {f.get('note','')[:60]}"
        for f in r3.get("findings", [])
    ) or "Nessun problema"

    data_row = [
        url,
        o1,
        len(r1.get("hidden_blocks", [])),
        r1.get("lazy_elements", 0),
        len(r1.get("js_patterns", [])),
        o2,
        r2.get("h1_count", 0),
        r2.get("h2_count", 0),
        r2.get("h3_count", 0),
        len(r2.get("level_jumps", [])),
        len(r2.get("orphan_blocks", [])),
        o3,
        findings_short,
    ]

    row = 6
    for col, val in enumerate(data_row, 1):
        c = ws.cell(row=row, column=col, value=val)
        c.font = Font(size=10)
        c.alignment = ALIGN_WRAP
        c.border = BORDER
        # Colorazione celle esito
        if col == 2:
            c.fill = PatternFill("solid", fgColor=esito_fill(o1))
            c.font = Font(bold=True, size=10)
            c.alignment = ALIGN_CTR
        elif col == 6:
            c.fill = PatternFill("solid", fgColor=esito_fill(o2))
            c.font = Font(bold=True, size=10)
            c.alignment = ALIGN_CTR
        elif col == 12:
            c.fill = PatternFill("solid", fgColor=esito_fill(o3))
            c.font = Font(bold=True, size=10)
            c.alignment = ALIGN_CTR
        # Colorazione metriche numeriche: rosso se > 0 per i problemi
        elif col in (3, 4, 5) and isinstance(val, int) and val > 0:
            c.fill = PatternFill("solid", fgColor="FFF0CC")
            c.alignment = ALIGN_CTR
        elif col in (3, 4, 5):
            c.fill = PatternFill("solid", fgColor="EBF9EE")
            c.alignment = ALIGN_CTR
        elif col == 7:  # H1 ideale = 1
            ok = val == 1
            c.fill = PatternFill("solid", fgColor="EBF9EE" if ok else "FFF0CC")
            c.alignment = ALIGN_CTR
        elif col in (8, 9):  # H2, H3
            c.alignment = ALIGN_CTR
        elif col in (10, 11) and isinstance(val, int):  # salti, orfani
            c.fill = PatternFill("solid", fgColor="FFF0CC" if val > 0 else "EBF9EE")
            c.alignment = ALIGN_CTR
    ws.row_dimensions[row].height = 50

    # ── Sezione dettaglio heading map ──
    row += 2
    ws.merge_cells(f"A{row}:M{row}")
    c = ws[f"A{row}"]
    c.value = "Dettaglio: Mappa Heading della pagina"
    c.font = Font(bold=True, size=11, color="FFFFFF")
    c.fill = PatternFill("solid", fgColor=COLOR_HEAD)
    c.alignment = ALIGN_WRAP
    ws.row_dimensions[row].height = 20
    row += 1

    for col, hdr in enumerate(["Livello", "Tag", "Testo Heading"], 1):
        c = ws.cell(row=row, column=col, value=hdr)
        c.fill = PatternFill("solid", fgColor="1E3A52")
        c.font = Font(bold=True, size=9, color="CBD5E1")
        c.alignment = ALIGN_CTR
        c.border = BORDER
    ws.row_dimensions[row].height = 20
    row += 1

    level_colors = {1: "DDEEFF", 2: "E8F4FF", 3: "F0F8FF"}
    for h in r2.get("heading_map", []):
        indent = "  " * (h["level"] - 1)
        bg = level_colors.get(h["level"], "FFFFFF")
        for col, val in enumerate([h["level"], h["tag"], indent + h["text"]], 1):
            c = ws.cell(row=row, column=col, value=val)
            c.fill = PatternFill("solid", fgColor=bg)
            c.font = Font(bold=(h["level"] <= 2), size=10)
            c.alignment = ALIGN_CTR if col <= 2 else ALIGN_WRAP
            c.border = BORDER
        ws.merge_cells(f"C{row}:M{row}")
        ws.row_dimensions[row].height = 22
        row += 1

    # ── Sezione findings robots.txt ──
    row += 1
    ws.merge_cells(f"A{row}:M{row}")
    c = ws[f"A{row}"]
    c.value = "Dettaglio: Findings Robots.txt"
    c.font = Font(bold=True, size=11, color="FFFFFF")
    c.fill = PatternFill("solid", fgColor=COLOR_HEAD)
    c.alignment = ALIGN_WRAP
    ws.row_dimensions[row].height = 20
    row += 1

    for col, hdr in enumerate(["Severity", "Tipo", "Direttiva", "Linea", "Nota"], 1):
        c = ws.cell(row=row, column=col, value=hdr)
        c.fill = PatternFill("solid", fgColor="1E3A52")
        c.font = Font(bold=True, size=9, color="CBD5E1")
        c.alignment = ALIGN_CTR
        c.border = BORDER
    ws.row_dimensions[row].height = 20
    row += 1

    for f in r3.get("findings", []):
        sev = f.get("severity", "INFO")
        for col, val in enumerate(
            [sev, f.get("type",""), f.get("directive",""), str(f.get("line","—")), f.get("note","")], 1
        ):
            c = ws.cell(row=row, column=col, value=val)
            c.fill = PatternFill("solid", fgColor={"OK": COLOR_OK, "WARN": COLOR_WARN,
                                                    "FAIL": COLOR_FAIL}.get(sev, "FFFFFF") if col == 1 else "FFFFFF")
            c.font = Font(bold=(col == 1), size=10)
            c.alignment = ALIGN_CTR if col <= 4 else ALIGN_WRAP
            c.border = BORDER
        ws.merge_cells(f"E{row}:M{row}")
        ws.row_dimensions[row].height = 36
        row += 1

    if not r3.get("findings"):
        c = ws.cell(row=row, column=1, value="Nessun problema rilevato ✓")
        c.font = Font(bold=True, size=10, color="22C55E")
        ws.merge_cells(f"A{row}:M{row}")


def write_pages_sheet(wb, grouped_results):
    """Sheet: riepilogo pagine analizzate con esito per area."""
    ws = wb.create_sheet("📊 Pagine Analizzate")
    ws.sheet_view.showGridLines = False

    headers = ["#", "URL", "Area 1 — Visibilità", "Area 2 — Heading", "Area 3 — Robots.txt", "Score Globale"]
    col_widths = [5, 60, 22, 22, 22, 16]
    for i, (h, w) in enumerate(zip(headers, col_widths), 1):
        c = ws.cell(row=1, column=i, value=h)
        c.fill = PatternFill("solid", fgColor="1E3A5F")
        c.font = Font(bold=True, size=10, color="FFFFFF")
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border = BORDER
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.row_dimensions[1].height = 28

    sev_color = {"OK": "C6EFCE", "WARN": "FFEB9C", "FAIL": "FFC7CE", "ERROR": "E0E0E0"}
    sev_font = {"OK": "276221", "WARN": "7D4D00", "FAIL": "9C0006", "ERROR": "444444"}
    score_map = {"OK": 100, "WARN": 50, "FAIL": 0, "ERROR": 0}
    score_labels = {"OK": "✓ OK", "WARN": "⚠ WARN", "FAIL": "✗ FAIL", "ERROR": "ERR"}

    for idx, page_results in enumerate(grouped_results, 1):
        row = idx + 1
        url = page_results[0].get("url", "—") if page_results else "—"
        r1 = next((r for r in page_results if "1." in r.get("area", "")), {})
        r2 = next((r for r in page_results if "2." in r.get("area", "")), {})
        r3 = next((r for r in page_results if "3." in r.get("area", "")), {})

        o1 = r1.get("overall", "ERROR")
        o2 = r2.get("overall", "ERROR")
        o3 = r3.get("overall", "ERROR")
        global_score = round((score_map.get(o1, 0) + score_map.get(o2, 0) + score_map.get(o3, 0)) / 3)
        score_color = "C6EFCE" if global_score >= 80 else ("FFEB9C" if global_score >= 50 else "FFC7CE")

        row_data = [idx, url, score_labels.get(o1, o1), score_labels.get(o2, o2), score_labels.get(o3, o3), f"{global_score}/100"]
        row_colors = [None, None, sev_color.get(o1), sev_color.get(o2), sev_color.get(o3), score_color]
        row_ffonts = [None, None, sev_font.get(o1), sev_font.get(o2), sev_font.get(o3), None]

        for ci, (val, fc, ff) in enumerate(zip(row_data, row_colors, row_ffonts), 1):
            c = ws.cell(row=row, column=ci, value=val)
            if fc:
                c.fill = PatternFill("solid", fgColor=fc)
            c.font = Font(bold=(ci in [3, 4, 5, 6]), size=10,
                          color=(ff if ff else "000000"))
            c.alignment = Alignment(horizontal="center" if ci != 2 else "left",
                                    vertical="center", wrap_text=True)
            c.border = BORDER
        ws.row_dimensions[row].height = 22


def write_problems_sheet(wb, grouped_results):
    """Sheet: elenco problemi per area con URL, severity, note."""
    ws = wb.create_sheet("⚠ Problemi per Area")
    ws.sheet_view.showGridLines = False

    headers = ["Area", "URL", "Tipo Problema", "Severity", "Dettaglio"]
    col_widths = [24, 48, 28, 12, 60]
    for i, (h, w) in enumerate(zip(headers, col_widths), 1):
        c = ws.cell(row=1, column=i, value=h)
        c.fill = PatternFill("solid", fgColor="1E3A5F")
        c.font = Font(bold=True, size=10, color="FFFFFF")
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.border = BORDER
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.row_dimensions[1].height = 28

    sev_color = {"OK": "C6EFCE", "WARN": "FFEB9C", "FAIL": "FFC7CE", "ERROR": "E0E0E0", "INFO": "DDEEFF"}
    row = 2

    for page_results in grouped_results:
        r1 = next((r for r in page_results if "1." in r.get("area", "")), {})
        r2 = next((r for r in page_results if "2." in r.get("area", "")), {})
        r3 = next((r for r in page_results if "3." in r.get("area", "")), {})
        pg_url = r1.get("url") or r2.get("url") or r3.get("url") or "—"

        problems = []

        # Area 1 problems
        for b in r1.get("hidden_blocks", []):
            problems.append(("1. Visibilità Contenuti", pg_url, "Contenuto nascosto", "WARN",
                             f"Selettore: {b['selector']} — «{b['text_preview'][:80]}»"))
        if r1.get("lazy_elements", 0) > 0:
            problems.append(("1. Visibilità Contenuti", pg_url, "Lazy-load / scroll-trigger", "WARN",
                             f"{r1['lazy_elements']} elementi con attributi lazy-load rilevati"))
        for pat in r1.get("js_patterns", []):
            problems.append(("1. Visibilità Contenuti", pg_url, "Pattern JS iniezione", "WARN",
                             f"Pattern: {pat}"))

        # Area 2 problems
        if r2.get("h1_count", 1) == 0:
            problems.append(("2. Struttura Heading", pg_url, "H1 assente", "FAIL",
                             "Nessun H1 trovato nella pagina"))
        elif r2.get("h1_count", 1) > 1:
            problems.append(("2. Struttura Heading", pg_url, "H1 multipli", "WARN",
                             f"{r2['h1_count']} H1 trovati — deve essere uno solo"))
        for jump in r2.get("level_jumps", []):
            problems.append(("2. Struttura Heading", pg_url, "Salto di livello heading", "WARN", jump))
        for orphan in r2.get("orphan_blocks", []):
            problems.append(("2. Struttura Heading", pg_url, "Blocco orfano >300 parole", "WARN",
                             orphan[:100]))

        # Area 3 problems
        for f in r3.get("findings", []):
            problems.append(("3. Robots.txt", pg_url, f.get("type", "—"), f.get("severity", "INFO"),
                             f"L.{f.get('line','—')} {f.get('directive','')} — {f.get('note','')}"))

        for area, pg_url_w, tipo, sev, note in problems:
            fc = sev_color.get(sev, "FFFFFF")
            row_data = [area, pg_url_w, tipo, sev, note]
            for ci, val in enumerate(row_data, 1):
                c = ws.cell(row=row, column=ci, value=val)
                if ci == 4:
                    c.fill = PatternFill("solid", fgColor=fc)
                    c.font = Font(bold=True, size=10)
                else:
                    c.font = Font(size=10)
                c.alignment = Alignment(horizontal="center" if ci == 4 else "left",
                                        vertical="top", wrap_text=True)
                c.border = BORDER
            ws.row_dimensions[row].height = 32
            row += 1

    if row == 2:
        c = ws.cell(row=2, column=1, value="Nessun problema rilevato ✓")
        c.font = Font(bold=True, size=11, color="276221")
        ws.merge_cells("A2:E2")


def export_excel(url, grouped_results, output_path):
    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    timestamp = datetime.now().strftime("%d/%m/%Y %H:%M")

    # Usa i risultati del primo URL per i sheet di dettaglio legacy
    first = grouped_results[0] if grouped_results else []
    write_summary_sheet(wb, url, first, timestamp)
    write_sintesi_sheet(wb, url, first, timestamp)

    # Nuovi sheet multi-URL
    write_pages_sheet(wb, grouped_results)
    write_problems_sheet(wb, grouped_results)

    # Sheet di dettaglio per il primo URL (comportamento originale)
    for r in first:
        area = r.get("area", "")
        if "1." in area:
            write_visibility_sheet(wb, r)
        elif "2." in area:
            write_heading_sheet(wb, r)
        elif "3." in area:
            write_robots_sheet(wb, r)

    wb.save(output_path)
    print(f"\n✅ Report Excel salvato: {output_path}")


# ─── STAMPA CONSOLE ───────────────────────────────────────────────────────────

ICONS = {"OK": "✅", "WARN": "⚠️ ", "FAIL": "❌", "ERROR": "💥"}

def print_result(r):
    icon = ICONS.get(r["overall"], "•")
    print(f"\n{icon} {r['area']} — {r['overall']}")
    print(f"   {r.get('summary', '')}")
    if r.get("hidden_blocks"):
        print(f"   Blocchi nascosti: {len(r['hidden_blocks'])}")
    if r.get("findings"):
        for f in r["findings"][:5]:
            print(f"   [{f['severity']}] L.{f['line']} {f['directive']}: {f['note'][:80]}")


# ─── SITEMAP PARSER ───────────────────────────────────────────────────────────

def get_urls_from_sitemap(sitemap_url, session, max_urls=20):
    urls = []
    try:
        r = session.get(sitemap_url, timeout=REQUEST_TIMEOUT)
        soup = BeautifulSoup(r.text, "lxml-xml")
        locs = soup.find_all("loc")
        for loc in locs[:max_urls]:
            url = loc.get_text(strip=True)
            if url:
                urls.append(url)
    except Exception as e:
        print(f"⚠ Errore sitemap: {e}")
    return urls


# ─── MAIN ─────────────────────────────────────────────────────────────────────

def run_audit(url, use_playwright=True, verbose=True):
    session = get_session()
    print(f"\n🔍 Analisi: {url}")

    # Fetch
    print("  → Fetch statico…")
    static_html, status, final_url = fetch_html_static(url, session)
    if not static_html:
        print(f"  ❌ Impossibile recuperare la pagina: {final_url}")
        return None

    rendered_html = None
    if use_playwright:
        print("  → Rendering JS (Playwright)…")
        rendered_html = fetch_html_rendered(url)

    # Audit
    print("  → Area 1: visibilità contenuti…")
    r1 = audit_content_visibility(url, static_html, rendered_html)

    print("  → Area 2: struttura heading…")
    r2 = audit_heading_structure(url, static_html)

    print("  → Area 3: robots.txt…")
    r3 = audit_robots_txt(url, session)

    print("  → Area 8: E-E-A-T signals…")
    r8 = audit_eeat_signals(url, static_html)

    print("  → Area 9: performance signals…")
    r9 = audit_performance_signals(url, static_html)

    print("  → Area 10: autorevolezza topica…")
    r10 = audit_topical_authority(url, static_html, session)

    for r in [r1, r2, r3, r8, r9, r10]:
        r["url"] = url

    results = [r1, r2, r3, r8, r9, r10]

    if verbose:
        for r in results:
            print_result(r)

    return results


def main():
    parser = argparse.ArgumentParser(
        description="SEO Audit: visibilità contenuti, heading, robots.txt"
    )
    parser.add_argument("--url", help="URL da analizzare")
    parser.add_argument("--sitemap", help="URL sitemap per analisi multipla")
    parser.add_argument("--max", type=int, default=10, help="Max URL da sitemap (default: 10)")
    parser.add_argument("--output", default="seo_audit_report.xlsx", help="File Excel di output")
    parser.add_argument("--no-playwright", action="store_true", help="Disabilita rendering JS")
    args = parser.parse_args()

    use_playwright = not args.no_playwright

    if args.sitemap:
        session = get_session()
        print(f"📄 Lettura sitemap: {args.sitemap}")
        urls = get_urls_from_sitemap(args.sitemap, session, args.max)
        print(f"   Trovati {len(urls)} URL da analizzare.")

        grouped_results = []
        for url in urls:
            res = run_audit(url, use_playwright=use_playwright, verbose=True)
            if res:
                grouped_results.append(res)

        if grouped_results:
            export_excel(args.sitemap, grouped_results, args.output)
            html_path = args.output.replace(".xlsx", ".html") if args.output.endswith(".xlsx") else args.output + ".html"
            export_html(args.sitemap, grouped_results, html_path)

    elif args.url:
        results = run_audit(args.url, use_playwright=use_playwright)
        if results:
            grouped_results = [results]
            export_excel(args.url, grouped_results, args.output)
            html_path = args.output.replace(".xlsx", ".html") if args.output.endswith(".xlsx") else args.output + ".html"
            export_html(args.url, grouped_results, html_path)
    else:
        parser.print_help()
        print("\nEsempio:")
        print("  python seo_audit.py --url https://esempio.com")
        print("  python seo_audit.py --url https://esempio.com --output report.xlsx")
        print("  python seo_audit.py --sitemap https://esempio.com/sitemap.xml --max 5")
        sys.exit(1)


if __name__ == "__main__":
    main()
