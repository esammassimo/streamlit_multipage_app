"""
NVL Agency · SEO/SEA Analysis Tool
Streamlit app — integra Google Ads + Search Console e produce report Excel
"""

import io
import re
from pathlib import Path

import numpy as np
import pandas as pd
import streamlit as st
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

# ─────────────────────────────────────────────────────────────
# PAGE CONFIG
# ─────────────────────────────────────────────────────────────
# ─────────────────────────────────────────────
# PAGE CONFIG
# set_page_config() va chiamato SOLO in Home.py.
# Decommentare solo per esecuzione standalone.
# ─────────────────────────────────────────────
# st.set_page_config(
#     page_title="SEO/SEA Analysis · NVL Agency",
#     page_icon="📊",
#     layout="wide",
#     initial_sidebar_state="expanded",
# )



# ─────────────────────────────────────────────────────────────
# CORE LOGIC  (ported from seo_sea_analysis.py)
# ─────────────────────────────────────────────────────────────

COL_MAP = {
    "keyword":           ["search keyword", "keyword", "parola chiave"],
    "keyword_status":    ["search keyword status", "keyword status", "stato parola chiave", "status"],
    "match_type":        ["search keyword match type", "match type", "tipo di corrispondenza", "tipo corrispondenza"],
    "brand_nobrand":     ["brand / no-brand", "brand / no brand", "brand/no-brand", "brand/no brand"],
    "territorio":        ["territorio", "[new] territorio 2025", "territory"],
    "campaign":          ["campaign", "campagna"],
    "ad_group":          ["ad group", "gruppo di annunci"],
    "currency":          ["currency code", "codice valuta"],
    "max_cpc":           ["keyword max cpc", "max. cpc", "max cpc", "cpc max", "max, cpc"],
    "final_url":         ["keyword final url", "final url", "url finale", "url"],
    "impressions":       ["impr.", "impressions", "impressioni", "impr,"],
    "clicks":            ["clicks", "clic", "click"],
    "ctr":               ["ctr", "interaction rate", "tasso di interazione"],
    "avg_cpc":           ["avg. cpc", "avg, cpc", "avg cpc", "cpc medio", "avg. cost", "avg, cost"],
    "cost":              ["cost", "costo"],
    "search_impr_share": ["search impr. share", "search impr, share", "quota imp. rete ricerca", "search impression share"],
    "search_lost_is_rank": ["search lost is (rank)", "search lost is rank", "quota imp. perse classif.", "quota imp. perse (rank)"],
    "quality_score":     ["quality score", "punteggio di qualità", "qs"],
    "exp_ctr":           ["exp. ctr", "exp, ctr", "ctr previsto", "expected ctr"],
    "landing_page_exp":  ["landing page exp.", "landing page exp,", "esperienza pagina dest.", "landing page experience"],
    "ad_relevance":      ["ad relevance", "pertinenza annuncio"],
    "conv_rate":         ["conv. rate", "conv, rate", "tasso di conversione"],
    "conv_value":        ["conv. value", "conv, value", "valore conv.", "conversions value"],
    "conversions":       ["conversions", "conversioni"],
    "cost_per_conv":     ["cost / conv.", "cost / conv,", "costo / conv.", "cost per conversion"],
    "pos_organica":      ["posizione organica gsc", "posizione organica", "organic position", "position", "avg. position", "avg position", "posizione media"],
    "url_posizionata":   ["url posizionata", "url posizionata gsc", "landing page", "page", "pagina"],
}

QS_TEXT_MAP = {"above average": 8, "average": 5, "below average": 3}


def normalize_col(name: str) -> str:
    return str(name).strip().lower().replace("  ", " ")


def detect_and_rename(df: pd.DataFrame) -> pd.DataFrame:
    col_lower = {normalize_col(c): c for c in df.columns}
    rename_map = {}
    used_internals = set()
    for internal, variants in COL_MAP.items():
        if internal in used_internals:
            continue
        for v in variants:
            if normalize_col(v) in col_lower:
                orig = col_lower[normalize_col(v)]
                if orig not in rename_map:
                    rename_map[orig] = internal
                    used_internals.add(internal)
                    break
    return df.rename(columns=rename_map)


def parse_ads_file(uploaded_file) -> pd.DataFrame:
    name = uploaded_file.name
    suffix = Path(name).suffix.lower()
    buf = io.BytesIO(uploaded_file.read())

    target_sheet = 0
    if suffix in (".xlsx", ".xls"):
        buf.seek(0)
        xf = pd.ExcelFile(buf)
        all_sheets = xf.sheet_names
        if len(all_sheets) > 1:
            # Priorità: sheet che ha una colonna chiamata esattamente "keyword" o "parola chiave"
            # (non solo che contiene la parola nel testo), cercando nelle prime 6 righe
            best_sheet = None
            for s in all_sheets:
                buf.seek(0)
                probe = pd.read_excel(buf, sheet_name=s, header=None, nrows=6, dtype=str)
                for _, prow in probe.iterrows():
                    vals = [normalize_col(str(v)) for v in prow if pd.notna(v) and str(v).strip()]
                    # Deve avere una colonna con nome ESATTO "keyword" o "parola chiave"
                    # (non url, non search term, non ad status)
                    has_keyword_col = any(
                        v in ("keyword", "parola chiave", "search keyword")
                        for v in vals
                    )
                    if has_keyword_col and len(vals) >= 5:
                        best_sheet = s
                        break
                if best_sheet == s:
                    break
            if best_sheet:
                target_sheet = best_sheet

        buf.seek(0)
        raw = pd.read_excel(buf, sheet_name=target_sheet, header=None, dtype=str)
    else:
        buf.seek(0)
        raw = pd.read_csv(buf, header=None, dtype=str, encoding="utf-8-sig")

    header_row = None
    for i, row in raw.iterrows():
        vals = [normalize_col(str(v)) for v in row if pd.notna(v) and str(v).strip()]
        if len(vals) >= 3 and any("keyword" in v or "parola chiave" in v for v in vals):
            header_row = i
            break

    if header_row is None:
        raise ValueError(f"Impossibile trovare la riga header in '{name}'")

    buf.seek(0)
    if suffix in (".xlsx", ".xls"):
        df = pd.read_excel(buf, sheet_name=target_sheet, header=header_row, dtype=str)
    else:
        df = pd.read_csv(buf, header=header_row, dtype=str, encoding="utf-8-sig")

    df = df.dropna(how="all")
    df = detect_and_rename(df)

    # Fallback: se "keyword" non riconosciuta, cerca colonna residua con "keyword"/"parola chiave"
    if "keyword" not in df.columns:
        for c in df.columns:
            if "keyword" in normalize_col(c) or "parola chiave" in normalize_col(c):
                df = df.rename(columns={c: "keyword"})
                break
    # Ultimo fallback: usa la prima colonna non-vuota come keyword
    if "keyword" not in df.columns:
        first_text_col = next((c for c in df.columns if df[c].dtype == object), None)
        if first_text_col:
            df = df.rename(columns={first_text_col: "keyword"})

    if "keyword" in df.columns:
        df["keyword"] = (df["keyword"].astype(str).str.strip()
                         .str.replace(r'^\[|\]$', '', regex=True)
                         .str.replace(r'^"|"$', '', regex=True)
                         .str.strip())

    if "match_type" in df.columns:
        mt_map = {
            "exact match": "Exact match", "exact": "Exact match", "corrispondenza esatta": "Exact match",
            "broad match": "Broad match", "broad": "Broad match", "corrispondenza generica": "Broad match",
            "phrase match": "Phrase match", "phrase": "Phrase match", "corrispondenza a frase": "Phrase match",
        }
        df["match_type"] = df["match_type"].str.strip().str.lower().map(
            lambda x: mt_map.get(x, x.title() if isinstance(x, str) else x))

    if "quality_score" in df.columns:
        def parse_qs(v):
            if pd.isna(v) or str(v).strip() in ("--", " --", ""):
                return np.nan
            v_low = str(v).strip().lower()
            if v_low in QS_TEXT_MAP:
                return QS_TEXT_MAP[v_low]
            try:
                return float(v)
            except ValueError:
                return np.nan
        df["quality_score"] = df["quality_score"].apply(parse_qs)

    numeric_cols = ["impressions", "clicks", "cost", "avg_cpc", "max_cpc",
                    "conv_rate", "conv_value", "conversions", "cost_per_conv"]
    for col in numeric_cols:
        if col in df.columns:
            df[col] = pd.to_numeric(
                df[col].astype(str).str.replace(",", ".").str.replace(" ", "")
                .str.replace("--", "").str.replace(" --", ""), errors="coerce")

    for col in ["ctr", "search_impr_share", "search_lost_is_rank"]:
        if col in df.columns:
            df[col] = (df[col].astype(str)
                       .str.replace("< 10%", "0.09")
                       .str.replace("%", "").str.replace("--", "").str.replace(" --", "")
                       .str.replace(",", "."))
            df[col] = pd.to_numeric(df[col], errors="coerce")
            mask = df[col] > 1
            df.loc[mask, col] = df.loc[mask, col] / 100

    return df


# Varianti colonna query GSC (IT + EN)
GSC_QUERY_VARIANTS  = ("query", "top queries", "search query", "keyword", "parola chiave", "query di ricerca")
GSC_POS_VARIANTS    = ("position", "avg. position", "avg, position", "posizione media", "posizione", "pos. media")
GSC_PAGE_VARIANTS   = ("page", "top pages", "pagina", "landing page", "url")
GSC_CLICKS_VARIANTS = ("clicks", "clic", "click")
GSC_IMPR_VARIANTS   = ("impressions", "impressioni")


def parse_gsc_file(uploaded_file) -> pd.DataFrame:
    """
    Parser dedicato per export Search Console.
    Non usa detect_and_rename (pensato per Ads) — lavora sulle colonne raw
    così da evitare conflitti sul nome 'keyword'.
    """
    name = uploaded_file.name
    suffix = Path(name).suffix.lower()
    buf = io.BytesIO(uploaded_file.read())

    if suffix in (".xlsx", ".xls"):
        raw = pd.read_excel(buf, sheet_name=0, header=None, dtype=str)
    else:
        buf.seek(0)
        raw = pd.read_csv(buf, header=None, dtype=str, encoding="utf-8-sig")

    # Trova riga header: contiene "query" o "keyword" o "clicks"
    header_row = 0
    for i, row in raw.iterrows():
        vals = [normalize_col(str(v)) for v in row if pd.notna(v) and str(v).strip()]
        if any(v in GSC_QUERY_VARIANTS or v in GSC_CLICKS_VARIANTS for v in vals):
            header_row = i
            break

    buf.seek(0)
    if suffix in (".xlsx", ".xls"):
        df = pd.read_excel(buf, sheet_name=0, header=header_row, dtype=str)
    else:
        df = pd.read_csv(buf, header=header_row, dtype=str, encoding="utf-8-sig")

    df = df.dropna(how="all")

    # Rename diretto senza detect_and_rename per evitare conflitti con COL_MAP Ads
    rename = {}
    for c in df.columns:
        cn = normalize_col(c)
        if cn in GSC_QUERY_VARIANTS and "gsc_keyword" not in rename.values():
            rename[c] = "gsc_keyword"
        elif cn in GSC_POS_VARIANTS and "pos_organica" not in rename.values():
            rename[c] = "pos_organica"
        elif cn in GSC_PAGE_VARIANTS and "url_posizionata" not in rename.values():
            rename[c] = "url_posizionata"
        elif cn in GSC_CLICKS_VARIANTS and "gsc_clicks" not in rename.values():
            rename[c] = "gsc_clicks"
        elif cn in GSC_IMPR_VARIANTS and "gsc_impressions" not in rename.values():
            rename[c] = "gsc_impressions"
    df = df.rename(columns=rename)

    if "gsc_keyword" not in df.columns:
        raise ValueError(
            f"Impossibile trovare la colonna query/keyword nel file GSC '{name}'. "
            f"Colonne rilevate: {list(df.columns)}"
        )

    for col in ["pos_organica", "gsc_clicks", "gsc_impressions"]:
        if col in df.columns:
            df[col] = pd.to_numeric(df[col].astype(str).str.replace(",", "."), errors="coerce")

    df["gsc_keyword"] = df["gsc_keyword"].str.strip().str.lower()

    return df


def join_with_gsc(df_ads: pd.DataFrame, df_gsc) -> pd.DataFrame:
    if df_gsc is None or df_gsc.empty:
        if "pos_organica" not in df_ads.columns:
            df_ads["pos_organica"] = np.nan
        if "url_posizionata" not in df_ads.columns:
            df_ads["url_posizionata"] = ""
        return df_ads

    if "keyword" not in df_ads.columns:
        # GSC caricato ma Ads non ha colonna keyword — skip join, aggiungi colonne vuote
        df_ads["pos_organica"] = np.nan
        df_ads["url_posizionata"] = ""
        return df_ads

    df_ads["_kw_join"] = df_ads["keyword"].fillna("").astype(str).str.strip().str.lower()
    gsc_cols = ["gsc_keyword", "pos_organica"]
    if "url_posizionata" in df_gsc.columns:
        gsc_cols.append("url_posizionata")

    df_gsc_agg = df_gsc[gsc_cols].copy()
    agg_dict = {"pos_organica": ("pos_organica", "mean")}
    if "url_posizionata" in df_gsc_agg.columns:
        agg_dict["url_posizionata"] = ("url_posizionata", "first")
    df_gsc_agg = df_gsc_agg.groupby("gsc_keyword").agg(**agg_dict).reset_index()

    df_ads = df_ads.merge(
        df_gsc_agg.rename(columns={"gsc_keyword": "_kw_join"}),
        on="_kw_join", how="left", suffixes=("", "_gsc"))
    df_ads = df_ads.drop(columns=["_kw_join"])

    for col in ["pos_organica", "url_posizionata"]:
        if f"{col}_gsc" in df_ads.columns:
            df_ads[col] = df_ads[col].fillna(df_ads[f"{col}_gsc"])
            df_ads = df_ads.drop(columns=[f"{col}_gsc"])

    return df_ads


# Varianti colonne Screaming Frog
SF_ADDRESS_VARIANTS = ("address", "url", "indirizzo")
SF_TITLE_VARIANTS   = ("title 1", "title", "titolo", "meta title", "page title")
SF_STATUS_VARIANTS  = ("status code", "status", "codice stato", "http status")


def parse_screaming_frog(uploaded_file) -> pd.DataFrame:
    """Legge export Screaming Frog. Restituisce df con sf_url, sf_title, sf_status."""
    name = uploaded_file.name
    suffix = Path(name).suffix.lower()
    buf = io.BytesIO(uploaded_file.read())

    if suffix in (".xlsx", ".xls"):
        raw = pd.read_excel(buf, sheet_name=0, header=None, nrows=3, dtype=str)
    else:
        buf.seek(0)
        raw = pd.read_csv(buf, header=None, nrows=3, dtype=str, encoding="utf-8-sig")

    header_row = 0
    for i, row in raw.iterrows():
        vals = [normalize_col(str(v)) for v in row if pd.notna(v) and str(v).strip()]
        if any(v in SF_ADDRESS_VARIANTS or v in SF_TITLE_VARIANTS for v in vals):
            header_row = i
            break

    buf.seek(0)
    if suffix in (".xlsx", ".xls"):
        df = pd.read_excel(buf, sheet_name=0, header=header_row, dtype=str)
    else:
        df = pd.read_csv(buf, header=header_row, dtype=str, encoding="utf-8-sig")

    df = df.dropna(how="all")
    rename = {}
    for c in df.columns:
        cn = normalize_col(c)
        if cn in SF_ADDRESS_VARIANTS and "sf_url" not in rename.values():
            rename[c] = "sf_url"
        elif cn in SF_TITLE_VARIANTS and "sf_title" not in rename.values():
            rename[c] = "sf_title"
        elif cn in SF_STATUS_VARIANTS and "sf_status" not in rename.values():
            rename[c] = "sf_status"
    df = df.rename(columns=rename)

    if "sf_url" not in df.columns:
        raise ValueError(
            f"Impossibile trovare colonna URL/Address nel file Screaming Frog '{name}'. "
            f"Colonne: {list(df.columns)}"
        )

    df["sf_url"] = df["sf_url"].astype(str).str.strip().str.rstrip("/").str.lower()
    if "sf_title" in df.columns:
        df["sf_title"] = df["sf_title"].astype(str).str.strip().replace("nan", "")

    return df[["sf_url"] + [c for c in ["sf_title", "sf_status"] if c in df.columns]]


def normalize_url(url):
    """Lowercase, no trailing slash, no parametri UTM/tracking."""
    if not url or str(url).strip() in ("", "nan", "--", " --"):
        return ""
    u = str(url).strip().lower().rstrip("/")
    if "?" in u:
        base, params = u.split("?", 1)
        kept = [p for p in params.split("&")
                if not any(p.startswith(k) for k in ("utm_", "gclid", "fbclid", "_ga"))]
        u = base + ("?" + "&".join(kept) if kept else "")
    return u


def url_depth(url):
    from urllib.parse import urlparse
    path = urlparse(url).path.rstrip("/")
    return len([s for s in path.split("/") if s])


def build_url_title_check(df_dopo, df_gsc, df_sf):
    """
    Confronto URL ADS vs GSC e check keyword/title da Screaming Frog.
    Ritorna DataFrame pronto per visualizzazione.
    """
    rows = []
    sf_lookup = {}
    if df_sf is not None and not df_sf.empty and "sf_title" in df_sf.columns:
        sf_lookup = dict(zip(df_sf["sf_url"], df_sf["sf_title"]))

    for _, row in df_dopo.iterrows():
        kw = str(row.get("keyword", "")).strip()
        if not kw or kw in ("", "nan"):
            continue

        ads_url_raw = str(row.get("final_url", "")).strip()
        gsc_url_raw = str(row.get("url_posizionata", "")).strip()

        # Salta righe senza nessun URL
        ads_empty = ads_url_raw in ("", "nan", "--", " --")
        gsc_empty = gsc_url_raw in ("", "nan", "--", " --")
        if ads_empty and gsc_empty:
            continue

        ads_url = normalize_url(ads_url_raw)
        gsc_url = normalize_url(gsc_url_raw)

        # URL match
        if not ads_url or not gsc_url:
            url_match = "N/D"
        elif ads_url == gsc_url:
            url_match = "OK"
        else:
            url_match = "KO"

        # GSC più generica?
        gsc_generica = ""
        if ads_url and gsc_url and ads_url != gsc_url:
            gsc_generica = "Sì" if url_depth(gsc_url) < url_depth(ads_url) else "No"

        # Title da SF
        title_ads = sf_lookup.get(ads_url, "") if ads_url else ""
        title_gsc = sf_lookup.get(gsc_url, "") if gsc_url else ""
        has_sf = bool(sf_lookup)

        kw_lower = kw.lower()
        kw_in_title_ads = ("Sì" if kw_lower in title_ads.lower() else "No") if (has_sf and title_ads) else ("N/D" if has_sf else "—")
        kw_in_title_gsc = ("Sì" if kw_lower in title_gsc.lower() else "No") if (has_sf and title_gsc) else ("N/D" if has_sf else "—")

        # Coerenza complessiva
        if not has_sf:
            coerenza = "—"
        elif title_ads and url_match == "OK" and kw_lower in title_ads.lower():
            coerenza = "OK"
        elif title_ads and kw_lower in title_ads.lower():
            coerenza = "Parziale"
        else:
            coerenza = "KO"

        rows.append({
            "Keyword":          kw,
            "Match Type":       str(row.get("match_type", "")),
            "URL ADS":          "" if ads_empty else ads_url_raw,
            "URL GSC":          "" if gsc_empty else gsc_url_raw,
            "URL Match":        url_match,
            "GSC generica":     gsc_generica,
            "Title ADS":        title_ads,
            "Title GSC":        title_gsc,
            "KW in title ADS":  kw_in_title_ads,
            "KW in title GSC":  kw_in_title_gsc,
            "Coerenza":         coerenza,
        })

    return pd.DataFrame(rows)

# Pattern per riconoscere le colonne speciali nel file di classificazione
BRAND_COL_PATTERNS    = ("brand",)
CATEGORIA_COL_PATTERNS = ("cluster", "territorio", "territory", "territories", "categoria", "category")

# Nomi interni normalizzati usati nel df e nell'Excel
INTERNAL_BRAND_COL    = "Brand/NoBrand"
INTERNAL_CATEGORIA_COL = "Categoria"


def detect_classif_col_type(col_name: str):
    """
    Ritorna 'brand', 'categoria' o None in base al nome della colonna.
    """
    cn = normalize_col(col_name)
    if any(p in cn for p in BRAND_COL_PATTERNS):
        return "brand"
    if any(p in cn for p in CATEGORIA_COL_PATTERNS):
        return "categoria"
    return None


def parse_classification_file(uploaded_file):
    """
    Legge un file di classificazione keyword.
    Prima colonna attesa: 'keyword'. Le colonne successive sono classificate come:
      - Brand/NoBrand  → colonna il cui nome contiene 'brand'
      - Categoria      → colonna il cui nome contiene 'cluster', 'territorio',
                         'territory', 'territories', 'categoria', 'category'
      - Altre          → aggiunte dinamicamente col nome originale

    Le colonne Brand e Categoria vengono rinominate ai nomi interni standard
    (INTERNAL_BRAND_COL, INTERNAL_CATEGORIA_COL) per uniformità tra file diversi.

    Ritorna (df, extra_cols, brand_col, categoria_col) dove:
      - extra_cols  : lista di tutti i nomi colonna nel df (già rinominati)
      - brand_col   : nome interno colonna brand (o None)
      - categoria_col: nome interno colonna categoria (o None)
    """
    name = uploaded_file.name
    suffix = Path(name).suffix.lower()
    buf = io.BytesIO(uploaded_file.read())

    if suffix in (".xlsx", ".xls"):
        df = pd.read_excel(buf, dtype=str)
    else:
        buf.seek(0)
        df = pd.read_csv(buf, dtype=str, encoding="utf-8-sig")

    df = df.dropna(how="all")

    # Trova colonna keyword
    kw_col = None
    for c in df.columns:
        if normalize_col(str(c)) in ("keyword", "parola chiave", "kw", "search keyword"):
            kw_col = c
            break
    if kw_col is None:
        kw_col = df.columns[0]

    extra_raw = [c for c in df.columns if c != kw_col]
    if not extra_raw:
        raise ValueError(
            f"Il file di classificazione deve avere almeno 2 colonne "
            f"(keyword + almeno una etichetta). Trovate: {list(df.columns)}"
        )

    # Rinomina colonne speciali ai nomi interni standard
    rename_map = {kw_col: "_classif_kw"}
    brand_col = None
    categoria_col = None
    for c in extra_raw:
        col_type = detect_classif_col_type(c)
        if col_type == "brand" and brand_col is None:
            rename_map[c] = INTERNAL_BRAND_COL
            brand_col = INTERNAL_BRAND_COL
        elif col_type == "categoria" and categoria_col is None:
            rename_map[c] = INTERNAL_CATEGORIA_COL
            categoria_col = INTERNAL_CATEGORIA_COL
        # Altre colonne: mantieni nome originale

    df = df.rename(columns=rename_map)
    df["_classif_kw"] = df["_classif_kw"].astype(str).str.strip().str.lower()
    df = df.drop_duplicates(subset=["_classif_kw"])

    # extra_cols: nomi colonne nel df dopo rename (escluso _classif_kw)
    extra_cols = [c for c in df.columns if c != "_classif_kw"]

    return df, extra_cols, brand_col, categoria_col


def join_classification(df_ads, df_classif, extra_cols):
    """
    Join left tra df_ads e df_classif sulla keyword normalizzata.
    Aggiunge le extra_cols al df_ads; keyword senza match → colonne vuote.
    """
    if df_classif is None or df_classif.empty:
        return df_ads

    df_ads = df_ads.copy()
    df_ads["_join_kw"] = df_ads["keyword"].fillna("").astype(str).str.strip().str.lower()

    df_ads = df_ads.merge(
        df_classif[["_classif_kw"] + extra_cols].rename(columns={"_classif_kw": "_join_kw"}),
        on="_join_kw", how="left"
    )
    df_ads = df_ads.drop(columns=["_join_kw"])

    # Riempi NaN con stringa vuota nelle colonne aggiunte
    for c in extra_cols:
        if c in df_ads.columns:
            df_ads[c] = df_ads[c].fillna("")

    return df_ads

def check_qs(qs, threshold=7) -> str:
    if pd.isna(qs):
        return ""
    return "OK" if float(qs) >= threshold else ("KO" if float(qs) > 0 else "")


def classify_qs_flag(df: pd.DataFrame, qs_brand=7, qs_nobrand=7,
                     brand_col=None, brand_values=None) -> pd.DataFrame:
    """
    Applica check QS con soglie separate per Brand e No-Brand se disponibili.
    brand_col: nome colonna brand nel df (es. "Brand/NoBrand")
    brand_values: set di valori considerati "brand" (case-insensitive)
    """
    if "quality_score" not in df.columns:
        df["check_qs"] = ""
        return df

    if brand_col and brand_col in df.columns and brand_values:
        def _check(row):
            qs = row["quality_score"]
            val = str(row.get(brand_col, "")).strip().lower()
            is_brand = any(bv in val for bv in brand_values)
            thr = qs_brand if is_brand else qs_nobrand
            return check_qs(qs, thr)
        df["check_qs"] = df.apply(_check, axis=1)
    else:
        thr = qs_brand  # unica soglia se non c'è distinzione
        df["check_qs"] = df["quality_score"].apply(lambda x: check_qs(x, thr))
    return df


def qs_stats(df):
    qs = df["quality_score"].dropna() if "quality_score" in df.columns else pd.Series(dtype=float)
    total = len(qs)
    ok = int((qs >= 7).sum())
    ko = int((qs < 7).sum())
    avg = float(qs.mean()) if total > 0 else 0.0
    pct_ok = ok / total if total > 0 else 0.0
    return {"total": total, "ok": ok, "ko": ko, "avg": avg, "pct_ok": pct_ok}


def active_stats(df):
    impr_col = df["impressions"].fillna(0) if "impressions" in df.columns else pd.Series(0, index=df.index)
    active = df[impr_col > 0]
    if "brand_nobrand" in df.columns:
        brand = active[active["brand_nobrand"].str.lower().str.contains("brand", na=False)
                       & ~active["brand_nobrand"].str.lower().str.contains("no", na=False)]
        nobrand = active[active["brand_nobrand"].str.lower().str.contains("no", na=False)]
    else:
        camp = active.get("campaign", pd.Series("", index=active.index)).fillna("")
        brand = active[camp.str.lower().str.contains("brand", na=False)]
        nobrand = active[~camp.str.lower().str.contains("brand", na=False)]
    cost = float(active["cost"].sum()) if "cost" in active.columns else 0.0
    clicks = int(active["clicks"].sum()) if "clicks" in active.columns else 0
    impr = int(active["impressions"].sum()) if "impressions" in active.columns else 0
    return {"total": len(active), "brand": len(brand), "nobrand": len(nobrand),
            "cost": cost, "clicks": clicks, "impr": impr}


def seo_coverage(df):
    if "pos_organica" not in df.columns:
        return {"pct_covered": 0.0, "pct_top10": 0.0, "avg_pos": float("nan")}
    pos = pd.to_numeric(df["pos_organica"], errors="coerce")
    has_seo = pos.notna() & (pos > 0)
    top10 = has_seo & (pos <= 10)
    return {
        "pct_covered": float(has_seo.mean()),
        "pct_top10": float(top10.sum() / len(df)) if len(df) > 0 else 0.0,
        "avg_pos": float(pos[has_seo].mean()) if has_seo.any() else float("nan"),
    }


# ─────────────────────────────────────────────────────────────
# EXCEL EXPORT
# ─────────────────────────────────────────────────────────────
C_HEADER_DARK = "1F3864"
C_HEADER_MID  = "2E5FAB"
C_OK    = "C6EFCE"; C_KO = "FFC7CE"; C_WARN = "FFEB9C"
C_WHITE = "FFFFFF"; C_GRAY  = "F5F5F5"

OUTPUT_COLS = [
    ("keyword_status",    "Stato KW"),
    ("keyword",           "Search Keyword"),
    ("brand_nobrand",     "Brand / No-Brand"),
    ("territorio",        "Territorio"),
    ("match_type",        "Match Type"),
    ("campaign",          "Campaign"),
    ("ad_group",          "Ad Group"),
    ("max_cpc",           "Max CPC"),
    ("final_url",         "Final URL"),
    ("quality_score",     "Quality Score"),
    ("check_qs",          "Check QS"),
    ("exp_ctr",           "Exp. CTR"),
    ("landing_page_exp",  "Landing Page Exp."),
    ("ad_relevance",      "Ad Relevance"),
    ("pos_organica",      "Pos. Organica GSC"),
    ("url_posizionata",   "URL Posizionata"),
    ("impressions",       "Impr."),
    ("clicks",            "Clicks"),
    ("ctr",               "CTR"),
    ("avg_cpc",           "Avg. CPC"),
    ("cost",              "Cost"),
    ("search_impr_share", "Search Impr. Share"),
    ("search_lost_is_rank", "Search Lost IS (rank)"),
]


def thin():
    s = Side(style="thin", color="BFBFBF")
    return Border(left=s, right=s, top=s, bottom=s)


def hdr(ws, row, color, txt_color="FFFFFF"):
    fill = PatternFill("solid", fgColor=color)
    for cell in ws[row]:
        if cell.value is not None:
            cell.fill = fill
            cell.font = Font(name="Arial", bold=True, color=txt_color, size=9)
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.border = thin()


def data_row(ws, row, alt=False):
    fill = PatternFill("solid", fgColor=C_GRAY if alt else C_WHITE)
    for cell in ws[row]:
        cell.fill = fill
        cell.font = Font(name="Arial", size=9)
        cell.border = thin()
        cell.alignment = Alignment(vertical="center")


def auto_width(ws):
    for col in ws.columns:
        w = max((len(str(c.value)) if c.value else 0) for c in col)
        ws.column_dimensions[get_column_letter(col[0].column)].width = min(max(w + 2, 8), 42)


def sanitize_sheet_title(title: str, max_len: int = 31) -> str:
    """Rimuove caratteri non validi per i nomi sheet Excel e tronca a 31 chars."""
    for ch in ["[", "]", ":", "*", "?", "/", "\\"]:
        title = title.replace(ch, "-")
    return title[:max_len]


def build_ads_sheet(wb, df, name, periodo, color, classif_cols=None):
    classif_cols = classif_cols or []
    # Colonne classificazione inserite dopo keyword e match_type (prime 2 di OUTPUT_COLS)
    EXTRA_COLS = [(c, c) for c in classif_cols if c in df.columns]
    # Costruisci lista colonne finale: fissi fino a match_type, poi classif, poi il resto
    FIXED_FIRST = [item for item in OUTPUT_COLS if item[0] in ("keyword_status", "keyword", "match_type")]
    FIXED_REST  = [item for item in OUTPUT_COLS if item[0] not in ("keyword_status", "keyword", "match_type")]
    ALL_COLS = FIXED_FIRST + EXTRA_COLS + FIXED_REST

    ws = wb.create_sheet(title=sanitize_sheet_title(name))
    n = len(ALL_COLS)
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=n)
    c = ws.cell(row=1, column=1)
    c.value = f"Google Ads · {name} · {periodo}"
    c.fill = PatternFill("solid", fgColor=color)
    c.font = Font(name="Arial", bold=True, color="FFFFFF", size=11)
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 22

    for j, (_, lbl) in enumerate(ALL_COLS, 1):
        ws.cell(row=2, column=j).value = lbl
    hdr(ws, 2, color)
    # Colonne classif con colore leggermente diverso per distinguerle
    if EXTRA_COLS:
        classif_start = len(FIXED_FIRST) + 1
        classif_end   = classif_start + len(EXTRA_COLS) - 1
        for jj in range(classif_start, classif_end + 1):
            ws.cell(row=2, column=jj).fill = PatternFill("solid", fgColor="7030A0")
    ws.row_dimensions[2].height = 30

    check_col = next((j for j, (k, _) in enumerate(ALL_COLS, 1) if k == "check_qs"), None)

    for i, (_, row) in enumerate(df.iterrows()):
        r = i + 3
        for j, (internal, _) in enumerate(ALL_COLS, 1):
            val = row.get(internal, "")
            if pd.isna(val):
                val = ""
            if internal == "quality_score" and val != "":
                try:
                    val = int(float(val))
                except Exception:
                    pass
            elif internal in ("cost", "avg_cpc") and val != "":
                try:
                    val = round(float(val), 2)
                except Exception:
                    pass
            elif internal == "pos_organica" and val != "":
                try:
                    val = round(float(val), 1)
                except Exception:
                    pass
            ws.cell(row=r, column=j).value = val
        data_row(ws, r, alt=(i % 2 == 1))
        if check_col:
            cell = ws.cell(row=r, column=check_col)
            if cell.value == "OK":
                cell.fill = PatternFill("solid", fgColor=C_OK)
                cell.font = Font(name="Arial", size=9, bold=True, color="276221")
            elif cell.value == "KO":
                cell.fill = PatternFill("solid", fgColor=C_KO)
                cell.font = Font(name="Arial", size=9, bold=True, color="9C0006")

    auto_width(ws)
    ws.freeze_panes = "C3"
    ws.auto_filter.ref = f"A2:{get_column_letter(n)}2"


def build_top10_sheet(wb, df_p, df_d, p_prima, p_dopo):
    ws = wb.create_sheet(title="TOP10 Kw per costo")
    headers = ["Search Keyword", "Match Type",
               f"QS {p_dopo}", f"QS {p_prima}", "Δ QS",
               f"Cost {p_dopo}", f"Cost {p_prima}"]

    def agg(df):
        d = df.copy()
        d["kk"] = d["keyword"].str.strip().str.lower()
        return d.groupby(["kk", "match_type"], dropna=False).agg(
            qs=("quality_score", "mean"), cost=("cost", "sum")).reset_index()

    ad = agg(df_d); ap = agg(df_p)
    merged = ad.merge(ap, on=["kk", "match_type"], how="outer", suffixes=("_d", "_p"))
    top10 = merged.nlargest(10, "cost_d").reset_index(drop=True)

    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(headers))
    c = ws.cell(row=1, column=1)
    c.value = f"TOP 10 Keyword per Costo · {p_dopo} vs {p_prima}"
    c.fill = PatternFill("solid", fgColor=C_HEADER_DARK)
    c.font = Font(name="Arial", bold=True, color="FFFFFF", size=11)
    c.alignment = Alignment(horizontal="center", vertical="center")
    for j, h in enumerate(headers, 1):
        ws.cell(row=2, column=j).value = h
    hdr(ws, 2, C_HEADER_DARK)

    for i, row in top10.iterrows():
        r = i + 3
        qs_d = row.get("qs_d", np.nan)
        qs_p = row.get("qs_p", np.nan)
        try:
            dq = float(qs_d) - float(qs_p)
        except Exception:
            dq = np.nan
        vals = [row["kk"].title(), row.get("match_type", ""),
                round(qs_d, 1) if pd.notna(qs_d) else "nd",
                round(qs_p, 1) if pd.notna(qs_p) else "nd",
                round(dq, 1) if pd.notna(dq) else "-",
                round(row.get("cost_d", 0), 2),
                round(row.get("cost_p", 0), 2)]
        for j, v in enumerate(vals, 1):
            ws.cell(row=r, column=j).value = v
        data_row(ws, r, alt=(i % 2 == 1))
        dc = ws.cell(row=r, column=5)
        if pd.notna(dq):
            dc.fill = PatternFill("solid", fgColor=C_OK if dq > 0 else (C_KO if dq < 0 else C_WHITE))
            dc.font = Font(name="Arial", size=9, color="276221" if dq > 0 else "9C0006")

    auto_width(ws)


def build_critiche_sheet(wb, df_d, p_dopo, qs_thr_brand=6.0, qs_thr_nobrand=6.0,
                         brand_col=None, brand_values=None):
    ws = wb.create_sheet(title="Keyword critiche")
    qs_col  = pd.to_numeric(df_d.get("quality_score", pd.Series(dtype=float)), errors="coerce")
    cost_col = df_d.get("cost", pd.Series(0.0, index=df_d.index)).fillna(0)
    if brand_col and brand_col in df_d.columns and brand_values:
        def _is_crit(row):
            val = str(row.get(brand_col, "")).strip().lower()
            is_brand = any(bv in val for bv in brand_values)
            thr = qs_thr_brand if is_brand else qs_thr_nobrand
            qs = pd.to_numeric(row.get("quality_score"), errors="coerce")
            return (not pd.isna(qs)) and float(qs) <= thr and float(row.get("cost") or 0) > 0
        mask = df_d.apply(_is_crit, axis=1)
    else:
        mask = (qs_col <= qs_thr_nobrand) & (cost_col > 0) & qs_col.notna()
    df_c = df_d[mask].sort_values("cost", ascending=False) if "cost" in df_d.columns else df_d[mask]
    thr_label = f"Brand ≤{int(qs_thr_brand)} / No-Brand ≤{int(qs_thr_nobrand)}" if brand_col else f"≤{int(qs_thr_nobrand)}"
    headers = ["Keyword", "Match Type", "Brand/No-Brand", "QS", "Check QS",
               "Exp. CTR", "Landing Page Exp.", "Ad Relevance", "Pos. GSC", "Cost", "IS"]
    cols = ["keyword", "match_type", "brand_nobrand", "quality_score", "check_qs",
            "exp_ctr", "landing_page_exp", "ad_relevance", "pos_organica", "cost", "search_impr_share"]

    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(headers))
    c = ws.cell(row=1, column=1)
    c.value = f"Keyword Critiche · QS {thr_label} con Spesa · {p_dopo}"
    c.fill = PatternFill("solid", fgColor="C00000")
    c.font = Font(name="Arial", bold=True, color="FFFFFF", size=11)
    c.alignment = Alignment(horizontal="center", vertical="center")
    for j, h in enumerate(headers, 1):
        ws.cell(row=2, column=j).value = h
    hdr(ws, 2, "C00000")

    for i, (_, row) in enumerate(df_c.iterrows()):
        r = i + 3
        for j, col in enumerate(cols, 1):
            v = row.get(col, "")
            ws.cell(row=r, column=j).value = "" if pd.isna(v) else v
        data_row(ws, r, alt=(i % 2 == 1))
        qc = ws.cell(row=r, column=5)
        if qc.value == "KO":
            qc.fill = PatternFill("solid", fgColor=C_KO)
            qc.font = Font(name="Arial", size=9, bold=True, color="9C0006")

    if df_c.empty:
        ws.cell(row=3, column=1).value = "✓ Nessuna keyword critica trovata."
    auto_width(ws)


def build_tabelle_sheet(wb, df_p, df_d, p_prima, p_dopo):
    ws = wb.create_sheet(title="TABELLE")

    sd = qs_stats(df_d); sp = qs_stats(df_p)
    ad = active_stats(df_d); ap = active_stats(df_p)
    seo_d = seo_coverage(df_d); seo_p = seo_coverage(df_p)

    ws.column_dimensions["A"].width = 34
    ws.column_dimensions["B"].width = 18
    ws.column_dimensions["C"].width = 18
    ws.column_dimensions["D"].width = 12

    def sec(ws, row, title, color=C_HEADER_DARK):
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=4)
        c = ws.cell(row=row, column=1)
        c.value = title
        c.fill = PatternFill("solid", fgColor=color)
        c.font = Font(name="Arial", bold=True, color="FFFFFF", size=10)
        c.alignment = Alignment(horizontal="left", vertical="center", indent=1)
        ws.row_dimensions[row].height = 20

    def kv(ws, row, lbl, vd, vp, fmt=None):
        ws.cell(row=row, column=1).value = lbl
        ws.cell(row=row, column=1).font = Font(name="Arial", size=9)
        ws.cell(row=row, column=1).border = thin()
        for col, val in [(2, vd), (3, vp)]:
            c = ws.cell(row=row, column=col)
            if fmt == "pct" and isinstance(val, float):
                c.value = val; c.number_format = "0.0%"
            elif fmt == "eur" and isinstance(val, float):
                c.value = val; c.number_format = '#,##0.00 "€"'
            elif fmt == "int":
                c.value = int(val) if not isinstance(val, str) else val
            elif fmt == "dec":
                c.value = round(float(val), 1) if not isinstance(val, str) else val
            else:
                c.value = val
            c.font = Font(name="Arial", size=9, bold=True)
            c.alignment = Alignment(horizontal="center")
            c.border = thin()
        try:
            dv = float(vd) - float(vp)
            dc = ws.cell(row=row, column=4)
            dc.value = round(dv, 3)
            if fmt == "pct":
                dc.number_format = "+0.0%;-0.0%;0.0%"
            dc.font = Font(name="Arial", size=9, color="276221" if dv >= 0 else "9C0006")
            dc.alignment = Alignment(horizontal="center")
            dc.border = thin()
        except Exception:
            pass

    for j, (v, clr) in enumerate([("Metrica", C_HEADER_DARK),
                                    (f"DOPO · {p_dopo}", C_HEADER_MID),
                                    (f"PRIMA · {p_prima}", C_HEADER_MID),
                                    ("Δ", C_HEADER_DARK)], 1):
        c = ws.cell(row=1, column=j)
        c.value = v
        c.fill = PatternFill("solid", fgColor=clr)
        c.font = Font(name="Arial", bold=True, color="FFFFFF", size=9)
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.border = thin()
    ws.row_dimensions[1].height = 25

    r = 2
    sec(ws, r, "📊  Quality Score"); r += 1
    kv(ws, r, "KW con QS disponibile", sd["total"], sp["total"], "int"); r += 1
    kv(ws, r, "KW con QS ≥ 7 (OK)", sd["ok"], sp["ok"], "int"); r += 1
    kv(ws, r, "KW con QS < 7 (KO)", sd["ko"], sp["ko"], "int"); r += 1
    kv(ws, r, "% KW con QS positivo (≥7)", sd["pct_ok"], sp["pct_ok"], "pct"); r += 1
    kv(ws, r, "QS medio", sd["avg"], sp["avg"], "dec"); r += 1

    r += 1
    sec(ws, r, "💰  Performance SEA (KW attive)"); r += 1
    kv(ws, r, "KW attive (impr. > 0)", ad["total"], ap["total"], "int"); r += 1
    kv(ws, r, "  → Brand", ad["brand"], ap["brand"], "int"); r += 1
    kv(ws, r, "  → No-Brand", ad["nobrand"], ap["nobrand"], "int"); r += 1
    kv(ws, r, "Costo totale (€)", ad["cost"], ap["cost"], "eur"); r += 1
    kv(ws, r, "Click totali", ad["clicks"], ap["clicks"], "int"); r += 1

    r += 1
    sec(ws, r, "🔍  Copertura SEO (GSC)"); r += 1
    kv(ws, r, "% KW con posizione organica", seo_d["pct_covered"], seo_p["pct_covered"], "pct"); r += 1
    kv(ws, r, "% KW in TOP 10 organico", seo_d["pct_top10"], seo_p["pct_top10"], "pct"); r += 1
    pos_d = seo_d["avg_pos"] if not np.isnan(seo_d["avg_pos"]) else 0.0
    pos_p = seo_p["avg_pos"] if not np.isnan(seo_p["avg_pos"]) else 0.0
    kv(ws, r, "Posizione media organica", pos_d, pos_p, "dec"); r += 1

    ws.freeze_panes = "A2"


def generate_excel(df_p, df_d, p_prima, p_dopo, qs_thr_brand=6, qs_thr_nobrand=6, classif_cols=None, brand_col=None, brand_values=None) -> bytes:
    wb = Workbook()
    wb.remove(wb.active)
    classif_cols = classif_cols or []
    build_tabelle_sheet(wb, df_p, df_d, p_prima, p_dopo)
    build_ads_sheet(wb, df_d, f"Kw {p_dopo} (DOPO)",  p_dopo,  C_HEADER_MID, classif_cols)
    build_ads_sheet(wb, df_p, f"Kw {p_prima} (PRIMA)", p_prima, "5B9BD5",     classif_cols)
    build_top10_sheet(wb, df_p, df_d, p_prima, p_dopo)
    build_critiche_sheet(wb, df_d, p_dopo, qs_thr_brand, qs_thr_nobrand, brand_col, brand_values)
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ─────────────────────────────────────────────────────────────
# STREAMLIT UI
# ─────────────────────────────────────────────────────────────

def fmt_pct(v):
    return f"{v:.1%}"

def fmt_eur(v):
    return f"€ {v:,.2f}"

def delta_color(v):
    return "normal" if v >= 0 else "inverse"


# ── SIDEBAR ───────────────────────────────────────────────────
# ─────────────────────────────────────────────
# SIDEBAR – parametri specifici di questa pagina.
# La sidebar globale (API key) è in Home.py.
# ─────────────────────────────────────────────
with st.sidebar:
    st.markdown("## 📊 SEO/SEA Analysis")
    st.divider()

    st.markdown("**Cliente**")
    cliente = st.text_input(
        "Nome cliente",
        value=st.session_state.get("seosea_cliente", "Cliente"),
        label_visibility="collapsed",
        key="seosea_cliente_input",
    )
    st.session_state["seosea_cliente"] = cliente

    st.markdown("**Periodo PRIMA**")
    periodo_prima = st.text_input(
        "Label periodo PRIMA",
        value=st.session_state.get("seosea_periodo_prima", "Periodo PRIMA"),
        label_visibility="collapsed",
        key="seosea_prima_input",
    )
    st.session_state["seosea_periodo_prima"] = periodo_prima

    st.markdown("**Periodo DOPO**")
    periodo_dopo = st.text_input(
        "Label periodo DOPO",
        value=st.session_state.get("seosea_periodo_dopo", "Periodo DOPO"),
        label_visibility="collapsed",
        key="seosea_dopo_input",
    )
    st.session_state["seosea_periodo_dopo"] = periodo_dopo

    st.markdown("---")
    st.markdown("**Screaming Frog** *(opzionale)*")
    file_sf = st.file_uploader(
        "Export internal HTML (Address + Title 1)",
        type=["xlsx", "xls", "csv"],
        key="sf",
        help="Esporta da Screaming Frog: Internal → Export. Abilita il tab URL & Title Check."
    )

    st.markdown("---")
    st.caption("File accettati: xlsx, xls, csv — header IT e EN riconosciuti automaticamente")


# ── MAIN ─────────────────────────────────────────────────────
st.markdown(f"# 📊 SEO/SEA Analysis — {cliente}")
st.divider()

# ── UPLOAD FILE ───────────────────────────────────────────────
st.markdown("### 📂 Carica i file")
st.caption("I file obbligatori sono ADS PRIMA e ADS DOPO. Gli altri arricchiscono l'analisi.")

row1_c1, row1_c2 = st.columns(2)
row2_c1, row2_c2 = st.columns(2)

with row1_c1:
    st.markdown("**1 · Google Ads PRIMA** ✱")
    file_prima = st.file_uploader(
        "Export keyword report periodo precedente",
        type=["xlsx", "xls", "csv"], key="ads_prima",
        help="Sheet 'Keyword' dell'export Google Ads. Obbligatorio."
    )

with row1_c2:
    st.markdown("**2 · Google Ads DOPO** ✱")
    file_dopo = st.file_uploader(
        "Export keyword report periodo corrente",
        type=["xlsx", "xls", "csv"], key="ads_dopo",
        help="Sheet 'Keyword' dell'export Google Ads. Obbligatorio."
    )

with row2_c1:
    st.markdown("**3 · Search Console** *(opzionale)*")
    file_gsc = st.file_uploader(
        "Export GSC con posizioni organiche per keyword",
        type=["xlsx", "xls", "csv"], key="gsc",
        help="Abilita la colonna Posizione Organica e il tab URL & Title Check."
    )

with row2_c2:
    st.markdown("**4 · Classificazione Keyword** *(opzionale)*")
    file_classif = st.file_uploader(
        "Prima colonna: keyword · Colonne successive: Brand/NoBrand, Cluster, Territorio…",
        type=["xlsx", "xls", "csv"], key="classif",
        help="Arricchisce le keyword con etichette di classificazione. Abilita soglie QS separate per Brand/No-Brand."
    )

st.caption("✱ obbligatorio")
st.divider()

# ── CONFIGURAZIONE SOGLIE QS ──────────────────────────────────
st.markdown("### ⚙️ Soglie Quality Score")
st.caption(
    "Imposta le soglie minime di QS per Brand e No-Brand. "
    "Le keyword con QS inferiore alla soglia saranno segnalate come critiche. "
    "Se carichi un file di classificazione con colonna Brand/NoBrand, "
    "le soglie verranno applicate per riga in base alla classificazione."
)

_qs_c1, _qs_c2 = st.columns(2)
with _qs_c1:
    qs_thr_brand = st.slider(
        "QS minimo Brand",
        min_value=1, max_value=10, value=7, step=1,
        help="Keyword Brand con QS inferiore a questa soglia → segnalate come critiche (KO)."
    )
with _qs_c2:
    qs_thr_nobrand = st.slider(
        "QS minimo No-Brand",
        min_value=1, max_value=10, value=6, step=1,
        help="Keyword No-Brand con QS inferiore a questa soglia → segnalate come critiche (KO)."
    )
st.divider()

# ── PROCESS ──────────────────────────────────────────────────
if file_prima and file_dopo:
    try:
        with st.spinner("Parsing file Google Ads..."):
            df_prima = parse_ads_file(file_prima)
            df_prima = classify_qs_flag(df_prima)

            df_dopo = parse_ads_file(file_dopo)
            df_dopo = classify_qs_flag(df_dopo)

            df_gsc = parse_gsc_file(file_gsc) if file_gsc else None
            df_prima = join_with_gsc(df_prima, df_gsc)
            df_dopo  = join_with_gsc(df_dopo,  df_gsc)
            df_sf = parse_screaming_frog(file_sf) if file_sf else None

        # Default variabili classificazione — sovrascritte se file caricato
        brand_col_detected = None
        brand_flag_values  = set()
        # qs_thr_brand e qs_thr_nobrand definiti dagli slider sopra

        # Avviso automatico se colonne chiave mancano nel DOPO
        missing_key = [c for c in ["quality_score", "cost", "impressions", "clicks"]
                       if c not in df_dopo.columns]
        if missing_key:
            st.warning(
                f"⚠️ Colonne non riconosciute nel file **DOPO**: `{'`, `'.join(missing_key)}`.\n\n"
                f"Colonne trovate nel file: `{'`, `'.join([str(c) for c in df_dopo.columns if not isinstance(c, float)][:20])}`\n\n"
                "Apri il pannello **🔍 Colonne riconosciute** qui sotto per il dettaglio completo."
            )

        st.success(f"✓ File caricati — {len(df_prima)} KW ({periodo_prima}) · {len(df_dopo)} KW ({periodo_dopo})")

        # Debug panel: colonne riconosciute
        with st.expander("🔍 Colonne riconosciute (debug)", expanded=False):
            known = ["keyword","keyword_status","match_type","brand_nobrand","territorio",
                     "campaign","ad_group","quality_score","check_qs","impressions","clicks",
                     "ctr","cost","avg_cpc","search_impr_share","search_lost_is_rank",
                     "pos_organica","url_posizionata","exp_ctr","landing_page_exp","ad_relevance"]
            dc1, dc2 = st.columns(2)
            with dc1:
                st.caption(f"**DOPO** — {len(df_dopo.columns)} colonne totali")
                rows = []
                for c in known:
                    rows.append({"Colonna": c, "Presente": "✅" if c in df_dopo.columns else "❌"})
                st.dataframe(pd.DataFrame(rows), hide_index=True, use_container_width=True)
            with dc2:
                st.caption(f"**PRIMA** — {len(df_prima.columns)} colonne totali")
                rows = []
                for c in known:
                    rows.append({"Colonna": c, "Presente": "✅" if c in df_prima.columns else "❌"})
                st.dataframe(pd.DataFrame(rows), hide_index=True, use_container_width=True)
            unknown = [str(c) for c in df_dopo.columns if c not in known and not isinstance(c, float)]
            st.caption("Colonne non riconosciute nel file DOPO: " + (", ".join(unknown) if unknown else "nessuna"))
        # ── KPI ROW ───────────────────────────────────────────
        st.markdown("### 📈 KPI Principali")
        sd = qs_stats(df_dopo);  sp = qs_stats(df_prima)
        ad = active_stats(df_dopo); ap = active_stats(df_prima)

        k1, k2, k3, k4, k5, k6 = st.columns(6)
        k1.metric("QS Medio (DOPO)",    f"{sd['avg']:.1f}",
                  delta=f"{sd['avg'] - sp['avg']:+.1f} vs {periodo_prima}")
        k2.metric("% QS OK (≥7)",       fmt_pct(sd['pct_ok']),
                  delta=f"{(sd['pct_ok'] - sp['pct_ok']):.1%}",
                  delta_color=delta_color(sd['pct_ok'] - sp['pct_ok']))
        k3.metric("KW critiche (QS KO)",
                  str(int((df_dopo.get("check_qs", pd.Series(dtype=str)) == "KO").sum())),
                  delta=None)
        k4.metric("KW attive (DOPO)",   str(ad['total']),
                  delta=f"{ad['total'] - ap['total']:+d}")
        k5.metric(f"Costo tot. (DOPO)", fmt_eur(ad['cost']),
                  delta=f"€{ad['cost'] - ap['cost']:+,.0f}",
                  delta_color=delta_color(-(ad['cost'] - ap['cost'])))  # costo: meno = meglio
        k6.metric("Click (DOPO)",       f"{ad['clicks']:,}",
                  delta=f"{ad['clicks'] - ap['clicks']:+,}")

        # ── TABS ──────────────────────────────────────────────
        st.divider()
        show_url_tab = (df_gsc is not None) or ("final_url" in df_dopo.columns)
        if show_url_tab:
            tab1, tab2, tab3, tab4, tab5 = st.tabs([
                f"📋 KW {periodo_dopo} (DOPO)",
                f"📋 KW {periodo_prima} (PRIMA)",
                "🏆 TOP 10 per Costo",
                "🚨 Keyword Critiche",
                "🔗 URL & Title Check",
            ])
        else:
            tab1, tab2, tab3, tab4 = st.tabs([
                f"📋 KW {periodo_dopo} (DOPO)",
                f"📋 KW {periodo_prima} (PRIMA)",
                "🏆 TOP 10 per Costo",
                "🚨 Keyword Critiche",
            ])
            tab5 = None

        # Colonne da mostrare in UI (subset leggibile)
        DISPLAY_COLS = [c for c in ["keyword", "match_type", "brand_nobrand", "territorio",
                                     "quality_score", "check_qs", "pos_organica",
                                     "impressions", "clicks", "ctr", "cost",
                                     "search_impr_share", "campaign"]
                        if c in df_dopo.columns or c in df_prima.columns]

        def style_df(df_raw, cols):
            df_show = df_raw[[c for c in cols if c in df_raw.columns]].copy()
            # Formattazione colonne
            if "ctr" in df_show.columns:
                df_show["ctr"] = df_show["ctr"].apply(lambda x: f"{float(x):.1%}" if pd.notna(x) and x != "" else "")
            if "search_impr_share" in df_show.columns:
                df_show["search_impr_share"] = df_show["search_impr_share"].apply(
                    lambda x: f"{float(x):.1%}" if pd.notna(x) and x != "" else "")
            if "cost" in df_show.columns:
                df_show["cost"] = df_show["cost"].apply(
                    lambda x: f"€ {float(x):,.2f}" if pd.notna(x) and x != "" else "")
            if "pos_organica" in df_show.columns:
                df_show["pos_organica"] = df_show["pos_organica"].apply(
                    lambda x: f"{float(x):.1f}" if pd.notna(x) and x != "" else "—")
            col_labels = {
                "keyword": "Keyword", "match_type": "Match Type", "brand_nobrand": "Brand",
                "territorio": "Territorio", "quality_score": "QS", "check_qs": "✓",
                "pos_organica": "Pos. GSC", "impressions": "Impr.", "clicks": "Clicks",
                "ctr": "CTR", "cost": "Cost", "search_impr_share": "IS",
                "campaign": "Campaign",
            }
            df_show = df_show.rename(columns={k: v for k, v in col_labels.items() if k in df_show.columns})
            return df_show

        with tab1:
            df_show = style_df(df_dopo, DISPLAY_COLS)
            # Filtri rapidi
            fc1, fc2 = st.columns([2, 2])
            with fc1:
                qs_filter = st.selectbox("Filtra QS", ["Tutti", "OK (≥7)", "KO (<7)", "N/D"], key="qs_f1")
            with fc2:
                search_kw = st.text_input("Cerca keyword", placeholder="es. vitamina c…", key="s1")

            df_f = df_dopo.copy()
            if qs_filter == "OK (≥7)":
                df_f = df_f[df_f["check_qs"] == "OK"]
            elif qs_filter == "KO (<7)":
                df_f = df_f[df_f["check_qs"] == "KO"]
            elif qs_filter == "N/D":
                df_f = df_f[df_f["check_qs"] == ""]
            if search_kw:
                df_f = df_f[df_f["keyword"].str.lower().str.contains(search_kw.lower(), na=False)]

            st.caption(f"{len(df_f)} righe")
            st.dataframe(style_df(df_f, DISPLAY_COLS), use_container_width=True, height=420)

        with tab2:
            df_show_p = style_df(df_prima, [c for c in DISPLAY_COLS if c in df_prima.columns])
            fc1p, fc2p = st.columns([2, 2])
            with fc1p:
                qs_filter_p = st.selectbox("Filtra QS", ["Tutti", "OK (≥7)", "KO (<7)", "N/D"], key="qs_f2")
            with fc2p:
                search_kw_p = st.text_input("Cerca keyword", placeholder="es. vitamina c…", key="s2")

            df_fp = df_prima.copy()
            if qs_filter_p == "OK (≥7)":
                df_fp = df_fp[df_fp["check_qs"] == "OK"]
            elif qs_filter_p == "KO (<7)":
                df_fp = df_fp[df_fp["check_qs"] == "KO"]
            elif qs_filter_p == "N/D":
                df_fp = df_fp[df_fp["check_qs"] == ""]
            if search_kw_p:
                df_fp = df_fp[df_fp["keyword"].str.lower().str.contains(search_kw_p.lower(), na=False)]

            st.caption(f"{len(df_fp)} righe")
            st.dataframe(style_df(df_fp, [c for c in DISPLAY_COLS if c in df_prima.columns]),
                         use_container_width=True, height=420)

        with tab3:
            def build_top10_df(df_p, df_d):
                def safe_agg(df):
                    d = df.copy()
                    if "keyword" not in d.columns:
                        return pd.DataFrame()
                    d["kk"] = d["keyword"].fillna("").astype(str).str.strip().str.lower()
                    group_keys = ["kk"] + (["match_type"] if "match_type" in d.columns else [])
                    agg_dict = {}
                    if "quality_score" in d.columns:
                        agg_dict["qs"] = ("quality_score", "mean")
                    if "cost" in d.columns:
                        agg_dict["cost"] = ("cost", "sum")
                    if not agg_dict:
                        return d[group_keys].drop_duplicates()
                    return d.groupby(group_keys, dropna=False).agg(**agg_dict).reset_index()

                ad = safe_agg(df_d)
                ap = safe_agg(df_p)

                if ad.empty:
                    st.info("Nessun dato disponibile per la TOP 10 (colonne keyword/cost non trovate).")
                    return pd.DataFrame()

                keys = ["kk"] + (["match_type"] if "match_type" in ad.columns and "match_type" in ap.columns else [])
                merged = ad.merge(ap, on=keys, how="outer", suffixes=("_d", "_p"))

                sort_col = "cost_d" if "cost_d" in merged.columns else (
                           "cost" if "cost" in merged.columns else None)
                top = merged.nlargest(10, sort_col).reset_index(drop=True) if sort_col else merged.head(10)

                qs_d = top["qs_d"] if "qs_d" in top.columns else pd.Series([None]*len(top))
                qs_p = top["qs_p"] if "qs_p" in top.columns else pd.Series([None]*len(top))
                cost_d = top["cost_d"] if "cost_d" in top.columns else pd.Series([None]*len(top))
                cost_p = top["cost_p"] if "cost_p" in top.columns else pd.Series([None]*len(top))

                delta_qs = []
                for i in range(len(top)):
                    try:
                        delta_qs.append(round(float(qs_d.iloc[i]) - float(qs_p.iloc[i]), 1))
                    except Exception:
                        delta_qs.append(None)

                out = pd.DataFrame({
                    "Keyword": top["kk"].str.title(),
                    f"QS {periodo_dopo}": qs_d.apply(lambda x: round(x, 1) if pd.notna(x) else "nd"),
                    f"QS {periodo_prima}": qs_p.apply(lambda x: round(x, 1) if pd.notna(x) else "nd"),
                    "Δ QS": delta_qs,
                    f"Cost {periodo_dopo}": cost_d.apply(lambda x: f"€ {x:,.2f}" if pd.notna(x) else "—"),
                    f"Cost {periodo_prima}": cost_p.apply(lambda x: f"€ {x:,.2f}" if pd.notna(x) else "—"),
                })
                if "match_type" in top.columns:
                    out.insert(1, "Match Type", top["match_type"])
                return out

            df_top10 = build_top10_df(df_prima, df_dopo)
            if not df_top10.empty:
                st.dataframe(df_top10, use_container_width=True, hide_index=True)

        with tab4:
            qs_col   = pd.to_numeric(df_dopo.get("quality_score", pd.Series(dtype=float)), errors="coerce")
            cost_col = df_dopo.get("cost", pd.Series(0.0, index=df_dopo.index)).fillna(0)
            if brand_col_detected and brand_col_detected in df_dopo.columns and brand_flag_values:
                def _crit_ui(row):
                    val = str(row.get(brand_col_detected, "")).strip().lower()
                    is_brand = any(bv in val for bv in brand_flag_values)
                    thr = qs_thr_brand if is_brand else qs_thr_nobrand
                    qs   = pd.to_numeric(row.get("quality_score"), errors="coerce")
                    cost = float(row.get("cost") or 0)
                    return (not pd.isna(qs)) and float(qs) <= thr and cost > 0
                mask = df_dopo.apply(_crit_ui, axis=1)
            else:
                mask = (qs_col <= qs_thr_nobrand) & (cost_col > 0) & qs_col.notna()
            df_crit = df_dopo[mask].sort_values("cost", ascending=False) if "cost" in df_dopo.columns else df_dopo[mask]
            thr_label = f"Brand ≤{qs_thr_brand} / No-Brand ≤{qs_thr_nobrand}" if brand_col_detected else f"≤{qs_thr_nobrand}"

            if df_crit.empty:
                st.success(f"✓ Nessuna keyword critica trovata (soglia QS {thr_label}) e costo > 0")
            else:
                st.warning(f"⚠️ {len(df_crit)} keyword critiche (QS {thr_label}) con spesa registrata")
                crit_cols = [c for c in ["keyword", "match_type", "quality_score", "check_qs",
                                          "exp_ctr", "landing_page_exp", "ad_relevance",
                                          "pos_organica", "cost", "brand_nobrand", "campaign"]
                             if c in df_crit.columns]
                st.dataframe(style_df(df_crit, crit_cols), use_container_width=True, height=380)

        # ── TAB 5: URL & TITLE CHECK ──────────────────────────
        if tab5 is not None:
            with tab5:
                if df_gsc is None:
                    st.info("Carica il file Search Console per abilitare il confronto URL GSC.")
                else:
                    with st.spinner("Analisi URL e title in corso..."):
                        df_url_check = build_url_title_check(df_dopo, df_gsc, df_sf)

                    if df_url_check.empty:
                        st.info("Nessuna keyword con URL disponibile per il confronto.")
                    else:
                        # KPI rapidi
                        has_sf_data = df_sf is not None
                        total_kw = len(df_url_check)
                        has_ads_url = (df_url_check["URL ADS"] != "").sum()
                        has_gsc_url = (df_url_check["URL GSC"] != "").sum()

                        uc1, uc2, uc3, uc4 = st.columns(4)
                        uc1.metric("KW analizzate", total_kw)
                        uc2.metric("Con URL ADS", int(has_ads_url))
                        uc3.metric("Con URL GSC", int(has_gsc_url))

                        matchable = df_url_check[
                            (df_url_check["URL ADS"] != "") & (df_url_check["URL GSC"] != "")
                        ]
                        if not matchable.empty:
                            n_match = (matchable["URL Match"] == "OK").sum()
                            uc4.metric("URL Match OK", f"{n_match}/{len(matchable)}",
                                       delta=f"{n_match/len(matchable):.0%}")

                        st.divider()

                        # Filtri
                        fc1, fc2, fc3 = st.columns([2, 2, 2])
                        with fc1:
                            url_filter = st.selectbox(
                                "Filtra URL Match",
                                ["Tutti", "OK", "KO", "N/D"],
                                key="url_f"
                            )
                        with fc2:
                            if has_sf_data:
                                title_filter = st.selectbox(
                                    "Filtra Coerenza",
                                    ["Tutti", "OK", "Parziale", "KO"],
                                    key="title_f"
                                )
                            else:
                                title_filter = "Tutti"
                                st.caption("Carica Screaming Frog per filtrare per coerenza title")
                        with fc3:
                            search_url = st.text_input("Cerca keyword", placeholder="es. vitamina c…", key="s_url")

                        df_uf = df_url_check.copy()
                        if url_filter != "Tutti":
                            df_uf = df_uf[df_uf["URL Match"] == url_filter]
                        if title_filter != "Tutti" and has_sf_data:
                            df_uf = df_uf[df_uf["Coerenza"] == title_filter]
                        if search_url:
                            df_uf = df_uf[df_uf["Keyword"].str.lower().str.contains(search_url.lower(), na=False)]

                        # Colonne da mostrare in base a disponibilità dati
                        show_cols = ["Keyword", "Match Type", "URL ADS", "URL GSC", "URL Match", "GSC generica"]
                        if has_sf_data:
                            show_cols += ["Title ADS", "Title GSC", "KW in title ADS", "KW in title GSC", "Coerenza"]

                        st.caption(f"{len(df_uf)} righe")
                        st.dataframe(
                            df_uf[show_cols],
                            use_container_width=True,
                            height=480,
                            hide_index=True,
                            column_config={
                                "URL ADS":  st.column_config.LinkColumn("URL ADS", display_text="🔗 ADS"),
                                "URL GSC":  st.column_config.LinkColumn("URL GSC", display_text="🔗 GSC"),
                                "URL Match": st.column_config.TextColumn("URL Match", width="small"),
                                "GSC generica": st.column_config.TextColumn("GSC generica", width="small"),
                                "KW in title ADS": st.column_config.TextColumn("KW ∈ Title ADS", width="small"),
                                "KW in title GSC": st.column_config.TextColumn("KW ∈ Title GSC", width="small"),
                                "Coerenza": st.column_config.TextColumn("Coerenza", width="small"),
                            }
                        )

                        if not has_sf_data:
                            st.caption("💡 Carica un export Screaming Frog nella sidebar per vedere i title e il check keyword/title.")

        # ── CLASSIFICAZIONE KEYWORD ────────────────────────────
        st.divider()
        st.markdown("### 🏷️ Classificazione Keyword")

        df_classif = None
        classif_extra_cols = []
        brand_col_detected = None
        brand_flag_values  = set()

        if file_classif:
            try:
                df_classif, classif_extra_cols, _brand_col, _cat_col = parse_classification_file(file_classif)
                # Join su entrambi i periodi
                df_dopo  = join_classification(df_dopo,  df_classif, classif_extra_cols)
                df_prima = join_classification(df_prima, df_classif, classif_extra_cols)
                # Stats match (usa prima colonna disponibile come proxy)
                _proxy = classif_extra_cols[0] if classif_extra_cols else None
                n_dopo  = int((df_dopo[_proxy]  != "").sum()) if _proxy else 0
                n_prima = int((df_prima[_proxy] != "").sum()) if _proxy else 0
                cs1, cs2, cs3 = st.columns(3)
                cs1.metric("Colonne aggiunte", len(classif_extra_cols))
                cs2.metric("Match DOPO",  f"{n_dopo}/{len(df_dopo)}")
                cs3.metric("Match PRIMA", f"{n_prima}/{len(df_prima)}")
                # Mostra colonne riconosciute con tipo
                col_labels = []
                for c in classif_extra_cols:
                    if c == INTERNAL_BRAND_COL:
                        col_labels.append(f"{c} (brand)")
                    elif c == INTERNAL_CATEGORIA_COL:
                        col_labels.append(f"{c} (categoria)")
                    else:
                        col_labels.append(c)
                st.success(f"✓ Classificazione applicata — {', '.join(col_labels)}")

                # Usa brand_col rilevato da parser
                if _brand_col and _brand_col in df_dopo.columns:
                    brand_col_detected = _brand_col
                    vals = df_dopo[_brand_col].dropna().unique()
                    brand_flag_values = {
                        str(v).strip().lower() for v in vals
                        if "brand" in str(v).lower()
                        and "no" not in str(v).lower()
                        and "non" not in str(v).lower()
                    }
            except Exception as ce:
                st.error(f"❌ Errore nel file classificazione: {ce}")

        # ── Ricalcola check_qs con soglie e colonna brand rilevata ──
        df_dopo  = classify_qs_flag(df_dopo,  qs_thr_brand, qs_thr_nobrand,
                                    brand_col_detected, brand_flag_values)
        df_prima = classify_qs_flag(df_prima, qs_thr_brand, qs_thr_nobrand,
                                    brand_col_detected, brand_flag_values)

        # ── DOWNLOAD ──────────────────────────────────────────
        st.divider()
        st.markdown("### 💾 Esporta Report Excel")

        col_dl1, col_dl2 = st.columns([2, 3])
        with col_dl1:
            with st.spinner("Generazione Excel..."):
                xlsx_bytes = generate_excel(
                    df_prima, df_dopo, periodo_prima, periodo_dopo,
                    qs_thr_brand, qs_thr_nobrand, classif_extra_cols,
                    brand_col_detected, brand_flag_values
                )

            filename = f"analisi_seo_sea_{cliente.lower().replace(' ', '_')}_{periodo_dopo.lower().replace(' ', '_')}.xlsx"
            st.download_button(
                label="⬇ Scarica Report Excel",
                data=xlsx_bytes,
                file_name=filename,
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
        with col_dl2:
            st.markdown("**Il file Excel include:**")
            st.markdown("📊 **TABELLE** — KPI riepilogo con delta PRIMA/DOPO")
            st.markdown("📋 **KW DOPO / PRIMA** — tutti i dati con colori QS e filtri")
            st.markdown("🏆 **TOP10 per costo** — con delta Quality Score")
            st.markdown("🚨 **Keyword critiche** — QS basso con spesa attiva")
            if classif_extra_cols:
                st.markdown(f"🏷️ **Classificazione** — colonne: {', '.join(classif_extra_cols)}")

    except Exception as e:
        st.error(f"❌ Errore durante il processing: {e}")
        with st.expander("Dettagli errore"):
            import traceback
            st.code(traceback.format_exc())

else:
    st.info("📂 Carica almeno i file Google Ads PRIMA e DOPO per generare l'analisi. Il file GSC è opzionale.")
