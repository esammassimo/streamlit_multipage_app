"""
23_llm_brand_monitor.py — LLM Brand Monitor
=============================================
Monitora la visibilità brand su LLM e AI Features.
Esporta i risultati in Excel nel formato standard del progetto
(Risposte / Brand / Fonti - Apps Script).

Dipendenze: già in requirements.txt (requests, openpyxl, pandas,
rapidfuzz, openai, anthropic, google-search-results).
La OpenAI API key viene letta da st.session_state['openai_api_key']
(sidebar globale dell'app). Le altre chiavi si inseriscono in questa pagina.
"""
from __future__ import annotations

import io
import re
import time
import json
import logging
from datetime import date
from typing import Callable, Optional

import pandas as pd
import requests
import streamlit as st
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from rapidfuzz import process, fuzz

log = logging.getLogger(__name__)

# ═══════════════════════════════════════════════════════════════════════════════
# ENGINE — chiamate API (no DB)
# ═══════════════════════════════════════════════════════════════════════════════

SYSTEM_PROMPT = (
    "Sei un esperto del settore. Rispondi in modo dettagliato e completo alla domanda, "
    "menzionando brand, aziende e servizi specifici quando rilevante. "
    "Includi fonti e URL quando possibile."
)

GEMINI_FALLBACK = ["gemini-2.5-flash", "gemini-2.0-flash"]

AVAILABLE_MODELS = {
    "ChatGPT":    ["gpt-4o", "gpt-4o-mini", "gpt-5.4", "gpt-5.5"],
    "Claude":     ["claude-sonnet-4-6", "claude-haiku-4-5-20251001", "claude-opus-4-8"],
    "Gemini":     ["gemini-2.5-flash", "gemini-3.5-flash", "gemini-2.5-pro"],
    "Perplexity": ["sonar-pro", "sonar", "sonar-reasoning-pro"],
}


def _call_chatgpt(question: str, keys: dict, model: str = "gpt-4o") -> tuple[str, list, str]:
    k = keys.get("openai", "")
    if not k:
        return "DISABLED", [], model
    r = requests.post(
        "https://api.openai.com/v1/chat/completions",
        headers={"Authorization": f"Bearer {k}", "Content-Type": "application/json"},
        json={"model": model,
              "messages": [{"role": "system", "content": SYSTEM_PROMPT},
                           {"role": "user", "content": question}],
              "max_tokens": 2000, "temperature": 0.7},
        timeout=60,
    )
    r.raise_for_status()
    text = r.json()["choices"][0]["message"]["content"]
    return text, re.findall(r'https?://[^\s\)\]\>\"\']+', text), model


def _call_claude(question: str, keys: dict, model: str = "claude-sonnet-4-6") -> tuple[str, list, str]:
    k = keys.get("anthropic", "")
    if not k:
        return "DISABLED", [], model
    r = requests.post(
        "https://api.anthropic.com/v1/messages",
        headers={"x-api-key": k, "anthropic-version": "2023-06-01",
                 "Content-Type": "application/json"},
        json={"model": model, "max_tokens": 2000, "system": SYSTEM_PROMPT,
              "messages": [{"role": "user", "content": question}]},
        timeout=60,
    )
    r.raise_for_status()
    text = r.json()["content"][0]["text"]
    return text, re.findall(r'https?://[^\s\)\]\>\"\']+', text), model


def _call_gemini(question: str, keys: dict, model: str | None = None) -> tuple[str, list, str]:
    k = keys.get("google", "")
    if not k:
        return "DISABLED", [], model or GEMINI_FALLBACK[0]
    models_to_try = [model] + [m for m in GEMINI_FALLBACK if m != model] if model else GEMINI_FALLBACK
    last_exc: Exception = Exception("No model tried")
    for m in models_to_try:
        try:
            r = requests.post(
                f"https://generativelanguage.googleapis.com/v1beta/models/{m}:generateContent?key={k}",
                headers={"Content-Type": "application/json"},
                json={"contents": [{"parts": [{"text": f"{SYSTEM_PROMPT}\n\n{question}"}]}],
                      "generationConfig": {"temperature": 0.7, "maxOutputTokens": 2000},
                      "tools": [{"google_search": {}}]},
                timeout=60,
            )
            if r.status_code in (400, 404):
                last_exc = Exception(f"HTTP {r.status_code}")
                continue
            r.raise_for_status()
            data = r.json()
            parts = data.get("candidates", [{}])[0].get("content", {}).get("parts", [])
            text = "\n".join(p.get("text", "") for p in parts).strip()
            chunks = (data.get("candidates", [{}])[0]
                      .get("groundingMetadata", {})
                      .get("groundingChunks", []))
            sources = [c.get("web", {}).get("uri", "") for c in chunks
                       if c.get("web", {}).get("uri")]
            return text, sources, m
        except Exception as exc:
            last_exc = exc
    return f"ERROR: {last_exc}", [], models_to_try[-1]


def _call_perplexity(question: str, keys: dict, model: str = "sonar-pro") -> tuple[str, list, str]:
    k = keys.get("pplx", "")
    if not k:
        return "DISABLED", [], model
    r = requests.post(
        "https://api.perplexity.ai/chat/completions",
        headers={"Authorization": f"Bearer {k}", "Content-Type": "application/json"},
        json={"model": model,
              "messages": [{"role": "system", "content": SYSTEM_PROMPT},
                           {"role": "user", "content": question}],
              "max_tokens": 2000},
        timeout=60,
    )
    r.raise_for_status()
    data = r.json()
    text = data["choices"][0]["message"]["content"]
    sources = data.get("citations", []) or re.findall(r'https?://[^\s\)\]\>\"\']+', text)
    return text, sources, model


def _parse_blocks(blocks: list) -> str:
    parts = []
    for b in blocks:
        s = b.get("snippet", "").strip()
        if s:
            parts.append(s)
        for item in b.get("list", []):
            t = item.get("snippet", "").strip()
            if t:
                parts.append(t)
    return "\n".join(parts).strip()


def _call_aio(question: str, keys: dict, language: str = "it", country: str = "it") -> tuple[str, list, str]:
    k = keys.get("serpapi", "")
    if not k:
        return "DISABLED", [], "google_aio"
    r = requests.get("https://serpapi.com/search",
                     params={"engine": "google", "q": question, "api_key": k,
                             "hl": language, "gl": country, "no_cache": "true"},
                     timeout=60)
    r.raise_for_status()
    aio = r.json().get("ai_overview")
    if not aio:
        return "", [], "google_aio"
    token = aio.get("page_token")
    if token:
        r2 = requests.get("https://serpapi.com/search",
                          params={"engine": "google_ai_overview", "page_token": token,
                                  "api_key": k, "no_cache": "true"}, timeout=60)
        r2.raise_for_status()
        aio = r2.json().get("ai_overview", r2.json())
    text = _parse_blocks(aio.get("text_blocks", [])) or (aio.get("text") or "").strip()
    sources = [ref["link"] for ref in aio.get("references", []) if ref.get("link")]
    return text, sources, "google_aio"


def _call_aim(question: str, keys: dict, language: str = "it", country: str = "it") -> tuple[str, list, str]:
    k = keys.get("serpapi", "")
    if not k:
        return "DISABLED", [], "google_aim"
    r = requests.get("https://serpapi.com/search",
                     params={"engine": "google_ai_mode", "q": question, "api_key": k,
                             "hl": language, "gl": country, "no_cache": "true"},
                     timeout=60)
    r.raise_for_status()
    data = r.json()
    text = _parse_blocks(data.get("text_blocks", [])) or data.get("reconstructed_markdown", "").strip()
    sources = [ref["link"] for ref in data.get("references", []) if ref.get("link")]
    return text, sources, "google_aim"


def _is_valid(text: str) -> bool:
    return (bool(text)
            and not str(text).startswith(("ERROR:", "DISABLED"))
            and len(str(text).strip()) > 20)


# ─── Brand extraction ────────────────────────────────────────────────────────

_SW = {
    "il","lo","la","i","gli","le","un","una","del","della","dei","delle","degli",
    "al","alla","ai","alle","nel","nella","nei","nelle","sul","sulla","sui","sulle",
    "dal","dalla","dai","dalle","con","per","tra","fra","che","chi","cui","non","ma",
    "se","come","quando","dove","però","quindi","così","anche","già","ancora","sempre",
    "mai","molto","poco","tutto","questo","questa","questi","queste","the","a","an",
    "in","on","at","to","for","of","and","or","but","is","are","was","were","be",
    "with","by","from","as","it","its","this","that","also",
}

BRAND_PROMPT = (
    "Extract all brand names, company names, and product names from the text below.\n"
    "Normalize to most common short form (e.g. 'Nike Inc.' → 'Nike').\n"
    "Assign position = ordinal of first mention (1 = first).\n"
    "Return ONLY a valid JSON array, no markdown.\n"
    "Example: [{\"name\": \"Nike\", \"position\": 1}]\nText:\n{text}"
)


BRAND_METHOD_LABELS = {
    "regex":         "Regex (gratuito)",
    "llm_openai":    "LLM — GPT-4o-mini (OpenAI)",
    "llm_anthropic": "LLM — Claude Sonnet (Anthropic)",
    "ensemble":      "Ensemble regex + LLM (con confidenza)",
}


def _brand_method_options(keys: dict) -> list[str]:
    """Restituisce i metodi disponibili in base alle key configurate."""
    opts = ["regex"]
    if keys.get("openai"):
        opts.append("llm_openai")
    if keys.get("anthropic"):
        opts.append("llm_anthropic")
    if len(opts) > 1:
        opts.append("ensemble")
    return opts


def _brands_regex(text: str) -> list[dict]:
    found, seen, pos = [], set(), 1
    for b in re.findall(r'\*\*([A-Z][A-Za-zÀ-ÿ\s\'&\-\.]+?)\*\*', text):
        b = b.strip()
        if len(b) >= 3 and b.lower() not in seen:
            seen.add(b.lower())
            found.append({"name": b, "position": pos})
            pos += 1
    pattern = r'\b([A-Z][a-zA-ZÀ-ÖØ-öø-ÿ0-9&\-\'\.]{1,}(?:\s+[A-Z][a-zA-ZÀ-ÖØ-öø-ÿ0-9&\-\'\.]+){0,3})\b'
    for b in re.findall(pattern, text):
        b = b.strip().rstrip(".")
        tokens = b.split()
        if (len(tokens) > 4 or all(t.lower() in _SW for t in tokens)
                or len(b) < 3
                or b.lower() in {"http", "https", "www", "com", "org", "net", "url", "api"}
                or b.lower() in seen):
            continue
        seen.add(b.lower())
        found.append({"name": b, "position": pos})
        pos += 1
    return found


def _brands_llm_openai(text: str, openai_key: str) -> list[dict]:
    if not openai_key:
        return _brands_regex(text)
    try:
        r = requests.post(
            "https://api.openai.com/v1/chat/completions",
            headers={"Authorization": f"Bearer {openai_key}",
                     "Content-Type": "application/json"},
            json={"model": "gpt-4o-mini",
                  "messages": [{"role": "user",
                                "content": BRAND_PROMPT.format(text=text[:8000])}],
                  "max_tokens": 1000, "temperature": 0},
            timeout=30,
        )
        r.raise_for_status()
        raw = re.sub(r'^```(?:json)?\s*|\s*```$', '',
                     r.json()["choices"][0]["message"]["content"].strip())
        return [{"name": b.get("name", ""), "position": b.get("position", i+1)}
                for i, b in enumerate(json.loads(raw)) if b.get("name")]
    except Exception:
        return _brands_regex(text)


def _brands_llm_anthropic(text: str, anthropic_key: str,
                          model: str = "claude-sonnet-4-6") -> list[dict]:
    if not anthropic_key:
        return _brands_regex(text)
    try:
        r = requests.post(
            "https://api.anthropic.com/v1/messages",
            headers={"x-api-key": anthropic_key, "anthropic-version": "2023-06-01",
                     "Content-Type": "application/json"},
            json={"model": model, "max_tokens": 1000,
                  "messages": [{"role": "user",
                                "content": BRAND_PROMPT.format(text=text[:8000])}]},
            timeout=30,
        )
        r.raise_for_status()
        raw = re.sub(r'^```(?:json)?\s*|\s*```$', '',
                     r.json()["content"][0]["text"].strip())
        return [{"name": b.get("name", ""), "position": b.get("position", i+1)}
                for i, b in enumerate(json.loads(raw)) if b.get("name")]
    except Exception:
        return _brands_regex(text)


def _normalize(brands: list[dict], known: list[str], threshold: int = 85) -> list[dict]:
    if not known or not brands:
        return brands
    norm, seen = [], set()
    for b in brands:
        res = process.extractOne(b["name"], known, scorer=fuzz.token_sort_ratio)
        name = res[0] if res and res[1] >= threshold else b["name"]
        k = name.lower()
        if k not in seen:
            seen.add(k)
            norm.append({**b, "name": name})
    return norm


def _extract_brands_ensemble(
    text: str, keys: dict, known: list[str], threshold: int = 85,
    engines: tuple = ("regex", "llm_openai", "llm_anthropic"),
) -> list[dict]:
    """Combina più metodi di estrazione sullo stesso testo e assegna a ogni
    brand un punteggio di confidenza pari al numero di fonti che lo hanno
    trovato (utile per distinguere estrazioni robuste da falsi positivi)."""
    found: dict[str, dict] = {}
    runs: list[tuple[str, list[dict]]] = []
    if "regex" in engines:
        runs.append(("regex", _brands_regex(text)))
    if "llm_openai" in engines and keys.get("openai"):
        runs.append(("llm_openai", _brands_llm_openai(text, keys["openai"])))
    if "llm_anthropic" in engines and keys.get("anthropic"):
        runs.append(("llm_anthropic", _brands_llm_anthropic(text, keys["anthropic"])))

    for source, brands in runs:
        for b in (_normalize(brands, known, threshold) if known else brands):
            key = b["name"].lower()
            if key not in found:
                found[key] = {"name": b["name"], "positions": [], "sources": set()}
            found[key]["positions"].append(b["position"])
            found[key]["sources"].add(source)

    result = [
        {
            "name": v["name"],
            "position": min(v["positions"]),
            "confidence": len(v["sources"]),
            "sources": ", ".join(sorted(v["sources"])),
        }
        for v in found.values()
    ]
    result.sort(key=lambda x: x["position"])
    return result


def extract_brands(
    text: str, keys: dict, known: list[str],
    method: str = "regex", threshold: int = 85,
    engines: tuple = ("regex", "llm_openai", "llm_anthropic"),
) -> list[dict]:
    if method == "ensemble":
        return _extract_brands_ensemble(text, keys, known, threshold, engines)

    if method in ("llm_openai", "llm"):  # "llm" resta come alias retro-compatibile
        brands = _brands_llm_openai(text, keys.get("openai", ""))
    elif method == "llm_anthropic":
        brands = _brands_llm_anthropic(text, keys.get("anthropic", ""))
    else:
        brands = _brands_regex(text)

    normalized = _normalize(brands, known, threshold) if known else brands
    for b in normalized:
        b.setdefault("confidence", 1)
        b.setdefault("sources", method)
    return normalized


def reclassify_brands(
    risposte: list[dict],
    keys: dict,
    known: list[str],
    method: str = "regex",
    threshold: int = 85,
    engines: tuple = ("regex", "llm_openai", "llm_anthropic"),
    progress_cb: Optional[Callable] = None,
) -> list[dict]:
    """Ri-estrae i brand dalle risposte già raccolte (in memoria), senza
    richiamare di nuovo le API degli LLM che hanno generato le risposte.
    Utile per confrontare metodologie di estrazione diverse (regex / LLM
    OpenAI / LLM Anthropic / ensemble) o soglie di normalizzazione diverse
    sullo stesso dataset di risposte."""
    valid_rows = [r for r in risposte if r.get("Risposta")]
    meta_cols = ("Data", "AI Questions", "Keyword", "Cluster", "Subcluster",
                 "Volume", "Intent", "Tone")
    brand_rows: list[dict] = []
    total = len(valid_rows)
    for i, r in enumerate(valid_rows, 1):
        meta = {k: r.get(k, "") for k in meta_cols}
        for b in extract_brands(r["Risposta"], keys, known, method, threshold, engines):
            brand_rows.append({
                **meta, "LLM": r.get("LLM", ""), "Model": r.get("Model", ""),
                "Brand": b["name"], "Position": b["position"],
                "Confidence": b.get("confidence", 1), "Sources": b.get("sources", method),
            })
        if progress_cb:
            try:
                progress_cb(i, total)
            except Exception:
                pass
    return brand_rows


# ─── Main run ────────────────────────────────────────────────────────────────

def run_monitor(
    questions: list[dict],
    keys: dict,
    config: dict,
    known_brands: list[str],
    brand_method: str,
    language: str,
    country: str,
    progress_cb: Optional[Callable] = None,
    brand_threshold: int = 85,
) -> dict:
    llms       = config.get("llms", [])
    ai_feats   = config.get("ai_features", [])
    iterations = int(config.get("iterations", 1))
    models     = config.get("models", {})
    today      = date.today().isoformat()

    risposte, brand_rows, fonti_rows = [], [], []
    total = len(questions) * (len(llms) * iterations + len(ai_feats))
    done  = 0

    def _meta(q: dict) -> dict:
        return {
            "Data": today,
            "AI Questions": q.get("question", ""),
            "Keyword": q.get("keyword", ""),
            "Cluster": q.get("cluster", ""),
            "Subcluster": q.get("subcluster", ""),
            "Volume": q.get("volume", ""),
            "Intent": q.get("intent", ""),
            "Tone": q.get("tone", ""),
        }

    def _store(q: dict, llm_label: str, model_name: str,
               text: str, sources: list, elapsed: float):
        nonlocal done
        meta = _meta(q)
        valid = _is_valid(text)

        risposte.append({**meta, "LLM": llm_label, "Model": model_name,
                         "Risposta": text if valid else ""})
        if valid:
            for b in extract_brands(text, keys, known_brands, brand_method, brand_threshold):
                brand_rows.append({**meta, "LLM": llm_label, "Model": model_name,
                                   "Brand": b["name"], "Position": b["position"],
                                   "Confidence": b.get("confidence", 1),
                                   "Sources": b.get("sources", brand_method)})
            for url in sources:
                fonti_rows.append({**meta, "LLM": llm_label, "Model": model_name, "URL": url})

        done += 1
        if progress_cb:
            try:
                progress_cb(done, total, {
                    "llm": llm_label, "model": model_name,
                    "question": q.get("question", "")[:70],
                    "valid": valid, "text": text,
                    "sources": sources, "elapsed": elapsed,
                })
            except Exception:
                pass

    for q in questions:
        question = q.get("question", "")

        for llm_name in llms:
            sel = models.get(llm_name, "")
            for _ in range(iterations):
                t0 = time.time()
                try:
                    if llm_name == "ChatGPT":
                        txt, src, mn = _call_chatgpt(question, keys, sel or "gpt-4o")
                    elif llm_name == "Claude":
                        txt, src, mn = _call_claude(question, keys, sel or "claude-sonnet-4-6")
                    elif llm_name == "Gemini":
                        txt, src, mn = _call_gemini(question, keys, sel or None)
                    elif llm_name == "Perplexity":
                        txt, src, mn = _call_perplexity(question, keys, sel or "sonar-pro")
                    else:
                        txt, src, mn = f"ERROR: unknown LLM {llm_name}", [], ""
                except Exception as exc:
                    txt, src, mn = f"ERROR: {exc}", [], sel or ""
                _store(q, llm_name, mn, txt, src, round(time.time() - t0, 1))
                time.sleep(1)

        serp_q = q.get("keyword", "").strip() or question
        for feat in ai_feats:
            t0 = time.time()
            try:
                if feat == "AI Overviews":
                    txt, src, mn = _call_aio(serp_q, keys, language, country)
                elif feat == "AI Mode":
                    txt, src, mn = _call_aim(serp_q, keys, language, country)
                else:
                    txt, src, mn = f"ERROR: unknown feature {feat}", [], ""
            except Exception as exc:
                txt, src, mn = f"ERROR: {exc}", [], feat
            _store(q, feat, mn, txt, src, round(time.time() - t0, 1))
            time.sleep(1.5)

    return {"risposte": risposte, "brand": brand_rows, "fonti": fonti_rows}


# ═══════════════════════════════════════════════════════════════════════════════
# EXPORT EXCEL — formato standard del progetto
# ═══════════════════════════════════════════════════════════════════════════════

def _hstyle() -> dict:
    return {
        "font":      Font(name="Arial", bold=True, color="FFFFFF", size=10),
        "fill":      PatternFill("solid", start_color="1F4E79"),
        "alignment": Alignment(horizontal="center", vertical="center", wrap_text=True),
        "border":    Border(bottom=Side(style="thin", color="FFFFFF"),
                            right=Side(style="thin", color="FFFFFF")),
    }


def _dstyle(even: bool) -> dict:
    return {
        "font":      Font(name="Arial", size=9),
        "fill":      PatternFill("solid", start_color="EBF3FB" if even else "FFFFFF"),
        "alignment": Alignment(vertical="top", wrap_text=False),
    }


def _write_sheet(ws, df: pd.DataFrame, widths: dict, wrap_col: str | None = None):
    hs = _hstyle()
    for ci, col in enumerate(df.columns, 1):
        cell = ws.cell(row=1, column=ci, value=col)
        for attr, val in hs.items():
            setattr(cell, attr, val)
    for ri, (_, row) in enumerate(df.iterrows(), 2):
        ds = _dstyle(ri % 2 == 0)
        for ci, col in enumerate(df.columns, 1):
            val = row[col]
            if pd.isna(val):
                val = ""
            cell = ws.cell(row=ri, column=ci, value=val)
            for attr, v in ds.items():
                setattr(cell, attr, v)
            if wrap_col and col == wrap_col:
                cell.alignment = Alignment(vertical="top", wrap_text=True)
    ws.freeze_panes = ws["A2"]
    ws.auto_filter.ref = ws.dimensions
    ws.row_dimensions[1].height = 30
    for ci, col in enumerate(df.columns, 1):
        ws.column_dimensions[get_column_letter(ci)].width = widths.get(col, 18)


def build_excel(results: dict, questions: list[dict], run_date) -> bytes:
    wb = Workbook()
    COMMON = {"Data": 12, "AI Questions": 42, "Keyword": 22, "Cluster": 18,
              "Subcluster": 16, "Volume": 10, "LLM": 14, "Model": 22,
              "Intent": 14, "Tone": 12}

    def _prep(rows: list, extra: list[str]) -> pd.DataFrame:
        df = pd.DataFrame(rows) if rows else pd.DataFrame()
        cols = (["Data", "AI Questions", "Keyword", "Cluster", "Subcluster", "Volume",
                 "LLM", "Model"] + extra + ["Intent", "Tone"])
        for c in cols:
            if c not in df.columns:
                df[c] = ""
        return df[cols].fillna("")

    # Risposte
    ws_r = wb.active
    ws_r.title = "Risposte - Apps Script"
    df_r = _prep(results["risposte"], ["Risposta"])
    _write_sheet(ws_r, df_r, {**COMMON, "Risposta": 65}, wrap_col="Risposta")
    for ri in range(2, len(df_r) + 2):
        ws_r.row_dimensions[ri].height = 55

    # Brand
    ws_b = wb.create_sheet("Brand - Apps Script")
    df_b = _prep(results["brand"], ["Brand", "Position"])
    _write_sheet(ws_b, df_b, {**COMMON, "Brand": 28, "Position": 10})
    for ri in range(2, len(df_b) + 2):
        ws_b.row_dimensions[ri].height = 15

    # Fonti
    ws_f = wb.create_sheet("Fonti - Apps Script")
    df_f = _prep(results["fonti"], ["URL"])
    _write_sheet(ws_f, df_f, {**COMMON, "URL": 65})
    for ri in range(2, len(df_f) + 2):
        ws_f.row_dimensions[ri].height = 15

    # AI Questions (input)
    ws_q = wb.create_sheet("AI Questions")
    df_q = (pd.DataFrame(questions)
            .rename(columns={"question": "AI Questions", "keyword": "Keyword",
                             "cluster": "Cluster", "subcluster": "Subcluster",
                             "volume": "Volume", "intent": "Intent", "tone": "Tone"})
            .fillna(""))
    if df_q.empty:
        df_q = pd.DataFrame(columns=["AI Questions", "Keyword", "Cluster",
                                     "Subcluster", "Volume", "Intent", "Tone"])
    _write_sheet(ws_q, df_q,
                 {"AI Questions": 50, "Keyword": 25, "Cluster": 20,
                  "Subcluster": 18, "Volume": 10, "Intent": 14, "Tone": 12})
    for ri in range(2, len(df_q) + 2):
        ws_q.row_dimensions[ri].height = 15

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ═══════════════════════════════════════════════════════════════════════════════
# UI
# ═══════════════════════════════════════════════════════════════════════════════

st.title("📡 LLM Brand Monitor")
st.caption(
    "Monitora la visibilità brand su LLM e AI Features. "
    "Esporta i risultati in Excel nel formato standard del progetto."
)

# ─── Session State Init ───────────────────────────────────────────────────────
_DEFAULTS = {
    "lbm_questions": [],
    "lbm_keys":      {},
    "lbm_config": {
        "llms":        ["ChatGPT", "Claude", "Gemini", "Perplexity"],
        "ai_features": ["AI Overviews", "AI Mode"],
        "iterations":  1,
        "models": {
            "ChatGPT":    "gpt-4o",
            "Claude":     "claude-sonnet-4-6",
            "Gemini":     "gemini-2.5-flash",
            "Perplexity": "sonar-pro",
        },
    },
    "lbm_results":   None,
    "lbm_known":     [],
    "lbm_lang":      "it",
    "lbm_country":   "it",
    "lbm_bmethod":   "regex",
    "lbm_threshold": 85,
}
for k, v in _DEFAULTS.items():
    if k not in st.session_state:
        st.session_state[k] = v

# Eredita OpenAI key dalla sidebar globale dell'app SOLO se non è stata
# inserita direttamente in questa pagina (il campo locale ha priorità)
if not st.session_state.lbm_keys.get("openai") and st.session_state.get("openai_api_key"):
    st.session_state.lbm_keys["openai"] = st.session_state["openai_api_key"]

# ─── Tabs ─────────────────────────────────────────────────────────────────────
tab_input, tab_cfg, tab_run, tab_preview, tab_export = st.tabs(
    ["📋 Input", "⚙️ Configurazione", "🚀 Esecuzione", "🔍 Anteprima & Riclassifica", "📥 Export"]
)

# ══════════════════════════════════════════════════════════════════════════════
# TAB 1 — INPUT
# ══════════════════════════════════════════════════════════════════════════════
with tab_input:
    st.subheader("Keyword e Domande")

    mode = st.radio("Modalità", ["✏️ Manuale", "📂 Importa file"], horizontal=True)

    DISPLAY = {
        "question": "AI Questions", "keyword": "Keyword", "cluster": "Cluster",
        "subcluster": "Subcluster", "volume": "Volume", "intent": "Intent", "tone": "Tone",
    }

    if mode == "✏️ Manuale":
        st.caption(
            "Una riga per domanda. "
            "Formato: `Domanda | Keyword | Cluster | Subcluster | Volume | Intent | Tone`"
        )
        txt = st.text_area(
            "Domande", height=250,
            placeholder=(
                "Che differenza c'è tra TAN e TAEG? | tan | Informativa | Prestiti\n"
                "Chi concede prestiti a tutti? | prestiti"
            ),
        )
        if st.button("✅ Carica", type="primary", disabled=not txt.strip()):
            rows = []
            for line in txt.strip().splitlines():
                if not line.strip():
                    continue
                p = [x.strip() for x in line.split("|")]
                keys_ord = ["question", "keyword", "cluster", "subcluster",
                            "volume", "intent", "tone"]
                row = {k: (p[i] if len(p) > i else "") for i, k in enumerate(keys_ord)}
                if row["question"]:
                    rows.append(row)
            st.session_state.lbm_questions = rows
            st.success(f"✅ {len(rows)} domande caricate.")

    else:
        st.caption(
            "Accetta Excel (.xlsx) o CSV. "
            "Rileva automaticamente il foglio **AI Questions** se presente."
        )
        uploaded = st.file_uploader("Carica file", type=["xlsx", "xls", "csv"])
        if uploaded:
            try:
                if uploaded.name.endswith(".csv"):
                    df = pd.read_csv(uploaded)
                else:
                    xl = pd.ExcelFile(uploaded)
                    if "AI Questions" in xl.sheet_names:
                        sheet = "AI Questions"
                    else:
                        sheet = st.selectbox("Seleziona foglio", xl.sheet_names)
                    df = pd.read_excel(xl, sheet_name=sheet)

                st.dataframe(df.head(3), use_container_width=True, hide_index=True)
                available = ["(nessuna)"] + list(df.columns)

                def _auto(hint: str) -> str:
                    return next((c for c in df.columns if hint in c.lower()), "(nessuna)")

                defs = {
                    "question": _auto("question"), "keyword": _auto("keyword"),
                    "cluster": _auto("cluster"), "subcluster": _auto("sub"),
                    "volume": _auto("volume"), "intent": _auto("intent"),
                    "tone": _auto("tone"),
                }
                import_map_cols = st.columns(2)
                col_map = {}
                for idx, (key, label) in enumerate(DISPLAY.items()):
                    with import_map_cols[idx % 2]:
                        di = available.index(defs[key]) if defs[key] in available else 0
                        col_map[key] = st.selectbox(label, available, index=di,
                                                    key=f"lbm_cmap_{key}")

                if st.button("✅ Importa", type="primary"):
                    if col_map["question"] == "(nessuna)":
                        st.error("Seleziona la colonna AI Questions.")
                    else:
                        rows = []
                        for _, row in df.iterrows():
                            q = str(row[col_map["question"]]) if col_map["question"] != "(nessuna)" else ""
                            if not q or q == "nan":
                                continue
                            rows.append({
                                k: (str(row[col_map[k]]) if col_map[k] != "(nessuna)" else "")
                                for k in DISPLAY
                            })
                        st.session_state.lbm_questions = rows
                        st.success(f"✅ {len(rows)} domande importate.")
            except Exception as e:
                st.error(f"Errore lettura file: {e}")

    if st.session_state.lbm_questions:
        st.divider()
        st.caption(f"**{len(st.session_state.lbm_questions)} domande caricate**")
        df_prev = (pd.DataFrame(st.session_state.lbm_questions)
                   .rename(columns=DISPLAY)
                   .replace("nan", "")
                   .fillna(""))
        st.dataframe(df_prev, use_container_width=True, hide_index=True)
        if st.button("🗑️ Cancella tutto"):
            st.session_state.lbm_questions = []
            st.rerun()

    # Brand list
    st.divider()
    st.subheader("Brand list (opzionale)")
    st.caption("Un brand per riga. Usato per normalizzare i brand estratti via fuzzy match (RapidFuzz, soglia 85%).")
    brands_txt = st.text_area(
        "Brand noti", height=120,
        value="\n".join(st.session_state.lbm_known),
        placeholder="Compass\nFindomestic\nAgos\nYounited Credit",
    )
    if st.button("💾 Salva brand list"):
        st.session_state.lbm_known = [b.strip() for b in brands_txt.splitlines() if b.strip()]
        st.success(f"{len(st.session_state.lbm_known)} brand salvati.")

# ══════════════════════════════════════════════════════════════════════════════
# TAB 2 — CONFIGURAZIONE
# ══════════════════════════════════════════════════════════════════════════════
with tab_cfg:
    st.subheader("🔑 API Keys")
    st.caption(
        "La chiave OpenAI può arrivare dalla sidebar globale dell'app oppure "
        "essere inserita qui sotto: il valore inserito in questa pagina ha sempre la priorità."
    )

    ALL_KEYS = [
        ("openai",    "OpenAI (ChatGPT)",            "sk-..."),
        ("anthropic", "Anthropic (Claude)",          "sk-ant-..."),
        ("google",    "Google AI (Gemini)",          "AIza..."),
        ("pplx",      "Perplexity",                  "pplx-..."),
        ("serpapi",   "SerpAPI (AI Overviews/Mode)", "serpapi_key..."),
    ]
    cfg_key_cols = st.columns(2)
    for idx, (key, label, ph) in enumerate(ALL_KEYS):
        with cfg_key_cols[idx % 2]:
            cur = st.session_state.lbm_keys.get(key, "")
            source = ""
            if key == "openai" and not cur and st.session_state.get("openai_api_key"):
                cur = st.session_state["openai_api_key"]
                source = " (dalla sidebar)"
            st.caption(("✅" if cur else "❌") + f" {label}{source}" +
                       (f" `{cur[:6]}…{cur[-4:]}`" if cur else ""))
            val = st.text_input(label, type="password", placeholder=ph,
                                key=f"lbm_key_{key}")
            if val.strip():
                st.session_state.lbm_keys[key] = val.strip()

    st.divider()
    st.subheader("🤖 Piattaforme e Modelli")

    cfg = st.session_state.lbm_config
    c_llm, c_ai = st.columns(2)
    with c_llm:
        sel_llms = st.multiselect("LLM", list(AVAILABLE_MODELS.keys()),
                                  default=cfg["llms"], key="lbm_sel_llms")
    with c_ai:
        sel_ai = st.multiselect("AI Features", ["AI Overviews", "AI Mode"],
                                default=cfg["ai_features"], key="lbm_sel_ai")

    sel_models = dict(cfg.get("models", {}))
    if sel_llms:
        st.markdown("**Modello per LLM**")
        st.caption("Non si applica a AI Overviews / AI Mode.")
        mcols = st.columns(len(sel_llms))
        for col, llm_name in zip(mcols, sel_llms):
            with col:
                opts = AVAILABLE_MODELS.get(llm_name, [])
                cur  = sel_models.get(llm_name, opts[0])
                idx  = opts.index(cur) if cur in opts else 0
                sel_models[llm_name] = st.selectbox(llm_name, opts, index=idx,
                                                    key=f"lbm_m_{llm_name}")

    st.divider()
    st.subheader("🔁 Iterazioni")
    st.caption("AI Overviews e AI Mode hanno sempre 1 iterazione.")
    iterations = st.number_input("Iterazioni per LLM", min_value=1, max_value=20,
                                 value=cfg.get("iterations", 1), step=1)

    n_q = len(st.session_state.lbm_questions)
    est = n_q * (len(sel_llms) * int(iterations) + len(sel_ai))
    if n_q:
        st.info(
            f"{n_q} domande × ({len(sel_llms)} LLM × {int(iterations)} iter + "
            f"{len(sel_ai)} AI Features) = **{est} chiamate**"
        )

    st.divider()
    st.subheader("⚙️ Opzioni avanzate")
    c_a, c_b, c_c, c_d = st.columns(4)
    with c_a:
        language = st.selectbox("Lingua", ["it", "en", "de", "fr", "es", "ja"],
                               key="lbm_lang")
    with c_b:
        country = st.selectbox("Paese", ["it", "us", "gb", "de", "fr", "jp"],
                               key="lbm_country")
    with c_c:
        _bm_opts = _brand_method_options(st.session_state.lbm_keys)
        if st.session_state.get("lbm_bmethod") not in _bm_opts:
            st.session_state["lbm_bmethod"] = "regex"
        brand_method = st.selectbox(
            "Estrazione brand", _bm_opts,
            format_func=lambda x: BRAND_METHOD_LABELS.get(x, x),
            key="lbm_bmethod",
            help="Le opzioni LLM/ensemble compaiono solo se hai configurato la relativa API key.",
        )
    with c_d:
        brand_threshold = st.slider(
            "Soglia fuzzy match", min_value=50, max_value=100,
            key="lbm_threshold",
            help="Soglia RapidFuzz per normalizzare i brand estratti sulla brand list.",
        )

    if st.button("💾 Salva configurazione", type="primary"):
        st.session_state.lbm_config = {
            "llms": sel_llms, "ai_features": sel_ai,
            "iterations": int(iterations), "models": sel_models,
        }
        st.success("Configurazione salvata.")

# ══════════════════════════════════════════════════════════════════════════════
# TAB 3 — ESECUZIONE
# ══════════════════════════════════════════════════════════════════════════════
with tab_run:
    st.subheader("🚀 Esecuzione Run")

    questions   = st.session_state.lbm_questions
    keys_run    = dict(st.session_state.lbm_keys)
    if not keys_run.get("openai"):
        keys_run["openai"] = st.session_state.get("openai_api_key", "")
    config_run  = st.session_state.lbm_config
    known_run   = st.session_state.lbm_known
    bmethod_run = st.session_state.get("lbm_bmethod", "regex")
    lang_run    = st.session_state.get("lbm_lang", "it")
    ctry_run    = st.session_state.get("lbm_country", "it")
    thresh_run  = st.session_state.get("lbm_threshold", 85)

    warns = []
    if not questions:
        warns.append("⚠️ Nessuna domanda caricata. Vai al tab Input.")
    if not (config_run.get("llms") or config_run.get("ai_features")):
        warns.append("⚠️ Nessuna piattaforma selezionata. Vai al tab Configurazione.")
    for w in warns:
        st.warning(w)

    if not warns:
        llms_r = config_run.get("llms", [])
        ai_r   = config_run.get("ai_features", [])
        iters  = config_run.get("iterations", 1)
        mods   = config_run.get("models", {})
        total_c = len(questions) * (len(llms_r) * iters + len(ai_r))

        st.markdown(
            f"**Riepilogo:**  \n"
            f"- Domande: **{len(questions)}** · "
            f"LLM: **{', '.join(llms_r) or '—'}** × {iters} iter · "
            f"AI Features: **{', '.join(ai_r) or '—'}** × 1  \n"
            f"- Totale chiamate: **{total_c}** · "
            f"Brand list: **{len(known_run)}** brand"
        )
        if mods and llms_r:
            st.caption("Modelli: " +
                       "  ·  ".join(f"{k}: `{v}`" for k, v in mods.items() if k in llms_r))

        if st.button("🚀 Avvia Run", type="primary"):
            progress = st.progress(0, text="Avvio…")
            status   = st.status("🚀 Run in corso…", expanded=True)
            table_ph = st.empty()

            log_lines: list[str] = []
            table_rows: list[dict] = []
            counts = {"ok": 0, "inv": 0, "err": 0}

            def _cb(done: int, total: int, r: dict):
                pct = done / max(total, 1)
                progress.progress(pct, text=f"{done}/{total} ({pct:.0%})")
                valid = r["valid"]
                is_err = str(r["text"]).startswith("ERROR")
                icon = "✅" if valid else ("❌" if is_err else "⚠️")
                counts["ok" if valid else ("err" if is_err else "inv")] += 1
                log_lines.append(
                    f"{icon} [{done}/{total}] {r['llm']} ({r['model']}) "
                    f"— {r['question']} — {r['elapsed']}s"
                    + (f" — {len(r['sources'])} fonti" if valid
                       else f" — {str(r['text'])[:80]}")
                )
                status.update(
                    label=f"🚀 {done}/{total} — ✅ {counts['ok']}  ⚠️ {counts['inv']}  ❌ {counts['err']}"
                )
                status.text("\n".join(log_lines[-30:]))

                brands_preview = []
                if valid:
                    brands_preview = [b["name"] for b in
                                      extract_brands(r["text"], keys_run,
                                                     known_run, bmethod_run, thresh_run)]
                table_rows.append({
                    "LLM": r["llm"],
                    "Modello": r["model"],
                    "Domanda": r["question"],
                    "Anteprima": (r["text"][:150].replace("\n", " ") + "…")
                    if valid else str(r["text"])[:100],
                    "Brand estratti": ", ".join(brands_preview[:8]),
                    "Fonti": len(r["sources"]),
                    "Tempo": f"{r['elapsed']}s",
                })
                table_ph.dataframe(
                    pd.DataFrame(table_rows), use_container_width=True, hide_index=True
                )

            try:
                results = run_monitor(
                    questions=questions,
                    keys=keys_run,
                    config=config_run,
                    known_brands=known_run,
                    brand_method=bmethod_run,
                    language=lang_run,
                    country=ctry_run,
                    progress_cb=_cb,
                    brand_threshold=thresh_run,
                )
                st.session_state.lbm_results = results
                progress.progress(1.0, text="✅ Completato!")
                status.update(
                    label=(f"✅ Run completato — ✅ {counts['ok']}  "
                           f"⚠️ {counts['inv']}  ❌ {counts['err']}"),
                    state="complete", expanded=False,
                )
                n_r = len(results["risposte"])
                n_b = len(results["brand"])
                n_f = len(results["fonti"])
                st.success(
                    f"**{n_r}** risposte · **{n_b}** brand · **{n_f}** fonti "
                    f"— vai al tab Export."
                )
            except Exception as exc:
                status.update(label=f"❌ Errore: {exc}", state="error", expanded=False)
                st.error(str(exc))

    if st.session_state.lbm_results:
        st.divider()
        st.caption("👉 Vai al tab **Anteprima & Riclassifica** per esplorare i risultati "
                   "e provare metodologie di estrazione brand diverse.")

# ══════════════════════════════════════════════════════════════════════════════
# TAB 4 — ANTEPRIMA & RICLASSIFICA
# ══════════════════════════════════════════════════════════════════════════════
with tab_preview:
    st.subheader("🔍 Anteprima risultati & Riclassificazione brand")

    if not st.session_state.lbm_results:
        st.info("Nessun run completato. Vai al tab Esecuzione.")
    else:
        results_prev = st.session_state.lbm_results
        n_valid = sum(1 for r in results_prev["risposte"] if r.get("Risposta"))

        st.markdown(
            "Ri-estrai i brand dalle risposte **già raccolte**, con una metodologia diversa "
            "(regex / LLM OpenAI / LLM Anthropic / ensemble con confidenza) o una soglia di "
            "normalizzazione diversa, **senza richiamare di nuovo le API degli LLM che hanno "
            "generato le risposte** — stesso principio della re-estrazione usata nella "
            "dashboard AI Brand Monitoring."
        )

        rc_keys = st.session_state.lbm_keys
        rc_opts = _brand_method_options(rc_keys)
        if st.session_state.get("lbm_rc_method") not in rc_opts:
            st.session_state["lbm_rc_method"] = "regex"

        rc1, rc2, rc3 = st.columns(3)
        with rc1:
            rc_method = st.selectbox(
                "Metodo estrazione", rc_opts,
                format_func=lambda x: BRAND_METHOD_LABELS.get(x, x),
                key="lbm_rc_method",
                help="Ensemble combina più metodi e assegna un punteggio di confidenza "
                     "(quante fonti concordano) a ogni brand.",
            )
        with rc2:
            rc_threshold = st.slider(
                "Soglia fuzzy match", 50, 100,
                value=st.session_state.get("lbm_threshold", 85),
                key="lbm_rc_threshold",
            )
        with rc3:
            st.metric("Risposte riclassificabili", n_valid)

        rc_brands_txt = st.text_area(
            "Brand list per la riclassificazione", height=100,
            value="\n".join(st.session_state.lbm_known),
            key="lbm_rc_brands",
            help="Puoi modificarla qui senza toccare quella salvata nel tab Input.",
        )

        needs_openai = rc_method in ("llm_openai", "ensemble") and not rc_keys.get("openai")
        needs_anthropic = rc_method in ("llm_anthropic", "ensemble") and not rc_keys.get("anthropic")
        can_rc = rc_method == "regex" or rc_keys.get("openai") or rc_keys.get("anthropic")
        if rc_method == "llm_openai" and needs_openai:
            can_rc = False
            st.warning("⚠️ Nessuna API key OpenAI configurata (tab Configurazione).")
        if rc_method == "llm_anthropic" and needs_anthropic:
            can_rc = False
            st.warning("⚠️ Nessuna API key Anthropic configurata (tab Configurazione).")
        if rc_method == "ensemble" and not (rc_keys.get("openai") or rc_keys.get("anthropic")):
            can_rc = False
            st.warning("⚠️ L'ensemble richiede almeno una key OpenAI o Anthropic "
                       "(tab Configurazione) oltre alla regex.")

        if st.button("🔄 Riclassifica brand", type="primary", disabled=not can_rc):
            rc_known = [b.strip() for b in rc_brands_txt.splitlines() if b.strip()]
            progress = st.progress(0, text="Riclassificazione…")

            def _rc_cb(done: int, total: int):
                progress.progress(done / max(total, 1), text=f"{done}/{total} risposte")

            new_brand_rows = reclassify_brands(
                results_prev["risposte"], rc_keys, rc_known,
                rc_method, rc_threshold, progress_cb=_rc_cb,
            )
            st.session_state.lbm_results["brand"] = new_brand_rows
            progress.progress(1.0, text="✅ Completato!")
            st.success(
                f"Riclassificati **{len(new_brand_rows)}** brand con metodo "
                f"**{BRAND_METHOD_LABELS.get(rc_method, rc_method)}** (soglia {rc_threshold}). "
                f"Tabelle aggiornate qui sotto."
            )

        st.divider()

        df_b = pd.DataFrame(results_prev["brand"])
        if not df_b.empty:
            st.markdown("**Classifica brand — menzioni e posizione media**")
            agg = {"Menzioni": ("Brand", "count"), "Posizione_media": ("Position", "mean")}
            if "Confidence" in df_b.columns:
                agg["Confidenza_media"] = ("Confidence", "mean")
            rank = df_b.groupby("Brand").agg(**agg).reset_index()
            rank["Posizione_media"] = rank["Posizione_media"].round(2)
            if "Confidenza_media" in rank.columns:
                rank["Confidenza_media"] = rank["Confidenza_media"].round(2)
            rank = rank.sort_values("Menzioni", ascending=False)
            st.dataframe(rank, use_container_width=True, hide_index=True)

            st.markdown("**Brand più menzionati per LLM analizzato**")
            if "LLM" in df_b.columns:
                pivot = (
                    df_b.groupby(["Brand", "LLM"]).size()
                    .unstack(fill_value=0)
                    .assign(Totale=lambda d: d.sum(axis=1))
                    .sort_values("Totale", ascending=False)
                    .reset_index()
                )
                st.dataframe(pivot, use_container_width=True, hide_index=True)
            else:
                st.caption("Nessuna colonna LLM disponibile nei dati brand.")

        st.divider()
        t_r, t_b, t_f = st.tabs([
            f"Risposte ({len(results_prev['risposte'])})",
            f"Brand ({len(results_prev['brand'])})",
            f"Fonti ({len(results_prev['fonti'])})",
        ])
        with t_r:
            df = pd.DataFrame(results_prev["risposte"])
            if not df.empty:
                df = df.copy()
                df["Risposta"] = df["Risposta"].str[:150] + "…"
            st.dataframe(df, use_container_width=True, hide_index=True)
        with t_b:
            st.dataframe(pd.DataFrame(results_prev["brand"]),
                         use_container_width=True, hide_index=True)
        with t_f:
            st.dataframe(pd.DataFrame(results_prev["fonti"]),
                         use_container_width=True, hide_index=True)

# ══════════════════════════════════════════════════════════════════════════════
# TAB 5 — EXPORT
# ══════════════════════════════════════════════════════════════════════════════
with tab_export:
    st.subheader("📥 Export Excel")

    if not st.session_state.lbm_results:
        st.info("Nessun run completato. Vai al tab Esecuzione.")
    else:
        results_exp = st.session_state.lbm_results
        c1, c2 = st.columns(2)
        with c1:
            project_name = st.text_input("Nome progetto", value="LLM_Brand_Monitor")
        with c2:
            run_date = st.date_input("Data run", value=date.today())

        c_r, c_b, c_f = st.columns(3)
        c_r.metric("Risposte", len(results_exp["risposte"]))
        c_b.metric("Brand (righe)", len(results_exp["brand"]))
        c_f.metric("Fonti (righe)", len(results_exp["fonti"]))

        if st.button("📊 Genera e scarica Excel", type="primary"):
            with st.spinner("Generazione Excel…"):
                xlsx = build_excel(
                    results_exp, st.session_state.lbm_questions, run_date
                )
            fname = f"{run_date.isoformat()}_{project_name}_AI_Brand_Monitor.xlsx"
            st.download_button(
                "⬇️ Scarica Excel",
                data=xlsx,
                file_name=fname,
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
            st.success(f"`{fname}`")
            st.caption(
                "4 fogli: **Risposte - Apps Script** · "
                "**Brand - Apps Script** · **Fonti - Apps Script** · **AI Questions**"
            )
