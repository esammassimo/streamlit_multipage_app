"""
NVL Agency · Close Reading & Signal Cleaning
=============================================
Analisi semantica LLM di testi grezzi per:
1. Estrarre vocabolario reale del target e mappare l'intento
2. Produrre lista di negative keyword per pulire il segnale su Google Ads

Input:
  - File (txt, pdf, csv) con testi grezzi
  - Lista URL (incolla, txt, csv, xlsx) → scraping requests o Playwright

Output:
  - signal_map.xlsx   : mappa semantica intento → pattern → parole
  - negative_kw.csv   : lista query informazionali da uploadare su Google Ads
"""

import io
import json
import re
import time
import traceback
from pathlib import Path
from urllib.parse import urlparse

import anthropic
import numpy as np
import pandas as pd
import streamlit as st
from openpyxl import Workbook
try:
    from rapidfuzz import fuzz as _rfuzz
    _HAS_RAPIDFUZZ = True
except ImportError:
    _HAS_RAPIDFUZZ = False
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

# ─────────────────────────────────────────────────────────────
# PAGE CONFIG (commentato — va solo in Home.py)
# ─────────────────────────────────────────────────────────────
# st.set_page_config(
#     page_title="Close Reading · NVL Agency",
#     page_icon="🔍",
#     layout="wide",
#     initial_sidebar_state="expanded",
# )


# ─────────────────────────────────────────────────────────────
# SIDEBAR
# ─────────────────────────────────────────────────────────────
with st.sidebar:
    st.markdown("## 🔍 Close Reading")
    st.divider()

    st.markdown("**Cliente / Brand**")
    brand_name = st.text_input(
        "Nome brand",
        value=st.session_state.get("cr_brand", ""),
        label_visibility="collapsed",
        key="cr_brand_input",
        placeholder="es. Garnier",
    )
    st.session_state["cr_brand"] = brand_name

    st.markdown("**Settore / Categoria prodotto**")
    sector = st.text_input(
        "Settore",
        value=st.session_state.get("cr_sector", ""),
        label_visibility="collapsed",
        key="cr_sector_input",
        placeholder="es. skincare, haircare...",
    )
    st.session_state["cr_sector"] = sector

    st.markdown("**Lingua dei testi**")
    lang = st.selectbox(
        "Lingua",
        ["Italiano", "Inglese", "Francese", "Spagnolo", "Tedesco"],
        index=0,
        key="cr_lang",
        label_visibility="collapsed",
    )

    st.divider()

    st.markdown("**🔑 Anthropic API Key**")
    _ant_key = st.text_input(
        "Anthropic API Key",
        type="password",
        placeholder="sk-ant-...",
        value=st.session_state.get("anthropic_api_key", ""),
        label_visibility="collapsed",
        help="Usata da Claude per l'analisi semantica.",
    )
    if _ant_key:
        st.session_state["anthropic_api_key"] = _ant_key

    if st.session_state.get("anthropic_api_key"):
        st.success("API Key configurata")
    else:
        st.warning("Inserisci la Anthropic API Key per procedere.")

    st.divider()
    st.markdown("**Scraping**")
    scraper_mode = st.radio(
        "Modalità scraping URL",
        ["requests (veloce)", "Playwright (headless, per siti protetti)"],
        key="cr_scraper",
        label_visibility="collapsed",
        help="Usa requests per siti normali. Playwright per Trustpilot, Reddit, siti con JS."
    )
    use_playwright = "Playwright" in scraper_mode

    st.divider()
    st.caption("File accettati: txt, pdf, csv, xlsx")


# ─────────────────────────────────────────────────────────────
# SCRAPING FUNCTIONS
# ─────────────────────────────────────────────────────────────

def scrape_with_requests(url: str, timeout: int = 15) -> str:
    """Scraping semplice con requests + BeautifulSoup."""
    try:
        import requests
        from bs4 import BeautifulSoup
    except ImportError:
        return "ERRORE: librerie requests/beautifulsoup4 non installate."

    headers = {
        "User-Agent": (
            "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
            "AppleWebKit/537.36 (KHTML, like Gecko) "
            "Chrome/120.0.0.0 Safari/537.36"
        )
    }
    try:
        resp = requests.get(url, headers=headers, timeout=timeout)
        resp.raise_for_status()
        soup = BeautifulSoup(resp.text, "html.parser")
        # Rimuovi script, style, nav, footer
        for tag in soup(["script", "style", "nav", "footer", "header", "aside"]):
            tag.decompose()
        text = soup.get_text(separator="\n", strip=True)
        # Pulizia righe vuote
        lines = [l.strip() for l in text.splitlines() if len(l.strip()) > 30]
        return "\n".join(lines)
    except Exception as e:
        return f"ERRORE scraping {url}: {e}"


def ensure_playwright_browser_installed() -> tuple[bool, str]:
    """
    Verifica se Chromium è installato per Playwright; se manca, lo installa.
    Risultato cachato in session_state per non ripetere il check ogni volta.
    Ritorna (success, message).
    """
    import subprocess, sys

    if st.session_state.get("_playwright_browser_ready"):
        return True, "già verificato"

    try:
        # Verifica rapida: prova a lanciare il browser
        check = subprocess.run(
            [sys.executable, "-c",
             "from playwright.sync_api import sync_playwright\n"
             "with sync_playwright() as p:\n"
             "    b = p.chromium.launch(headless=True)\n"
             "    b.close()\n"
             "print('OK')"],
            capture_output=True, text=True, timeout=30
        )
        if check.returncode == 0 and "OK" in check.stdout:
            st.session_state["_playwright_browser_ready"] = True
            return True, "browser già presente"
    except Exception:
        pass

    # Browser non presente — tenta installazione
    try:
        install = subprocess.run(
            [sys.executable, "-m", "playwright", "install", "chromium", "--with-deps"],
            capture_output=True, text=True, timeout=180
        )
        if install.returncode == 0:
            st.session_state["_playwright_browser_ready"] = True
            return True, "installato ora"
        else:
            err = (install.stderr or install.stdout).strip()[-500:]
            return False, f"installazione fallita: {err}"
    except subprocess.TimeoutExpired:
        return False, "installazione: timeout (>180s) — riprova o usa requests"
    except Exception as e:
        return False, f"installazione: errore {e}"


def scrape_with_playwright(url: str, timeout: int = 20000) -> str:
    """
    Scraping headless con Playwright tramite subprocess isolato.
    Usa subprocess per evitare conflitti con il thread Streamlit.
    Verifica/installa il browser Chromium al primo utilizzo.
    """
    import subprocess, sys, tempfile, os, textwrap

    ready, msg = ensure_playwright_browser_installed()
    if not ready:
        return (f"ERRORE Playwright {url}: browser Chromium non disponibile ({msg}). "
                f"Passa a 'requests' nella sidebar oppure riprova.")

    # Script Playwright scritto come stringa senza triple-quote annidate
    js_cleanup = "['script','style','nav','footer','header','aside','iframe'].forEach(t=>{document.querySelectorAll(t).forEach(el=>el.remove())});"

    script_lines = [
        "import sys",
        "from playwright.sync_api import sync_playwright",
        f"url = {repr(url)}",
        f"timeout = {timeout}",
        "try:",
        "    with sync_playwright() as p:",
        "        browser = p.chromium.launch(headless=True, args=['--no-sandbox','--disable-setuid-sandbox','--disable-dev-shm-usage','--disable-gpu'])",
        "        ctx = browser.new_context(user_agent='Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36')",
        "        page = ctx.new_page()",
        "        try:",
        "            page.goto(url, timeout=timeout, wait_until='domcontentloaded')",
        "        except Exception:",
        "            pass",
        "        page.wait_for_timeout(2000)",
        f"        page.evaluate({repr(js_cleanup)})",
        "        text = page.inner_text('body')",
        "        browser.close()",
        "    lines = [l.strip() for l in text.splitlines() if len(l.strip()) > 30]",
        "    print('\\n'.join(lines))",
        "except Exception as e:",
        "    print(f'ERRORE Playwright: {e}')",
        "    sys.exit(1)",
    ]
    script = "\n".join(script_lines)

    try:
        with tempfile.NamedTemporaryFile(mode="w", suffix=".py",
                                         delete=False, encoding="utf-8") as tmp:
            tmp.write(script)
            tmp_path = tmp.name

        result = subprocess.run(
            [sys.executable, tmp_path],
            capture_output=True, text=True,
            timeout=timeout // 1000 + 15
        )
        os.unlink(tmp_path)

        output = result.stdout.strip()
        if result.returncode != 0 or output.startswith("ERRORE"):
            err = result.stderr.strip() or output
            return f"ERRORE Playwright {url}: {err[:400]}"
        if not output:
            return f"ERRORE Playwright {url}: nessun contenuto estratto"
        return output

    except subprocess.TimeoutExpired:
        return f"ERRORE Playwright {url}: timeout ({timeout//1000}s)"
    except Exception as e:
        return f"ERRORE Playwright {url}: {e}"

def scrape_url(url: str, use_playwright: bool = False) -> str:
    if use_playwright:
        result = scrape_with_playwright(url)
        if result.startswith("ERRORE"):
            # Fallback automatico a requests se Playwright non disponibile
            fallback = scrape_with_requests(url)
            if not fallback.startswith("ERRORE"):
                return fallback
            return f"{result}\n[Fallback requests fallito anche]: {fallback}"
        return result
    return scrape_with_requests(url)


# ─────────────────────────────────────────────────────────────
# FILE PARSING
# ─────────────────────────────────────────────────────────────

def extract_text_from_file(uploaded_file) -> str:
    """Estrae testo grezzo da txt, csv, pdf."""
    name = uploaded_file.name
    suffix = Path(name).suffix.lower()
    buf = io.BytesIO(uploaded_file.read())

    if suffix == ".txt":
        return buf.read().decode("utf-8", errors="replace")

    elif suffix == ".csv":
        buf.seek(0)
        df = pd.read_csv(buf, dtype=str, encoding="utf-8-sig", on_bad_lines="skip")
        return "\n".join(df.astype(str).apply(lambda r: " | ".join(r.dropna()), axis=1).tolist())

    elif suffix in (".xlsx", ".xls"):
        buf.seek(0)
        df = pd.read_excel(buf, dtype=str)
        return "\n".join(df.astype(str).apply(lambda r: " | ".join(r.dropna()), axis=1).tolist())

    elif suffix == ".pdf":
        try:
            import pdfplumber
            buf.seek(0)
            with pdfplumber.open(buf) as pdf:
                return "\n".join(
                    page.extract_text() or "" for page in pdf.pages
                )
        except ImportError:
            return "ERRORE: pdfplumber non installato. Esegui: pip install pdfplumber"
        except Exception as e:
            return f"ERRORE lettura PDF: {e}"

    return f"Formato non supportato: {suffix}"


def parse_url_list(uploaded_file) -> list[str]:
    """Estrae lista URL da file txt/csv/xlsx."""
    name = uploaded_file.name
    suffix = Path(name).suffix.lower()
    buf = io.BytesIO(uploaded_file.read())

    if suffix == ".txt":
        text = buf.read().decode("utf-8", errors="replace")
        lines = [l.strip() for l in text.splitlines() if l.strip().startswith("http")]
        return lines

    elif suffix == ".csv":
        buf.seek(0)
        df = pd.read_csv(buf, dtype=str, header=None, on_bad_lines="skip")
        urls = []
        for col in df.columns:
            urls += [v.strip() for v in df[col].dropna() if str(v).strip().startswith("http")]
        return list(dict.fromkeys(urls))

    elif suffix in (".xlsx", ".xls"):
        buf.seek(0)
        df = pd.read_excel(buf, dtype=str, header=None)
        urls = []
        for col in df.columns:
            urls += [v.strip() for v in df[col].dropna() if str(v).strip().startswith("http")]
        return list(dict.fromkeys(urls))

    return []


# ─────────────────────────────────────────────────────────────
# LLM ANALYSIS
# ─────────────────────────────────────────────────────────────

SYSTEM_PROMPT = """Sei un esperto di analisi semantica e strategia SEO/SEM.
Il tuo compito è analizzare testi grezzi (recensioni, forum, articoli) per estrarre
segnali di intento utili alla pianificazione di campagne Google Ads.
Rispondi SOLO in JSON valido, senza backtick, senza testo fuori dal JSON."""

def build_analysis_prompt(texts: list[dict], brand: str, sector: str, lang: str) -> str:
    # Tronca ogni testo a 3000 caratteri per gestire il context window
    corpus = ""
    for t in texts:
        snippet = t["text"][:3000].replace('"', "'")
        corpus += f"\n\n[FONTE: {t['label']} — {t['source']}]\n{snippet}"

    return f"""Analizza i seguenti testi grezzi relativi al brand "{brand}" nel settore "{sector}".
I testi sono in lingua: {lang}.

TESTI:
{corpus}

Produci un JSON con questa struttura esatta:

{{
  "vocabolario_target": [
    {{"frase": "...", "frequenza_stimata": "alta|media|bassa", "contesto": "..."}}
  ],
  "intenti_commerciali": [
    {{"pattern": "...", "esempio_query": "...", "segnale": "...", "priorita": "alta|media|bassa"}}
  ],
  "frizioni_dubbi": [
    {{"tema": "...", "frase_esatta": "...", "competitor_coinvolti": "...", "opportunita_ads": "..."}}
  ],
  "intenti_informazionali": [
    {{"pattern": "...", "esempio_query": "...", "motivazione_esclusione": "..."}}
  ],
  "negative_keywords": [
    {{"keyword": "...", "match_type": "broad|phrase|exact", "motivazione": "..."}}
  ],
  "insight_ai_brief": [
    {{"istruzione": "...", "razionale": "..."}}
  ]
}}

Regole:
- vocabolario_target: parole/frasi ESATTE dai testi, non parafrasate
- intenti_commerciali: pattern semantici con reale intenzione d acquisto/valutazione
- frizioni_dubbi: obiezioni, confronti con competitor, punti di attrito
- intenti_informazionali: pattern da escludere dalle campagne (guide, cos è, come funziona, tutorial)
- negative_keywords: lista pronta per Google Ads, focus su query puramente informazionali
- insight_ai_brief: istruzioni in linguaggio naturale per Google AI Brief (max 5)
"""


def run_claude_analysis(texts: list[dict], brand: str, sector: str,
                        lang: str, api_key: str) -> dict:
    client = anthropic.Anthropic(api_key=api_key)

    # Se i testi sono molti, analizza in batch da 5 e aggrega
    BATCH_SIZE = 5
    batches = [texts[i:i+BATCH_SIZE] for i in range(0, len(texts), BATCH_SIZE)]
    all_results = []

    for batch in batches:
        prompt = build_analysis_prompt(batch, brand, sector, lang)
        msg = client.messages.create(
            model="claude-opus-4-5",
            max_tokens=4000,
            system=SYSTEM_PROMPT,
            messages=[{"role": "user", "content": prompt}]
        )
        raw = msg.content[0].text.strip()
        # Pulizia backtick residui
        raw = re.sub(r"^```json\s*", "", raw)
        raw = re.sub(r"\s*```$", "", raw)
        try:
            all_results.append(json.loads(raw))
        except Exception:
            # Fallback: prova a estrarre il JSON dal testo
            match = re.search(r"\{.*\}", raw, re.DOTALL)
            if match:
                all_results.append(json.loads(match.group()))

    if not all_results:
        raise ValueError("Nessun risultato valido dall'analisi LLM.")

    # Aggrega i risultati dei batch
    if len(all_results) == 1:
        return all_results[0]

    merged = {k: [] for k in all_results[0].keys()}
    for r in all_results:
        for k in merged:
            if k in r and isinstance(r[k], list):
                merged[k].extend(r[k])
    return merged


# ─────────────────────────────────────────────────────────────
# EXCEL EXPORT
# ─────────────────────────────────────────────────────────────

def thin_border():
    s = Side(style="thin", color="BFBFBF")
    return Border(left=s, right=s, top=s, bottom=s)


def write_sheet(wb: Workbook, title: str, rows: list[dict],
                header_color: str = "1F3864") -> None:
    if not rows:
        return
    ws = wb.create_sheet(title=title[:31])
    headers = list(rows[0].keys())
    n = len(headers)

    # Titolo
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=n)
    c = ws.cell(row=1, column=1, value=title)
    c.fill = PatternFill("solid", fgColor=header_color)
    c.font = Font(name="Arial", bold=True, color="FFFFFF", size=11)
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 22

    # Header
    for j, h in enumerate(headers, 1):
        cell = ws.cell(row=2, column=j, value=h.replace("_", " ").title())
        cell.fill = PatternFill("solid", fgColor=header_color)
        cell.font = Font(name="Arial", bold=True, color="FFFFFF", size=9)
        cell.alignment = Alignment(horizontal="center", vertical="center",
                                   wrap_text=True)
        cell.border = thin_border()
    ws.row_dimensions[2].height = 28

    # Dati
    for i, row in enumerate(rows):
        r = i + 3
        for j, h in enumerate(headers, 1):
            val = row.get(h, "")
            cell = ws.cell(row=r, column=j, value=val)
            cell.font = Font(name="Arial", size=9)
            cell.alignment = Alignment(vertical="center", wrap_text=True)
            cell.border = thin_border()
            # Zebratura
            if i % 2 == 1:
                cell.fill = PatternFill("solid", fgColor="F5F5F5")
        ws.row_dimensions[r].height = 14

    # Larghezze colonne
    for col in ws.columns:
        w = max((len(str(c.value)) if c.value else 0) for c in col)
        ws.column_dimensions[get_column_letter(col[0].column)].width = min(
            max(w + 2, 12), 60
        )
    ws.freeze_panes = "A3"


SHEET_CONFIG = {
    "vocabolario_target": ("Vocabolario Target", "2E5FAB"),
    "intenti_commerciali": ("Intenti Commerciali", "375623"),
    "frizioni_dubbi": ("Frizioni e Dubbi", "833C00"),
    "intenti_informazionali": ("Intenti Informazionali", "7030A0"),
    "negative_keywords": ("Negative Keywords", "C00000"),
    "insight_ai_brief": ("AI Brief Insights", "0563C1"),
}


def build_signal_map_excel(result: dict, brand: str) -> bytes:
    wb = Workbook()
    wb.remove(wb.active)

    for key, (sheet_title, color) in SHEET_CONFIG.items():
        rows = result.get(key, [])
        if rows:
            write_sheet(wb, sheet_title, rows, header_color=color)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def build_negative_kw_csv(result: dict) -> bytes:
    rows = result.get("negative_keywords", [])
    if not rows:
        return b""
    df = pd.DataFrame(rows)
    # Formato compatibile Google Ads bulk upload
    if "keyword" in df.columns and "match_type" in df.columns:
        def _fmt(row):
            kw = str(row["keyword"]).strip()
            mt = str(row.get("match_type", "broad")).strip().lower()
            if mt == "exact":
                return f"[{kw}]"
            elif mt == "phrase":
                return f'"{kw}"'
            return kw
        df["keyword_formatted"] = df.apply(_fmt, axis=1)
    return df.to_csv(index=False, encoding="utf-8-sig").encode("utf-8-sig")


def build_classification_csv(result: dict) -> bytes:
    """
    Esporta i pattern degli intenti commerciali come file di classificazione
    pronto per l'upload in 18_seo_sea.py.

    Formato output:
      keyword  |  Intento  |  Priorità
    ─────────────────────────────────────
    (una riga per pattern, con eventuale esempio_query come keyword alternativa)

    Il file viene riconosciuto automaticamente da SEO/SEA:
    - colonna "keyword"  → chiave di join (fuzzy)
    - colonna "Intento"  → etichetta (riconosciuta come tipo "intento")
    - colonna "Priorità" → metadato aggiuntivo
    """
    rows_out = []

    # Da intenti_commerciali: usa esempio_query come keyword se disponibile
    for item in result.get("intenti_commerciali", []):
        pattern   = str(item.get("pattern", "")).strip()
        esempio   = str(item.get("esempio_query", "")).strip()
        priorita  = str(item.get("priorita", "media")).strip()
        segnale   = str(item.get("segnale", "")).strip()

        # Riga principale: il pattern
        if pattern:
            rows_out.append({
                "keyword":  pattern,
                "Intento":  segnale if segnale else pattern,
                "Priorità": priorita,
                "Tipo":     "pattern",
            })
        # Riga aggiuntiva: l'esempio di query (più facile da matchare con le keyword ADS)
        if esempio and esempio != pattern:
            rows_out.append({
                "keyword":  esempio,
                "Intento":  segnale if segnale else pattern,
                "Priorità": priorita,
                "Tipo":     "query_esempio",
            })

    # Da vocabolario_target: aggiungi le frasi ad alta frequenza
    for item in result.get("vocabolario_target", []):
        frase = str(item.get("frase", "")).strip()
        freq  = str(item.get("frequenza_stimata", "")).strip().lower()
        if frase and freq in ("alta", "media"):
            rows_out.append({
                "keyword":  frase,
                "Intento":  "vocabolario_target",
                "Priorità": "alta" if freq == "alta" else "media",
                "Tipo":     "vocabolario",
            })

    if not rows_out:
        return b""

    df = pd.DataFrame(rows_out).drop_duplicates(subset=["keyword"])
    return df.to_csv(index=False, encoding="utf-8-sig").encode("utf-8-sig")


# ─────────────────────────────────────────────────────────────
# MAIN UI
# ─────────────────────────────────────────────────────────────

st.markdown("# 🔍 Close Reading & Signal Cleaning")
st.caption(
    "Analisi semantica LLM di testi grezzi per mappare l'intento reale del target "
    "e produrre negative keyword per pulire il segnale su Google Ads."
)
st.divider()

# ── SEZIONE 1: INPUT TESTI ────────────────────────────────────────────────
st.markdown("### 📥 Input Testi")
st.caption(
    "Carica file con testi grezzi (recensioni, forum, articoli) "
    "oppure fornisci una lista di URL da scrappare — o entrambi."
)

input_tab1, input_tab2 = st.tabs(["📄 Carica File", "🌐 Lista URL"])

sources: list[dict] = []  # {"source": str, "label": str, "text": str}

# ── TAB 1: Upload file ────────────────────────────────────────────────────
with input_tab1:
    st.markdown("**Carica uno o più file** (txt, pdf, csv, xlsx)")
    uploaded_files = st.file_uploader(
        "File testi grezzi",
        type=["txt", "pdf", "csv", "xlsx", "xls"],
        accept_multiple_files=True,
        key="cr_files",
        label_visibility="collapsed",
    )

    if uploaded_files:
        st.markdown("**Etichetta ogni file:**")
        LABEL_OPTIONS = ["brand", "competitor", "forum", "review", "articolo", "altro"]
        file_labels = {}
        cols = st.columns(min(len(uploaded_files), 3))
        for i, f in enumerate(uploaded_files):
            with cols[i % 3]:
                file_labels[f.name] = st.selectbox(
                    f.name[:30],
                    LABEL_OPTIONS,
                    key=f"label_file_{i}",
                    index=3 if "review" in f.name.lower() else
                          1 if "competitor" in f.name.lower() else
                          2 if "forum" in f.name.lower() else 0,
                )

        for f in uploaded_files:
            text = extract_text_from_file(f)
            sources.append({
                "source": f.name,
                "label": file_labels.get(f.name, "altro"),
                "text": text,
            })
        st.success(f"✓ {len(uploaded_files)} file caricati — "
                   f"{sum(len(s['text']) for s in sources):,} caratteri totali")

# ── TAB 2: Lista URL ──────────────────────────────────────────────────────
with input_tab2:
    col_url1, col_url2 = st.columns([3, 2])

    with col_url1:
        st.markdown("**Incolla URL (uno per riga)**")
        url_text = st.text_area(
            "URL",
            height=180,
            key="cr_url_text",
            label_visibility="collapsed",
            placeholder="https://www.trustpilot.com/review/...\nhttps://www.reddit.com/r/...",
        )

    with col_url2:
        st.markdown("**Oppure carica file URL** (txt, csv, xlsx)")
        url_file = st.file_uploader(
            "File URL",
            type=["txt", "csv", "xlsx", "xls"],
            key="cr_url_file",
            label_visibility="collapsed",
        )

    # Raccogli URL da entrambe le fonti
    url_list: list[str] = []
    if url_text:
        url_list += [u.strip() for u in url_text.splitlines()
                     if u.strip().startswith("http")]
    if url_file:
        url_list += parse_url_list(url_file)
    url_list = list(dict.fromkeys(url_list))  # deduplication

    if url_list:
        st.markdown(f"**{len(url_list)} URL rilevati** — etichetta per categoria:")
        LABEL_OPTIONS = ["brand", "competitor", "forum", "review", "articolo", "altro"]
        url_labels = {}
        for i, url in enumerate(url_list):
            domain = urlparse(url).netloc.replace("www.", "")
            default = (
                "review"     if any(x in domain for x in ["trustpilot", "reviews", "tripadvisor"]) else
                "forum"      if any(x in domain for x in ["reddit", "forum", "quora"]) else
                "competitor" if brand_name and brand_name.lower() not in domain else
                "articolo"
            )
            idx = LABEL_OPTIONS.index(default) if default in LABEL_OPTIONS else 0
            url_labels[url] = st.selectbox(
                domain[:50],
                LABEL_OPTIONS,
                index=idx,
                key=f"label_url_{i}",
            )

        st.markdown(f"**Modalità:** `{'Playwright' if use_playwright else 'requests'}`")

        if st.button("🕷 Avvia scraping", key="btn_scrape"):
            scraped_sources = []
            prog = st.progress(0)
            status = st.empty()
            errors = []
            for i, url in enumerate(url_list):
                status.text(f"Scraping {i+1}/{len(url_list)}: {url[:60]}...")
                text = scrape_url(url, use_playwright=use_playwright)
                if text.startswith("ERRORE"):
                    errors.append(f"{url}: {text}")
                else:
                    scraped_sources.append({
                        "source": url,
                        "label": url_labels.get(url, "altro"),
                        "text": text,
                    })
                prog.progress((i + 1) / len(url_list))

            st.session_state["cr_scraped_sources"] = scraped_sources
            prog.empty(); status.empty()

            if errors:
                with st.expander(f"⚠️ {len(errors)} errori di scraping"):
                    for e in errors:
                        st.caption(e)
            if scraped_sources:
                st.success(f"✓ Scraping completato — {len(scraped_sources)} URL, "
                           f"{sum(len(s['text']) for s in scraped_sources):,} caratteri")

    # Aggiungi sorgenti scrappate al pool
    scraped = st.session_state.get("cr_scraped_sources", [])
    sources = sources + scraped

# ── RIEPILOGO SORGENTI ────────────────────────────────────────────────────
st.divider()
if sources:
    st.markdown(f"**Corpus totale: {len(sources)} fonti · "
                f"{sum(len(s['text']) for s in sources):,} caratteri**")
    with st.expander("Anteprima corpus", expanded=False):
        for s in sources:
            st.markdown(f"**[{s['label'].upper()}]** `{s['source'][:60]}`")
            st.caption(s["text"][:300] + "…" if len(s["text"]) > 300 else s["text"])
            st.divider()
else:
    st.info("Nessuna fonte caricata. Usa i tab sopra per aggiungere testi o URL.")

# ── SEZIONE 2: ANALISI LLM ────────────────────────────────────────────────
st.markdown("### 🧠 Analisi Semantica")

_ready = (
    bool(sources) and
    bool(brand_name) and
    bool(st.session_state.get("anthropic_api_key"))
)

if not _ready:
    missing = []
    if not sources:
        missing.append("testi o URL")
    if not brand_name:
        missing.append("nome brand (sidebar)")
    if not st.session_state.get("anthropic_api_key"):
        missing.append("Anthropic API Key (sidebar)")
    st.warning(f"Per procedere mancano: {', '.join(missing)}.")
else:
    col_run1, col_run2 = st.columns([2, 3])
    with col_run1:
        run_analysis = st.button(
            "▶ Avvia Close Reading",
            type="primary",
            key="btn_analysis",
        )
    with col_run2:
        st.caption(
            f"Claude analizzerà **{len(sources)} fonti** "
            f"({sum(len(s['text']) for s in sources):,} caratteri) "
            f"per brand **{brand_name}** · settore **{sector or 'non specificato'}** · "
            f"lingua **{lang}**"
        )

    if run_analysis:
        api_key = st.session_state["anthropic_api_key"]
        with st.spinner("Analisi in corso con Claude…"):
            try:
                result = run_claude_analysis(
                    sources, brand_name, sector, lang, api_key
                )
                st.session_state["cr_result"] = result
            except Exception as e:
                st.error(f"❌ Errore analisi: {e}")
                with st.expander("Dettagli"):
                    st.code(traceback.format_exc())

# ── SEZIONE 3: RISULTATI ─────────────────────────────────────────────────
result = st.session_state.get("cr_result")

if result:
    st.divider()
    st.markdown("### 📊 Risultati")

    # KPI rapidi
    k1, k2, k3, k4 = st.columns(4)
    k1.metric("Pattern commerciali", len(result.get("intenti_commerciali", [])))
    k2.metric("Frizioni rilevate",   len(result.get("frizioni_dubbi", [])))
    k3.metric("Negative keyword",    len(result.get("negative_keywords", [])))
    k4.metric("Insight AI Brief",    len(result.get("insight_ai_brief", [])))

    # Tab risultati
    rtabs = st.tabs([
        "💬 Vocabolario",
        "🎯 Intenti commerciali",
        "⚡ Frizioni",
        "🚫 Intenti informazionali",
        "❌ Negative Keywords",
        "🤖 AI Brief",
    ])

    def _show_table(tab, key, cols_order=None):
        with tab:
            rows = result.get(key, [])
            if not rows:
                st.caption("Nessun dato estratto per questa sezione.")
                return
            df = pd.DataFrame(rows)
            if cols_order:
                df = df[[c for c in cols_order if c in df.columns] +
                         [c for c in df.columns if c not in cols_order]]
            st.dataframe(df, use_container_width=True, hide_index=True, height=380)

    _show_table(rtabs[0], "vocabolario_target",
                ["frase", "frequenza_stimata", "contesto"])
    _show_table(rtabs[1], "intenti_commerciali",
                ["pattern", "esempio_query", "segnale", "priorita"])
    _show_table(rtabs[2], "frizioni_dubbi",
                ["tema", "frase_esatta", "competitor_coinvolti", "opportunita_ads"])
    _show_table(rtabs[3], "intenti_informazionali",
                ["pattern", "esempio_query", "motivazione_esclusione"])
    _show_table(rtabs[4], "negative_keywords",
                ["keyword", "match_type", "motivazione"])
    _show_table(rtabs[5], "insight_ai_brief",
                ["istruzione", "razionale"])

    # ── SEZIONE 4: EXPORT ────────────────────────────────────────────────
    st.divider()
    st.markdown("### 💾 Export")

    col_dl1, col_dl2, col_dl3 = st.columns(3)

    with col_dl1:
        signal_map = build_signal_map_excel(result, brand_name)
        fname_map = f"signal_map_{brand_name.lower().replace(' ', '_')}.xlsx"
        st.download_button(
            "⬇ Signal Map Excel",
            data=signal_map,
            file_name=fname_map,
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            help="Mappa semantica completa: vocabolario, intenti, frizioni, AI Brief."
        )

    with col_dl2:
        neg_csv = build_negative_kw_csv(result)
        fname_neg = f"negative_keywords_{brand_name.lower().replace(' ','_')}.csv"
        st.download_button(
            "⬇ Negative Keywords CSV",
            data=neg_csv,
            file_name=fname_neg,
            mime="text/csv",
            help="Lista pronta per upload su Google Ads (formato bulk).",
            disabled=not neg_csv,
        )

    with col_dl3:
        st.caption(
            "**Signal Map** → 6 sheet: Vocabolario, Intenti commerciali, "
            "Frizioni, Intenti informazionali, Negative Keywords, AI Brief\n\n"
            "**Negative Keywords CSV** → formato Google Ads bulk upload "
            "(keyword con `[exact]`, `\"phrase\"`, broad)"
        )

    # ── Export per SEO/SEA ────────────────────────────────────────────────
    st.divider()
    st.markdown("### 🔗 Workflow → SEO/SEA Analysis")
    st.markdown(
        "Esporta il file di classificazione e caricalo direttamente nella pagina "
        "**SEO/SEA Analysis** come *Classificazione Keyword*. "
        "Il match è fuzzy: i pattern vengono confrontati con le keyword ADS "
        "anche senza corrispondenza esatta."
    )

    classif_csv = build_classification_csv(result)
    fname_classif = f"classif_intenti_{brand_name.lower().replace(' ', '_')}.csv"

    col_w1, col_w2 = st.columns([2, 3])
    with col_w1:
        st.download_button(
            "⬇ Scarica file classificazione → SEO/SEA",
            data=classif_csv,
            file_name=fname_classif,
            mime="text/csv",
            disabled=not classif_csv,
            type="primary",
            help="Carica questo file in SEO/SEA Analysis → sezione 'Classificazione Keyword'."
        )
    with col_w2:
        n_patterns = len(result.get("intenti_commerciali", []))
        n_vocab    = sum(1 for v in result.get("vocabolario_target", [])
                        if v.get("frequenza_stimata", "").lower() in ("alta", "media"))
        st.markdown(f"""
**Contenuto del file:**
- `{n_patterns * 2}` righe da intenti commerciali (pattern + query esempio)
- `{n_vocab}` frasi da vocabolario (frequenza alta/media)
- Colonne: `keyword` · `Intento` · `Priorità` · `Tipo`

**Come usarlo in SEO/SEA:**
1. Scarica il file ↑
2. Vai su **SEO/SEA Analysis** → *Carica i file* → **Classificazione Keyword**
3. Il match fuzzy associa automaticamente le keyword ADS agli intenti
4. Nell'Excel di output trovi la colonna **Intento** in ogni sheet KW
        """)
